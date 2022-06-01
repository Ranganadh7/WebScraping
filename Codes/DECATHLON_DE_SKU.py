import re
import openpyxl
import requests
import lxml
from lxml import html
import pandas as pd
import time
import general
from datetime import datetime
import json
import random
from random import choice
import os
from bs4 import BeautifulSoup
from csv import DictWriter



headers = {'authority': 'www.decathlon.de',
           'method': 'GET',
           'path': '/p/dartscheibe-ed-310-elektronisch-mit-6-pfeilen/_/R-p-12794?mc=8295139 ',
           'cheme': 'https',
           'accept': '*/*',
           'accept-encoding': 'gzip, deflate,br',
           'accept-language': 'en-US,en;q=0.9',
           'cookie': 'ACTIVE_USER =y; nfs-traffic-split=2109_atcpopin%3Datc2%2C2109_newpaycta%3Dv1%2Ccart'
                     '-notification%3Dn2%2C2110_divided-navigation%3Dd2; PLAY_LANG=de; '
                     'visid_incap_989911=MQ5i7SZfQu2+3/vHOd+YAwJHyWEAAAAAQUIPAAAAAAA/ndKNCKPAVzZugnW/Owc3; '
                     'nlbi_989911=0+2oSARuuGkBCSQnrLFxRQAAAADSGIWE3E6D/Ro86NvXCeSU; '
                     'nlbi_989911_2465335=pNg9fugukjq2uRKWrLFxRQAAAAA9OV9I3k3JIx/MvZPrwjr6; '
                     'incap_ses_713_989911=GPHHe6OSASRMZrqnNhblCQdHyWEAAAAAMuA5wAkB3I27PUpKjeGCkg==; '
                     'DKT_SESSION=Cg2nd2PwnGntjp478CWRPC/sdtlNDGllKDggqLIwok8R597ikHYD'
                     '+PoaYFBumnERt9AHEO35rU2fjcGYtbpV1FV'
                     '/0PTboXmg10tovWHCSiNFwsR3UbgPPcFyZTcGnEAhKDk4hxPEqNhauCk8wONoyFpW15BGP1wJKr/RW3m9WxQ=; '
                     'CUSTOMER_POSITION=false; '
                     'incap_ses_708_989911=xzAWAKXYhlGq8TNsxFLTCQ5HyWEAAAAAW7jttMRufYXsR1MkifeufA==; '
                     'tc_cj_v2=%5Ecl_%5Dny%5B%5D%5D_mmZZZZZZKPNJORJRSSOSMZZZ%5D; didomi_cookies=essential; '
                     'rxVisitor=16405809012209S2G595A50SRNBHE3SC233FR5DRJ8323; '
                     'baqend-speedkit-ab-test-info=%7B%22group%22%3A%22A%22%2C%22testId%22%3A%22100%22%7D; '
                     'baqend-speedkit-user-id=CwLuDXae38uCzSu7StZ17eSYF; '
                     'dtCookie=v_4_srv_6_sn_5ED6C322B102BA808C33DBB2B2A89B6D_app'
                     '-3A8397d8276578eb80_0_ol_0_perc_100000_mul_1; '
                     'didomi_token'
                     '=eyJ1c2VyX2lkIjoiMTdkZmEzZGUtNjE5OS02YWNlLTliNDYtYjEzMGZkNTc4YzVlIiwiY3JlYXRlZCI6IjIwMjEtMTItMjdUMDU6MDU6NTQuNjIzWiIsInVwZGF0ZWQiOiIyMDIxLTEyLTI3VDA1OjA1OjU0LjYyM1oiLCJ2ZW5kb3JzIjp7ImVuYWJsZWQiOlsiYzpzcGVlZGtpdC1kZUM2a01SRyIsImM6eXNhbmNlLTN4d0Z4OWU3IiwiYzpjb21lY29nZXItaE1BREx5RlgiLCJjOmRpYWxvZ3VlcG8tSFd4a0ZteGYiLCJjOnBpbnRlcmVzdC1WSkdGOTJGYSIsImM6YWJ0YXN0eWctWEVqQ0pnRUwiLCJjOmNvbm5leGl0eS1hQ1RDQjI0TCIsImM6dXNhYmlsbGFnLWFIZ2VDa3BpIiwiYzpmaXRhbmFseXQtUVpSQjd4NlYiLCJjOm1wdWxzZ2VybS1haHpXVlVQeCIsImM6Y29va2llc2FuLWFZN1dla0tyIiwiYzpleGNlbnRvcy1UQmZHeXdGZyIsImM6ZmFjZWJvb2tnLTZrR2NYaHIzIiwiYzpnb29nbGVhZHMtQWszMlRQREEiLCJjOmF3aW5nZXJtYS1qZlkzZHBnNCIsImM6bHVja3lvcmFuLXJXbkFkWFFVIiwiYzpkeW5hbWljeWktRVRlYXhIY2siLCJjOmViYXljb21tZS1OTlVqRlZtRiIsImM6aHVycmEtVnlrQkNKSkQiXX0sInB1cnBvc2VzIjp7ImVuYWJsZWQiOlsibGVzY29va2llLUhYMmg2VGpiIiwibWFya2V0aW5nLU1SWlZwcmVhIiwicGVyc29uYWxpcy10ZjhwWlRWSCIsImFuYWx5dGljcy1HTWdCV0dUaCJdfSwidmVyc2lvbiI6Mn0=; euconsent-v2=CPR3OD6PR3OD6AHABBENB5CgAAAAAAAAAAAAAAAAAAAA.YAAAAAAAAAAA; _dy_df_geo=India..Pune; _dy_geo=IN.AS.IN_MH.IN_MH_Pune; _dycnst=dg; _dycst=dk.w.c.ms.; _dyid=3102981008187279107; didomiVendorsConsent=c:speedkit-deC6kMRG,c:ysance-3xwFx9e7,c:comecoger-hMADLyFX,c:dialoguepo-HWxkFmxf,c:pinterest-VJGF92Fa,c:abtastyg-XEjCJgEL,c:connexity-aCTCB24L,c:usabillag-aHgeCkpi,c:fitanalyt-QZRB7x6V,c:mpulsgerm-ahzWVUPx,c:cookiesan-aY7WekKr,c:excentos-TBfGywFg,c:facebookg-6kGcXhr3,c:googleads-Ak32TPDA,c:awingerma-jfY3dpg4,c:luckyoran-rWnAdXQU,c:dynamicyi-ETeaxHck,c:ebaycomme-NNUjFVmF,c:hurra-VykBCJJD,; __ywtfpcvuid=38754017781640581555085; _ga=GA1.2.DKT-8e250a0c-ca72-4bd5-b3ff-2aba999adff3; _gid=GA1.2.2122066197.1640581555; _gcl_au=1.1.1469076819.1640581555; _fbp=fb.1.1640581555990.1341673303; _pin_unauth=dWlkPVpXVXlNekkyT1RrdE0yVTBOeTAwWVRVNExUbGpZVGt0T0dWbE5qRTJZVEpqT0RrMA; _HC_uu=MQZ8ck3r3kM-i-08zzzzzzzz; _HC_4147=MQZ8d-Mr3kM-lU0Ezzzzzzzz:C4147G108:1643173540; _HC_fr=:::1640581540; _dy_c_exps=; _dyid_server=3102981008187279107; _dy_c_att_exps=; _dyfs=1640581567767; _dy_toffset=-17; dtLatC=10; _uetsid=ad42371066d211eca3f7b97c7e97e34f; _uetvid=ad43c5d066d211ecbee3972eb3e2b4a5; _gat_transcript=1; rxvt=1640585190879|1640583372643; dtSa=true%7CC%7C-1%7CElectronic%20dartboard%20ED%20310%20with%206%20arrows%7C-%7C1640583390972%7C181569594_191%7Chttps%3A%2F%2Fwww.decathlon.de%2Fbrowse%2Fc0-alle-sportarten-a-z%2Fc1-darts%2Fc2-dartscheiben%2F_5F%2FN-eioci7%7C%7C%7C%7C; dtPC=6$181569594_191h-vRBEPMLPKLPMOUUDQCOTWDVCBESIWRAFA-0e0; RT="sl=0&ss=1640581564387&tt=0&obo=0&bcn=%2F%2F684d0d4b.akstat.io%2F&sh=&dm=decathlon.de&si=546c6e0a-d043-4811-b73f-7017b00fe9d9&nu=https%3A%2F%2Fwww.decathlon.de%2Fp%2Fdartscheibe-ed-310-elektronisch-mit-6-pfeilen%2F_%2FR-p-12794%3Fmc%3D8295139&cl=1640583390969&r=https%3A%2F%2Fwww.decathlon.de%2Fbrowse%2Fc0-alle-sportarten-a-z%2Fc1-darts%2Fc2-dartscheiben%2F_%2FN-eioci7&ul=1640583390998&hd=1640583391677"; CUSTOMER_LVP=1773995; incap_ses_704_989911=wX8ba8GdhnLrbt514xzFCdBQyWEAAAAA+C5iajUIK49bzTwAQ2lq3w==; incap_ses_706_989911=127QHfAcxirGLt/KyzfMCdFQyWEAAAAAuUc549aMS1gr9vpdm8gOQQ==; _dyjsession=7jhiy4bhie3ldnf9ptiahnbh6jo8bxup; dy_fs_page=www.decathlon.de%2Fp%2Fdartscheibe-ed-310-elektronisch-mit-6-pfeilen%2F_%2Fr-p-12794%3Fmc%3D8295139; _dy_csc_ses=7jhiy4bhie3ldnf9ptiahnbh6jo8bxup; _dy_soct=1048902.1117754.1640583395.7jhiy4bhie3ldnf9ptiahnbh6jo8bxup*1068821.1185234.1640583395*1029547.1057513.1640583395*1074786.1206052.1640583395; incap_ses_709_989911=ffXZNE0LoSdTjO4UPODWCdJQyWEAAAAAKXey2ZSuEQQNp7PhKEOy8g==; PLAY_SESSION=1707d69f8dc7f9bad5fb51128c5e85743cd30935-JSESSIONID=KIL6Y7sa9zBwFYIgq6dwp7v9v9sPfrM3yKaKc59ScHyBNqN-ezyd%21-954757393&APISERVER=API07; incap_ses_710_989911=NwzIdIMck3AqLki9s23aCdNQyWEAAAAAw6T8P1aubCHf7n6PLHoTLw==',
           'referer': 'https://www.decathlon.de/browse/c0-alle-sportarten-a-z/c1-darts/c2-dartscheiben/_/N-eioci7',
           'sec-ch-ua': '"Not A;Brand";v="99","Chromium";v="96","Google Chrome";v="96"',
           'sec-ch-ua-mobile': '?0',
           'sec-ch-ua-platform': '"Windows"',
           'sec-fetch-dest': 'empty',
           'sec-fetch-mode': 'no-cors',
           'sec-fetch-site': 'same-origin',
           'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64 ; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                         'Chrome/96.0.4664.110 Safari/537.36 '
           }
p1 =['154.28.67.106', '154.28.67.111', '154.28.67.116', '154.28.67.117', '154.28.67.125',
              '154.28.67.131',
              '154.28.67.133', '154.28.67.142', '154.28.67.156', '154.28.67.163', '154.28.67.173',
              '154.28.67.18',
              '154.28.67.182', '154.28.67.184', '154.28.67.20', '154.28.67.200', '154.28.67.210',
              '154.28.67.218',
              '154.28.67.222', '154.28.67.223', '154.28.67.231', '154.28.67.240', '154.28.67.243',
              '154.28.67.253',
              '154.28.67.39', '154.28.67.4', '154.28.67.49', '154.28.67.5', '154.28.67.61', '154.28.67.80',
              '154.28.67.81', '154.28.67.87', '154.28.67.88', '154.28.67.96', '154.28.67.99', '154.7.230.100',
              '154.7.230.101', '154.7.230.103', '154.7.230.107', '154.7.230.109', '154.7.230.130',
              '154.7.230.132',
              '154.7.230.14', '154.7.230.140', '154.7.230.147', '154.7.230.151', '154.7.230.156',
              '154.7.230.163',
              '154.7.230.170', '154.7.230.18', '154.7.230.183', '154.7.230.188', '154.7.230.189',
              '154.7.230.19',
              '154.7.230.190', '154.7.230.198', '154.7.230.204', '154.7.230.209', '154.7.230.235',
              '154.7.230.238',
              '154.7.230.246', '154.7.230.29', '154.7.230.41', '154.7.230.42', '154.7.230.51', '154.7.230.55',
              '154.7.230.60', '154.7.230.61', '154.7.230.74', '154.7.230.82', '154.7.230.89', '23.131.8.112',
              '23.131.8.115', '23.131.8.117', '23.131.8.12', '23.131.8.121', '23.131.8.124', '23.131.8.150',
              '23.131.8.161', '23.131.8.166', '23.131.8.171', '23.131.8.173', '23.131.8.176', '23.131.8.177',
              '23.131.8.181', '23.131.8.19', '23.131.8.192', '23.131.8.194', '23.131.8.199', '23.131.8.202',
              '23.131.8.203', '23.131.8.204', '23.131.8.207', '23.131.8.209', '23.131.8.213', '23.131.8.216',
              '23.131.8.225', '23.131.8.228', '23.131.8.231', '23.131.8.238', '23.131.8.254', '23.131.8.36',
              '23.131.8.5', '23.131.8.76', '23.131.8.93', '23.131.8.95', '23.131.8.99', '23.131.88.105',
              '23.131.88.12', '23.131.88.137', '23.131.88.139', '23.131.88.140', '23.131.88.145',
              '23.131.88.150',
              '23.131.88.151', '23.131.88.153', '23.131.88.154', '23.131.88.156', '23.131.88.165',
              '23.131.88.18',
              '23.131.88.191', '23.131.88.192', '23.131.88.194', '23.131.88.198', '23.131.88.202',
              '23.131.88.206',
              '23.131.88.220', '23.131.88.223', '23.131.88.228', '23.131.88.233', '23.131.88.24',
              '23.131.88.242',
              '23.131.88.244', '23.131.88.47', '23.131.88.63', '23.131.88.67', '23.131.88.73', '23.131.88.80',
              '23.131.88.81', '23.131.88.82', '23.131.88.88', '23.131.88.97', '23.170.144.149',
              '23.170.144.209',
              '23.170.144.212', '23.170.144.242', '23.170.144.83', '23.170.145.117', '23.170.145.167',
              '23.170.145.182', '23.170.145.19', '23.170.145.203', '23.226.17.101', '23.226.17.109',
              '23.226.17.112', '23.226.17.113', '23.226.17.115', '23.226.17.123', '23.226.17.129',
              '23.226.17.143',
              '23.226.17.148', '23.226.17.165', '23.226.17.186', '23.226.17.199', '23.226.17.201',
              '23.226.17.207',
              '23.226.17.210', '23.226.17.219', '23.226.17.220', '23.226.17.222', '23.226.17.229',
              '23.226.17.250',
              '23.226.17.254', '23.226.17.26', '23.226.17.33', '23.226.17.4', '23.226.17.49', '23.226.17.5',
              '23.226.17.55', '23.226.17.66', '23.226.17.7', '23.226.17.72', '23.226.17.78', '23.226.17.8',
              '23.226.17.86', '23.226.17.90', '23.226.17.93', '23.230.177.105', '23.230.177.110',
              '23.230.177.113',
              '23.230.177.121', '23.230.177.130', '23.230.177.14', '23.230.177.143', '23.230.177.15',
              '23.230.177.150', '23.230.177.154', '23.230.177.165', '23.230.177.173', '23.230.177.191',
              '23.230.177.196', '23.230.177.203', '23.230.177.206', '23.230.177.208', '23.230.177.217',
              '23.230.177.220', '23.230.177.221', '23.230.177.224', '23.230.177.228', '23.230.177.231',
              '23.230.177.235', '23.230.177.237', '23.230.177.241', '23.230.177.27', '23.230.177.38',
              '23.230.177.52', '23.230.177.61', '23.230.177.67', '23.230.177.72', '23.230.177.80',
              '23.230.177.88',
              '23.230.177.94', '23.230.177.99', '23.230.197.103', '23.230.197.106', '23.230.197.109',
              '23.230.197.11', '23.230.197.12', '23.230.197.122', '23.230.197.124', '23.230.197.146',
              '23.230.197.155', '23.230.197.156', '23.230.197.174', '23.230.197.179', '23.230.197.181',
              '23.230.197.196', '23.230.197.2', '23.230.197.201', '23.230.197.207', '23.230.197.208',
              '23.230.197.225', '23.230.197.227', '23.230.197.233', '23.230.197.236', '23.230.197.239',
              '23.230.197.240', '23.230.197.244', '23.230.197.251', '23.230.197.50', '23.230.197.52',
              '23.230.197.54', '23.230.197.60', '23.230.197.71', '23.230.197.80', '23.230.197.81',
              '23.230.197.84',
              '23.230.197.97', '23.230.74.102', '23.230.74.110', '23.230.74.116', '23.230.74.125',
              '23.230.74.133',
              '23.230.74.135', '23.230.74.14', '23.230.74.141', '23.230.74.149', '23.230.74.15',
              '23.230.74.157',
              '23.230.74.16', '23.230.74.170', '23.230.74.172', '23.230.74.174', '23.230.74.183',
              '23.230.74.187',
              '23.230.74.19', '23.230.74.198', '23.230.74.208', '23.230.74.212', '23.230.74.215',
              '23.230.74.23',
              '23.230.74.230', '23.230.74.231', '23.230.74.252', '23.230.74.30', '23.230.74.41', '23.230.74.57',
              '23.230.74.58', '23.230.74.59', '23.230.74.6', '23.230.74.75', '23.230.74.81', '23.230.74.88',
              '23.230.74.91', '23.27.222.108', '23.27.222.109', '23.27.222.134', '23.27.222.138',
              '23.27.222.159',
              '23.27.222.161', '23.27.222.164', '23.27.222.166', '23.27.222.178', '23.27.222.19',
              '23.27.222.195',
              '23.27.222.201', '23.27.222.202', '23.27.222.203', '23.27.222.208', '23.27.222.21',
              '23.27.222.211',
              '23.27.222.218', '23.27.222.223', '23.27.222.228', '23.27.222.234', '23.27.222.236',
              '23.27.222.242',
              '23.27.222.251', '23.27.222.253', '23.27.222.34', '23.27.222.61', '23.27.222.62', '23.27.222.69',
              '23.27.222.70', '23.27.222.72', '23.27.222.73', '23.27.222.74', '23.27.222.81', '23.27.222.93',
              '38.131.131.110', '38.131.131.114', '38.131.131.123', '38.131.131.125', '38.131.131.137',
              '38.131.131.142', '38.131.131.145', '38.131.131.147', '38.131.131.15', '38.131.131.154',
              '38.131.131.16', '38.131.131.17', '38.131.131.173', '38.131.131.18', '38.131.131.193',
              '38.131.131.204', '38.131.131.207', '38.131.131.227', '38.131.131.229', '38.131.131.233',
              '38.131.131.238', '38.131.131.246', '38.131.131.248', '38.131.131.250', '38.131.131.31',
              '38.131.131.36', '38.131.131.50', '38.131.131.58', '38.131.131.64', '38.131.131.70',
              '38.131.131.71',
              '38.131.131.74', '38.131.131.83', '38.131.131.94', '38.131.131.99', '38.75.75.104',
              '38.75.75.111',
              '38.75.75.112', '38.75.75.119', '38.75.75.120', '38.75.75.123', '38.75.75.127', '38.75.75.139',
              '38.75.75.14', '38.75.75.143', '38.75.75.155', '38.75.75.156', '38.75.75.158', '38.75.75.170',
              '38.75.75.179', '38.75.75.188', '38.75.75.2', '38.75.75.201', '38.75.75.231', '38.75.75.232',
              '38.75.75.241', '38.75.75.246', '38.75.75.251', '38.75.75.26', '38.75.75.29', '38.75.75.4',
              '38.75.75.44', '38.75.75.49', '38.75.75.56', '38.75.75.58', '38.75.75.62', '38.75.75.72',
              '38.75.75.76', '38.75.75.79', '38.75.75.88', '38.96.156.108', '38.96.156.112', '38.96.156.128',
              '38.96.156.131', '38.96.156.14', '38.96.156.142', '38.96.156.143', '38.96.156.149',
              '38.96.156.16',
              '38.96.156.163', '38.96.156.165', '38.96.156.169', '38.96.156.186', '38.96.156.188',
              '38.96.156.190',
              '38.96.156.192', '38.96.156.194', '38.96.156.199', '38.96.156.218', '38.96.156.236',
              '38.96.156.240',
              '38.96.156.252', '38.96.156.28', '38.96.156.32', '38.96.156.35', '38.96.156.56', '38.96.156.57',
              '38.96.156.6', '38.96.156.67', '38.96.156.77', '38.96.156.80', '38.96.156.83', '38.96.156.84',
              '38.96.156.89', '38.96.156.92', '45.238.157.100', '45.238.157.104', '45.238.157.106',
              '45.238.157.110', '45.238.157.116', '45.238.157.118', '45.238.157.119', '45.238.157.12',
              '45.238.157.123', '45.238.157.132', '45.238.157.14', '45.238.157.149', '45.238.157.15',
              '45.238.157.183', '45.238.157.186', '45.238.157.189', '45.238.157.2', '45.238.157.212',
              '45.238.157.214', '45.238.157.217', '45.238.157.22', '45.238.157.228', '45.238.157.23',
              '45.238.157.247', '45.238.157.43', '45.238.157.48', '45.238.157.51', '45.238.157.52',
              '45.238.157.53',
              '45.238.157.56', '45.238.157.61', '45.238.157.65', '45.238.157.72', '45.238.157.79',
              '45.238.157.8',
              '45.238.159.103', '45.238.159.107', '45.238.159.110', '45.238.159.114', '45.238.159.116',
              '45.238.159.123', '45.238.159.126', '45.238.159.144', '45.238.159.148', '45.238.159.15',
              '45.238.159.156', '45.238.159.165', '45.238.159.167', '45.238.159.183', '45.238.159.20',
              '45.238.159.208', '45.238.159.217', '45.238.159.220', '45.238.159.23', '45.238.159.230',
              '45.238.159.235', '45.238.159.237', '45.238.159.238', '45.238.159.24', '45.238.159.249',
              '45.238.159.251', '45.238.159.32', '45.238.159.34', '45.238.159.51', '45.238.159.6',
              '45.238.159.66',
              '45.238.159.77', '45.238.159.79', '45.238.159.82', '45.238.159.91']

start=time.time()
data = []
stock_condition = []

stock_message = []
stock_count1 = []
inputdata=[]

now = datetime.now()
f_s_d = datetime.now().strftime("%d-%m-")
def store_data_in_csv(dict):
    #df1=pd.DataFrame.from_dict(dict)
    #df1.to_csv('output1_0104_rehit1.csv', mode='a+', index=False, encoding="utf-8-sig", header=False)
    fieldnames = ['SKU_ID','Website','Country','RPC','MPC','Product_ID','Product_URL','Product_Name','Category_Path','Specification','Description','Currency','List_Price','Promo_Price','Discount','Brand','Rating_Count','Review_Count','Image_URLs','Variant','Variant_ID','Colour_of_Variant','Colour_Grouping','Seller_Name','Stock_Count','Stock_Condition','Stock_Message','Sustainability_Badge','Reason_Code','Crawling_TimeStamp','Cache_Page_Link','Extra1','Extra2','Extra3','Extra4','Extra5']

    with open(f'{f_s_d}Decathlon_de-1.csv', 'a+', encoding='utf-8-sig',newline='') as file:  # encoding='UTF8',
        writer = DictWriter(file, fieldnames=fieldnames)  # encoding='UTF8',encoding="cp1252",
        #writer.writeheader()
        writer.writerows(dict)
        file.close()

def get_csv():
    global varient_id
    file_path = r'input1.xlsx'
    # file_path = 'input_rehit.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet =workbook.active
    all_rows = sheet.max_row
    for i in range(1, all_rows):
        ID = sheet.cell(row=i + 1, column=1)
        ID2 = sheet.cell(row=i + 1, column=2)
        ID3 = sheet.cell(row=i + 1, column=3)
        ID4 = sheet.cell(row=i + 1, column=4)

        u_data = [str(ID.value),str(ID2.value),str(ID3.value),str(ID4.value)]
        inputdata.append(u_data)
    # print(len(inputdata))
    #return u_data
    # print(inputdata)
    return inputdata

def scrap():
    try:
        for i in inputdata:
            try:
                print("url",i)
                now = datetime.now()
                output_date = datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ")
                #links = 'https://www.decathlon.de/search?Ntt=' + str(u_data)
                links = i[3]
                skuid = i[2]
                country = i[1]
                website = i[0]

                Url_Product = links
                for _ in range(10):
                    """
                    p_auth = str("csimonra:h19VA2xZ")
                    p_host = random.choice(p1)
                    p_port = "29842"
                    proxy = {
                        'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                        'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
                    }
                    """

                    p_auth = str("csimonra:h19VA2xZ")
                    p_host = random.choice(p1)
                    p_port = "29842"
                    proxy1 = {
                        'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                        'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
                    }
                    prx = {'http': 'http://PRX001:w1t5ubvr@204.217.130.145:5432',
                           'https': 'https://PRX001:w1t5ubvr@204.217.130.145:5432'}

                    response = requests.get(url=links, headers=headers, proxies=proxy1,timeout=30)
                    # geturl_hit = requests.get(product_id_url,proxies=proxy, headers=headers, timeout=30)
                    if response.status_code == 200:
                        saving_data = response.text
                        break

                # response = requests.get(url=links, headers=headers, proxies=proxy,timeout=30)
                tree2 = lxml.html.fromstring(response.content)
                s=BeautifulSoup(response.text,'lxml')
                print(('url==', links))

                # --- Cachec Page Code
                # datazone = datetime.datetime.now()
                saving_data = response.text
                f_date = datetime.now().strftime("%d_%m_%Y")
                # strdate = datazone.day
                # strm = datazone.month
                # stry = datazone.year
                pageid = skuid
                cpid = pageid + '_' + f_date
                # ASS_folder = f"E:\ADIDAS_SavePages\Jdsports_uk\ASS"

                ASS_folder = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Decathlon_de\\PDP"
                sos_date_wise_folder = ASS_folder + f"\\{f_date}"
                if os.path.exists(sos_date_wise_folder):
                    pass
                else:
                    os.mkdir(sos_date_wise_folder)
                sos_filename = sos_date_wise_folder + "\\" + cpid + ".html"
                sos_filename = sos_filename.replace("+", "_").replace("-", "_")
                page_path = sos_filename.replace('/', '')
                print(page_path)
                page_path = page_path.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\',
                                              'https:////ecxus440.eclerx.com//cachepages//').replace('\\',
                                                                                                     '//').replace(
                    '//', '/')
                print(page_path)
                # Cache_Page_Link = page_path

                if os.path.exists(sos_filename):
                    with open(sos_filename, 'w', encoding='utf-8') as f:
                        f.write(saving_data)
                else:
                    with open(sos_filename, 'w', encoding='utf-8') as f:
                        f.write(saving_data)
                try:
                    if response.status_code == 200:
                        reason_Code = 'Success-PF'
                        product_id2=links.split('mc=')[1]
                        # product_id2=product_id2.split("&")[0]
                        # print(product_id2)
                        product_id1=''.join(tree2.xpath('//div[@class="product-summary-infos svelte-1y5ko7f"]/div/div[2]/span/text()'))
                        if len(product_id1)==0:
                            product_id1 = '-'
                        # print(product_id1) #
                        name = ''.join(tree2.xpath('//h1[@class="vtmn-typo_title-4 svelte-1y5ko7f"]/text()'))
                        if len(name) == 0:
                            name = '-'
                        # category_path1 = '>'.join(tree2.xpath('//nav[@class="breadcrumbs svelte-gxh85q"]/ol/li/a/text()'))
                        # category_path = category_path1.rstrip().replace('\n', '')
                        # if len(category_path)==0:
                        #     category_path = '-'
                        category_path = general.xpath(tree2, '//ol[@class="svelte-1xqc499"]/li', mode="set_tc", sep='>')
                        # print(('Category_Path', category_path))
                        discount1 = ''.join(tree2.xpath(
                            '//span[@class="prc__rate"]/text()'))[::]
                        if len(discount1) > 0:
                            try:
                                discount = discount1[1] + discount1[2] + discount1[3]
                               # list_price1 = ''.join(
                                #    tree2.xpath('//article[@class="product-main-infos svelte-9yk964"]/div/div/span[1]/span[1]/text()'))
                                list_price1=''.join(tree2.xpath('//*[@id="app"]/main/article/div[3]/div[1]/span[1]/span[1]/text()'))
                                list_price = list_price1.replace('\n', '').replace('€', '')
                                #promoprice1 = ''.join(
                                #    tree2.xpath('//article[@class="product-main-infos svelte-9yk964"]/div[3]/div[1]/div/text()'))
                                promoprice1=''.join(tree2.xpath('//*[@id="app"]/main/article/div[3]/div[1]/div/text()'))
                                promoprice2 = promoprice1.replace('€', '').replace('\n', '').strip()
                            except Exception as e:
                                discount = discount1[1] + discount1[2]
                                list_price1 = ''.join(
                                    tree2.xpath('//*[@id="app"]/main/article/div[3]/div[1]/span[1]/span[1]/text()'))
                                list_price = list_price1.replace('\n', '').replace('€', '')
                                # promoprice1 = ''.join(
                                #    tree2.xpath('//article[@class="product-main-infos svelte-9yk964"]/div[3]/div[1]/div/text()'))
                                promoprice1 = ''.join(
                                    tree2.xpath('//*[@id="app"]/main/article/div[3]/div[1]/div/text()'))
                                promoprice2 = promoprice1.replace('€', '').replace('\n', '').strip()
                        else:
                            # //article[@class="product-main-infos svelte-9yk964"]/div[3]/div[1]/div/text()
                            #list_price1 = ''.join(
                            #    tree2.xpath('//article[@class="product-main-infos svelte-9yk964"]/div[3]/div[1]/div/text()'))

                            list_price1=''.join(tree2.xpath('//*[@id="app"]/main/article/div[3]/div[1]/div/text()'))
                            list_price = list_price1.replace('€', '').replace('\n', '').strip()
                            promoprice2 = list_price
                            discount = '-'
                        color = ' '.join(tree2.xpath('//span[@class="current-model-color svelte-1gybc6v"]/text()'))
                        if len(color) == 0:
                            color = '-'
                        color_grouping = ' | '.join(tree2.xpath('//div[@class="model-choice svelte-1gybc6v"]/button/@aria-label'))
                        if len(color_grouping) == 0:
                            color_grouping = '-'

                        #rating_count = ''.join(tree2.xpath(
                        #    '//article[@class="product-main-infos svelte-9yk964"]/div/div/div/div/a/span[1]/strong/text()'))
                        rating_count=''.join(tree2.xpath('//strong[@class="svelte-3gle04"]/text()'))
                        if len(rating_count) == 0:
                            rating_count = '-'

                        #review_Count = ''.join(tree2.xpath('//article[@class="product-main-infos svelte-9yk964"]/div/div/div/div/a/span[2]/text()'))
                        review_Count=''.join(tree2.xpath('//span[@class="count svelte-3gle04"]/text()'))
                        if len(review_Count)>0:
                            review_Count=review_Count[1]+review_Count[2]
                        if len(review_Count)==0:
                            review_Count = '-'

                        sellername = tree2.xpath('//a[@class="seller-link seller-btn"]/text()')
                        if sellername:
                            sellername = sellername[0]
                        sellername = general.clean(sellername)



                        # specification1 = ' '.join(tree2.xpath('//div[@id="product-composition"]/p/text()'))
                        # specification = specification1.replace("\n", '').replace(',', '|').replace('\xa0', '')
                        # if len(specification) == 0:
                        #     specification = '-'
                        spectext = ''
                        specval = general.xpath(tree2,'//div[@class="content"]/div[@class="tech-info svelte-16jroli"]',mode='set')
                        if specval:
                            for eachspec in specval:
                                speheader = general.xpath(eachspec,'.//h3',mode='tc')
                                specval = general.xpath(eachspec, './/p', mode='tc')
                                spevlen = len(specval)
                                if spevlen> 200:
                                    pass
                                else:
                                    if spectext == '':
                                        spectext = speheader + ':' + specval
                                    else:
                                        spectext = spectext + '|' + speheader + ':' + specval
                            specification = spectext
                        else:
                            specification = ''



                        description = ' '.join(
                            tree2.xpath('//p[@class="svelte-1y5ko7f"]/text()'))
                        description = description.replace('\xa0', '')
                        if len(description)==0:
                            description = '-'

                        image_url1 = ''.join(tree2.xpath('//div[@id="thumbnails-slider"]/div/section/button/img/@srcset'))
                        image_url2 = re.findall(
                            'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', image_url1)
                        image_url = '   '.join(image_url2).replace('  ', ' |')
                        if len(image_url) == 0:
                            try:
                                image_url = ''.join(tree2.xpath('//*[@id="app"]/main/article/div[2]/div[2]/div[2]/div/section[1]/img/@srcset'))
                            except:
                                image_url='-'
                            #image_url="|".join(image_url)

                        varient = [x.replace('\n    ', '') for x in
                                       (tree2.xpath('//select[@class="vh"]/option[position()>1]/text()'))]
                        if len(varient)==0:
                            varient = '[-]'
                        if varient == '-':
                            varient = [x.replace('\n    ', '') for x in
                                       (tree2.xpath('//div[@class="size-option only-size svelte-1dkais7"]/span[@class="size"]/text()'))]


                        varient_id = re.findall(r'"Offer","sku":"(.*?)","price":', response.text)
                        stock_count1.clear()
                        stock_condition.clear()
                        stock_message.clear()
                        for sku_id in varient_id:
                            url1 = 'https://www.decathlon.de/de/ajax/nfs/stocks/online?skuIds=' + sku_id
                            # varient_id = sku_id
                            for i in range(8):
                                try:
                                    p_auth = str("csimonra:h19VA2xZ")
                                    p_host = random.choice(p1)
                                    p_port = "29842"
                                    proxy1 = {
                                        'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                                        'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
                                    }
                                    response1 = requests.get(url=url1, headers=headers,proxies=proxy1)
                                    if response1.status_code == 200:
                                        break
                                except Exception as e:
                                    print('retrying for 2nd response')
                            stock1 = response1.json()
                            stock1 = list(stock1.values())[0]
                            stock_count = list(stock1.values())[0]
                            stock_count1.append(stock_count)

                            if stock_count > 10:
                                stock_condition.append('In Stock')
                                stock_message.append('verfügbar')
                            elif 0 < stock_count and stock_count <= 10:
                                stock_condition.append('In Stock')  # In Stock
                                stock_message.append(str(stock_count) + ' verfügbar')
                            elif stock_count == 0:
                                stock_message.append('Momentan nicht lieferbar')
                                stock_condition.append('Out of Stock')
                        if len(varient_id) == 0:
                                # print(len(varient_id))
                                varient_id = '-'
                        try:
                            brand = s.find('a', class_='brand-logo brand-logo-link svelte-12flqeg')['aria-label']

                        except:
                            brand =tree2.xpath('//p[@class="product-brand-label svelte-1oztfpt"]/text()')

                        specification=general.clean(specification)
                        description=general.clean(description)
                        category_path=general.clean(category_path)
                        if varient_id != '-':
                            for i, v, stc, stm, stc1 in zip(varient, varient_id, stock_condition, stock_message, stock_count1):
                                data.clear()
                                prod_data = {
                                    'SKU_ID': skuid,
                                    'Website': website,
                                    'Country': country,
                                    'RPC': product_id2,
                                    'MPC': '-',
                                    'Product_ID': product_id2,
                                    'Product_URL': Url_Product,
                                    'Product_Name': name,
                                    'Category_Path': category_path,
                                    'Specification': specification,
                                    'Description': description,
                                    'Currency': 'EURO',
                                    'List_Price': list_price,  # .replace('€','').replace('\n','').strip(),
                                    'Promo_Price': promoprice2,
                                    'Discount': discount,
                                    'Brand': brand,
                                    'Rating_Count': rating_count,
                                    'Review_Count': review_Count,
                                    'Image_URLs': image_url,
                                    'Variant': i,
                                    'Variant_ID': v,
                                    'Colour_of_Variant': color,
                                    'Colour_Grouping': color_grouping,
                                    'Seller_Name': sellername,
                                    'Stock_Count': stc1,
                                    'Stock_Condition': stc,
                                    'Stock_Message': stm,
                                    'Sustainability_Badge': 'Not Available',
                                    'Reason_Code': reason_Code,
                                    'Crawling_TimeStamp': output_date,
                                    'Cache_Page_Link':page_path,
                                    'Extra1': '-',
                                    'Extra2': '-',
                                    'Extra3': '-',
                                    'Extra4': '-',
                                    'Extra5': '-'

                                }
                                data.append(prod_data)
                                print(data)
                                store_data_in_csv(data)
                        else:
                            data.clear()
                            prod_data = {
                                'SKU_ID': skuid,
                                'Website': website,
                                'Country': country,
                                'RPC': product_id2,
                                'MPC': '-',
                                'Product_ID': product_id2,
                                'Product_URL': Url_Product,
                                'Product_Name': name,
                                'Category_Path': category_path,
                                'Specification': specification,
                                'Description': description,
                                'Currency': 'EURO',
                                'List_Price': list_price,  # .replace('€','').replace('\n','').strip(),
                                'Promo_Price': promoprice2,
                                'Discount': discount,
                                'Brand': brand,
                                'Rating_Count': rating_count,
                                'Review_Count': review_Count,
                                'Image_URLs': image_url,
                                'Variant': '-',
                                'Variant_ID': '-',
                                'Colour_of_Variant': color,
                                'Colour_Grouping': color_grouping,
                                'Seller_Name': 'Not Available',
                                'Stock_Count': '-',
                                'Stock_Condition': '-',
                                'Stock_Message': '-',
                                'Sustainability_Badge': 'Not Available',
                                'Reason_Code': reason_Code,
                                'Crawling_TimeStamp': output_date,
                                'Cache_Page_Link': page_path,
                                'Extra1': '-',
                                'Extra2': '-',
                                'Extra3': '-',
                                'Extra4': '-',
                                'Extra5': '-'

                            }
                            data.append(prod_data)
                            print(data)
                            store_data_in_csv(data)
                        # else:
                        #     data.clear()
                        #     prod_data = {
                        #         'SKU_ID': skuid,
                        #         'Website': website,
                        #         'Country': country,
                        #         'RPC': product_id2,
                        #         'MPC': '-',
                        #         'Product_ID': product_id2,
                        #         'Product_URL': Url_Product,
                        #         'Product_Name': name,
                        #         'Category_Path': category_path,
                        #         'Specification': specification,
                        #         'Description': description,
                        #         'Currency': 'EURO',
                        #         'List_Price': list_price,  # .replace('€','').replace('\n','').strip(),
                        #         'Promo_Price': promoprice2,
                        #         'Discount': discount,
                        #         'Brand': brand,
                        #         'Rating_Count': rating_count,
                        #         'Review_Count': review_Count,
                        #         'Image_URLs': image_url,
                        #         'Variant': '-',
                        #         'Variant_ID': '-',
                        #         'Colour_of_Variant': color,
                        #         'Colour_Grouping': color_grouping,
                        #         'Seller_Name': 'Not Available',
                        #         'Stock_Count': '-',
                        #         'Stock_Condition': 'In Stock',
                        #         'Stock_Message': '-',
                        #         'Sustainability_Badge': 'Not Available',
                        #         'Reason_Code': reason_Code,
                        #         'Crawling_TimeStamp': output_date,
                        #         'Cache_Page_Link': page_path,
                        #         'Extra1': '-',
                        #         'Extra2': '-',
                        #         'Extra3': '-',
                        #         'Extra4': '-',
                        #         'Extra5': '-'
                        #
                        #     }
                        #     data.append(prod_data)
                        #     print(data)
                        #     store_data_in_csv(data)
                    else:
                        reason_Code = 'Success-PNF'
                        # print(response.status_code)
                        data.clear()
                        prod_data = {
                            'SKU_ID': skuid,
                            'Website': website,
                            'Country': country,
                            'RPC': '-',
                            'MPC': '-',
                            'Product_ID': '-',
                            'Product_URL': links,
                            'Product_Name': '-',
                            'Category_Path': '-',
                            'Specification': '-',
                            'Description': '-',
                            'Currency': '-',
                            'List_Price': '-',  # .replace('€','').replace('\n','').strip(),
                            'Promo_Price': '-',
                            'Discount': '-',
                            'Brand': '-',
                            'Rating_Count': '-',
                            'Review_Count': '-',
                            'Image_URLs': '-',
                            'Variant': '-',
                            'Variant_ID': '-',
                            'Colour_of_Variant': '-',
                            'Colour_Grouping': '-',
                            'Seller_Name': '-',
                            'Stock_Count': '-',
                            'Stock_Condition': '-',
                            'Stock_Message': '-',
                            'Sustainability_Badge': '-',
                            'Reason_Code': 'Success_PNF',
                            'Crawling_TimeStamp': output_date,
                            'Cache_Page_Link': '_',
                            'Extra1': '-',
                            'Extra2': '-',
                            'Extra3': '-',
                            'Extra4': '-',
                            'Extra5': '-'

                        }
                        data.append(prod_data)
                        print(data)
                        store_data_in_csv(data)
                        # data_to_csv(data)

                except Exception as e:
                    # print(e)
                    #break
                    prod_data = {
                        'SKU_ID': skuid,
                        'Website': website,
                        'Country': country,
                        'RPC': '-',
                        'MPC': '-',
                        'Product_ID': '-',
                        'Product_URL': links,
                        'Product_Name': '-',
                        'Category_Path': '-',
                        'Specification': '-',
                        'Description': '-',
                        'Currency': '-',
                        'List_Price': '-',  # .replace('€','').replace('\n','').strip(),
                        'Promo_Price': '-',
                        'Discount': '-',
                        'Brand': '-',
                        'Rating_Count': '-',
                        'Review_Count': '-',
                        'Image_URLs': '-',
                        'Variant': '-',
                        'Variant_ID': '-',
                        'Colour_of_Variant': '-',
                        'Colour_Grouping': '-',
                        'Seller_Name': '-',
                        'Stock_Count': '-',
                        'Stock_Condition': '-',
                        'Stock_Message': '-',
                        'Sustainability_Badge': '-',
                        'Reason_Code': 'Blocked',
                        'Crawling_TimeStamp': output_date,
                        'Cache_Page_Link': '_',
                        'Extra1': '-',
                        'Extra2': '-',
                        'Extra3': '-',
                        'Extra4': '-',
                        'Extra5': '-'
                    }
                    data.append(prod_data)
                    print(data)
                    store_data_in_csv(data)
            except Exception as e:
                print(e)
                print("response isuue")
                continue
            # end = time.time()
            # total_time = end - start
            # print(total_time)
    except Exception as e:
        print("blocking issue")
if __name__ == '__main__':
    get_csv()
    scrap()
    end = time.time()
    total_time = end - start
    print(total_time)
