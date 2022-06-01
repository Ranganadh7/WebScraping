import os
import re
import  random
import general
import requests
import xlrd
import  openpyxl
import json
import pandas as pd
from datetime import datetime
from scrapy.http import HtmlResponse
from lxml import  html
from bs4 import  BeautifulSoup
from csv import DictWriter

headers={
"Accept": 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
'Accept-Encoding': 'gzip, deflate, br',
'Accept-Language': 'en-US,en;q=0.9',
'Cache-Control': 'max-age=0',
'Connection': 'keep-alive',
'Cookie': 'consentTimeStamp=2022011217; consentStateNew=accept_all; consentState=accepted; consentTime=2021110410; userconsent_googletagmanager=true; userconsent_dynamicyield=true; sizeSystem=EU; basestore-id=deichmann-de-children; countryPreference=de-DE; _shop=deichmann_DE; _shopentity=dsd; _shopversion=H6; isInternalIP=false; _devicetype=d; _currency=EUR; _language=de; _country=DE; _gcl_au=1.1.789027178.1641990204; _ga=GA1.2.1812772137.1641990205; _gid=GA1.2.1294935352.1641990205; _cs_c=0; _fbp=fb.1.1641990205690.1515340812; _dyjsession=incsa29x786b4k0mo3c7rjsnwp0vagoj; _dy_csc_ses=incsa29x786b4k0mo3c7rjsnwp0vagoj; _dy_c_exps=; _dycnst=dg; _dyid=3436247513758158137; _dycst=dk.w.c.ms.; _dy_geo=IN.AS.IN_.IN__; _dy_df_geo=India..; _dyid_server=3436247513758158137; _dy_c_att_exps=; _dy_toffset=0; _cs_mk=0.9390299788844851_1642058847730; BIGipServer~DMZ-ESHOP-PUBLIC-01~POOL_HYB6-VAJSON-DEICH-DE_HTTP_DE-DE=rd3o00000000000000000000ffffac12360eo80; TS018eb9f9=0162db5bb2a9d00190d3b9f1604a783f6532dc234bd5de3b292dfa67b2aeb5bf8ab487eba4c9efa04b55cb3598c78bd09e5d898b66aeb4ac1f3a934ebf1152edaf81a04b311ad32f6d92283f430e15ac4a5570c918; _gasessionid=1642058982426.xlp2hvp7l; _originalLocation=https://www.deichmann.com/de-de/kinder-schuhe/sportschuhe/fussballschuhe/c-css3a; _uetsid=fa9ec800743d11eca09f334a52b118ea; _uetvid=fa9ed3e0743d11ecbec5d5fefdaacd59; _dy_soct=1073424.1200856.1642057015.incsa29x786b4k0mo3c7rjsnwp0vagoj*1076800.1212539.1642057015.incsa29x786b4k0mo3c7rjsnwp0vagoj*1040338.1129343.1642058061.incsa29x786b4k0mo3c7rjsnwp0vagoj*1040338.1129342.1642058958.incsa29x786b4k0mo3c7rjsnwp0vagoj*1076800.1220482.1642058982.incsa29x786b4k0mo3c7rjsnwp0vagoj*1080760.1227235.1642058982*1080761.1227237.1642058982*1080762.1227238.1642058982*1080763.1227239.1642058982*1080764.1227240.1642058982*1080765.1227241.1642058983*1080766.1227242.1642058983*1080767.1227243.1642058983*1080768.1227244.1642058983*1080769.1227245.1642058983*1009113.1014840.1642058983; _cs_id=89bc40eb-0cc8-a2c3-89df-523ff2f95de8.1641990205.2.1642058983.1642057020.1.1676154205092; _cs_s=12.0.0.1642060783385',
'Host': 'www.deichmann.com',
'If-None-Match': 'W/"1886a0-/ZzDNoenC4zkfqlD8cIWdZljOKg"',
'sec-ch-ua': '" Not;A Brand";v="99", "Microsoft Edge";v="97", "Chromium";v="97"',
'sec-ch-ua-mobile': '?0',
'sec-ch-ua-platform': "Windows",
'Sec-Fetch-Dest': 'document',
'Sec-Fetch-Mode': 'navigate',
'Sec-Fetch-Site': 'same-origin',
'Sec-Fetch-User': '?1',
'Upgrade-Insecure-Requests': '1',
'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36 Edg/97.0.1072.55'
}
datazone = datetime.now()
f_date = datazone.strftime("%d_%m_")
def store_data_in_csv(dict):
    fieldnames = ['SKU_ID', 'Website', 'Country', 'RPC', 'MPC', 'Product_ID', 'Product_URL', 'Product_Name',
                  'Category_Path', 'Specification', 'Description', 'Currency', 'List_Price', 'Promo_Price',
                  'Discount', 'Brand', 'Rating_Count', 'Review_Count', 'Image_URLs', 'Variant', 'Variant_ID',
                  'Colour_of_Variant', 'Colour_Grouping', 'Seller_Name', 'Stock_Count', 'Stock_Condition',
                  'Stock_Message', 'Sustainability_Badge', 'Reason_Code', 'Crawling_TimeStamp',
                  'Cache_Page_Link', 'Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5']

    with open(f'{f_date}Deichmann_de-1.csv', 'a+' , encoding='utf-8-sig', newline='') as file:
        writer = DictWriter(file, fieldnames=fieldnames)
        writer.writerows(dict)
        file.close()

# file_path=r'ASS-1.xlsx'
file_path=r'ASS-1.xlsx'
# file_path='Pending_One.xlsx'
def extract_xl():
    workbook= openpyxl.load_workbook(file_path)
    sheet_obj= workbook.active
    all_rows=sheet_obj.max_row

    inputdata=[]
    for i in range(1,all_rows):
        website_name = sheet_obj.cell(row=i + 1, column=1)
        country = sheet_obj.cell(row=i + 1, column=2)
        sku_id=sheet_obj.cell(row=i+1,column=4)
        urls= sheet_obj.cell(row=i+1,column=3)

        u_data=[website_name.value, country.value, sku_id.value, urls.value]
        inputdata.append(u_data)
    return inputdata

input_data = extract_xl()


all_product_data = []
data_record = []
pos_count = 1
c = 0
request_failed_arr = []
for idx, i_data in enumerate(input_data):
    website_name = i_data[0]
    country = i_data[1]
    sku_id = str(i_data[2])
    pdp_url = i_data[3]
    #pdp_url = 'https://www.deichmann.com/de-de/p-m01909091/01926436'
    #pdp_url = 'https://www.deichmann.com/de-de/p-m01906176/01731364'

    try:
        for _ in range(15):
            p1 = ['154.28.67.106', '154.28.67.111', '154.28.67.116', '154.28.67.117', '154.28.67.125',
              '154.28.67.131',
              '154.28.67.133', '154.28.67.142', '154.28.67.156', '154.28.67.163', '154.28.67.173',
              '154.28.67.18',
              '154.28.67.182', '154.28.67.184', '154.28.67.20', '154.28.67.200', '154.28.67.210',
              '154.28.67.218',
              '154.28.67.222', '154.28.67.223', '154.28.67.231', '154.28.67.240', '154.28.67.243',
              '154.28.67.253',
              '154.28.67.39', '154.28.67.4', '154.28.67.49', '154.28.67.5', '154.28.67.61', '154.28.67.80',
              '154.28.67.81', '154.28.67.87', '154.28.67.88', '154.28.67.96', '154.28.67.99', '154.7.230.100',
              '154.7.230.101', '154.7.230.103', '154.7.230.107', '154.7.230.109', '154.7.230.130',
              '154.7.230.132',
              '154.7.230.14', '154.7.230.140', '154.7.230.147', '154.7.230.151', '154.7.230.156',
              '154.7.230.163',
              '154.7.230.170', '154.7.230.18', '154.7.230.183', '154.7.230.188', '154.7.230.189',
              '154.7.230.19',
              '154.7.230.190', '154.7.230.198', '154.7.230.204', '154.7.230.209', '154.7.230.235',
              '154.7.230.238',
              '154.7.230.246', '154.7.230.29', '154.7.230.41', '154.7.230.42', '154.7.230.51', '154.7.230.55',
              '154.7.230.60', '154.7.230.61', '154.7.230.74', '154.7.230.82', '154.7.230.89', '23.131.8.112',
              '23.131.8.115', '23.131.8.117', '23.131.8.12', '23.131.8.121', '23.131.8.124', '23.131.8.150',
              '23.131.8.161', '23.131.8.166', '23.131.8.171', '23.131.8.173', '23.131.8.176', '23.131.8.177',
              '23.131.8.181', '23.131.8.19', '23.131.8.192', '23.131.8.194', '23.131.8.199', '23.131.8.202',
              '23.131.8.203', '23.131.8.204', '23.131.8.207', '23.131.8.209', '23.131.8.213', '23.131.8.216',
              '23.131.8.225', '23.131.8.228', '23.131.8.231', '23.131.8.238', '23.131.8.254', '23.131.8.36',
              '23.131.8.5', '23.131.8.76', '23.131.8.93', '23.131.8.95', '23.131.8.99', '23.131.88.105',
              '23.131.88.12', '23.131.88.137', '23.131.88.139', '23.131.88.140', '23.131.88.145',
              '23.131.88.150',
              '23.131.88.151', '23.131.88.153', '23.131.88.154', '23.131.88.156', '23.131.88.165',
              '23.131.88.18',
              '23.131.88.191', '23.131.88.192', '23.131.88.194', '23.131.88.198', '23.131.88.202',
              '23.131.88.206',
              '23.131.88.220', '23.131.88.223', '23.131.88.228', '23.131.88.233', '23.131.88.24',
              '23.131.88.242',
              '23.131.88.244', '23.131.88.47', '23.131.88.63', '23.131.88.67', '23.131.88.73', '23.131.88.80',
              '23.131.88.81', '23.131.88.82', '23.131.88.88', '23.131.88.97', '23.170.144.149',
              '23.170.144.209',
              '23.170.144.212', '23.170.144.242', '23.170.144.83', '23.170.145.117', '23.170.145.167',
              '23.170.145.182', '23.170.145.19', '23.170.145.203', '23.226.17.101', '23.226.17.109',
              '23.226.17.112', '23.226.17.113', '23.226.17.115', '23.226.17.123', '23.226.17.129',
              '23.226.17.143',
              '23.226.17.148', '23.226.17.165', '23.226.17.186', '23.226.17.199', '23.226.17.201',
              '23.226.17.207',
              '23.226.17.210', '23.226.17.219', '23.226.17.220', '23.226.17.222', '23.226.17.229',
              '23.226.17.250',
              '23.226.17.254', '23.226.17.26', '23.226.17.33', '23.226.17.4', '23.226.17.49', '23.226.17.5',
              '23.226.17.55', '23.226.17.66', '23.226.17.7', '23.226.17.72', '23.226.17.78', '23.226.17.8',
              '23.226.17.86', '23.226.17.90', '23.226.17.93', '23.230.177.105', '23.230.177.110',
              '23.230.177.113',
              '23.230.177.121', '23.230.177.130', '23.230.177.14', '23.230.177.143', '23.230.177.15',
              '23.230.177.150', '23.230.177.154', '23.230.177.165', '23.230.177.173', '23.230.177.191',
              '23.230.177.196', '23.230.177.203', '23.230.177.206', '23.230.177.208', '23.230.177.217',
              '23.230.177.220', '23.230.177.221', '23.230.177.224', '23.230.177.228', '23.230.177.231',
              '23.230.177.235', '23.230.177.237', '23.230.177.241', '23.230.177.27', '23.230.177.38',
              '23.230.177.52', '23.230.177.61', '23.230.177.67', '23.230.177.72', '23.230.177.80',
              '23.230.177.88',
              '23.230.177.94', '23.230.177.99', '23.230.197.103', '23.230.197.106', '23.230.197.109',
              '23.230.197.11', '23.230.197.12', '23.230.197.122', '23.230.197.124', '23.230.197.146',
              '23.230.197.155', '23.230.197.156', '23.230.197.174', '23.230.197.179', '23.230.197.181',
              '23.230.197.196', '23.230.197.2', '23.230.197.201', '23.230.197.207', '23.230.197.208',
              '23.230.197.225', '23.230.197.227', '23.230.197.233', '23.230.197.236', '23.230.197.239',
              '23.230.197.240', '23.230.197.244', '23.230.197.251', '23.230.197.50', '23.230.197.52',
              '23.230.197.54', '23.230.197.60', '23.230.197.71', '23.230.197.80', '23.230.197.81',
              '23.230.197.84',
              '23.230.197.97', '23.230.74.102', '23.230.74.110', '23.230.74.116', '23.230.74.125',
              '23.230.74.133',
              '23.230.74.135', '23.230.74.14', '23.230.74.141', '23.230.74.149', '23.230.74.15',
              '23.230.74.157',
              '23.230.74.16', '23.230.74.170', '23.230.74.172', '23.230.74.174', '23.230.74.183',
              '23.230.74.187',
              '23.230.74.19', '23.230.74.198', '23.230.74.208', '23.230.74.212', '23.230.74.215',
              '23.230.74.23',
              '23.230.74.230', '23.230.74.231', '23.230.74.252', '23.230.74.30', '23.230.74.41', '23.230.74.57',
              '23.230.74.58', '23.230.74.59', '23.230.74.6', '23.230.74.75', '23.230.74.81', '23.230.74.88',
              '23.230.74.91', '23.27.222.108', '23.27.222.109', '23.27.222.134', '23.27.222.138',
              '23.27.222.159',
              '23.27.222.161', '23.27.222.164', '23.27.222.166', '23.27.222.178', '23.27.222.19',
              '23.27.222.195',
              '23.27.222.201', '23.27.222.202', '23.27.222.203', '23.27.222.208', '23.27.222.21',
              '23.27.222.211',
              '23.27.222.218', '23.27.222.223', '23.27.222.228', '23.27.222.234', '23.27.222.236',
              '23.27.222.242',
              '23.27.222.251', '23.27.222.253', '23.27.222.34', '23.27.222.61', '23.27.222.62', '23.27.222.69',
              '23.27.222.70', '23.27.222.72', '23.27.222.73', '23.27.222.74', '23.27.222.81', '23.27.222.93',
              '38.131.131.110', '38.131.131.114', '38.131.131.123', '38.131.131.125', '38.131.131.137',
              '38.131.131.142', '38.131.131.145', '38.131.131.147', '38.131.131.15', '38.131.131.154',
              '38.131.131.16', '38.131.131.17', '38.131.131.173', '38.131.131.18', '38.131.131.193',
              '38.131.131.204', '38.131.131.207', '38.131.131.227', '38.131.131.229', '38.131.131.233',
              '38.131.131.238', '38.131.131.246', '38.131.131.248', '38.131.131.250', '38.131.131.31',
              '38.131.131.36', '38.131.131.50', '38.131.131.58', '38.131.131.64', '38.131.131.70',
              '38.131.131.71',
              '38.131.131.74', '38.131.131.83', '38.131.131.94', '38.131.131.99', '38.75.75.104',
              '38.75.75.111',
              '38.75.75.112', '38.75.75.119', '38.75.75.120', '38.75.75.123', '38.75.75.127', '38.75.75.139',
              '38.75.75.14', '38.75.75.143', '38.75.75.155', '38.75.75.156', '38.75.75.158', '38.75.75.170',
              '38.75.75.179', '38.75.75.188', '38.75.75.2', '38.75.75.201', '38.75.75.231', '38.75.75.232',
              '38.75.75.241', '38.75.75.246', '38.75.75.251', '38.75.75.26', '38.75.75.29', '38.75.75.4',
              '38.75.75.44', '38.75.75.49', '38.75.75.56', '38.75.75.58', '38.75.75.62', '38.75.75.72',
              '38.75.75.76', '38.75.75.79', '38.75.75.88', '38.96.156.108', '38.96.156.112', '38.96.156.128',
              '38.96.156.131', '38.96.156.14', '38.96.156.142', '38.96.156.143', '38.96.156.149',
              '38.96.156.16',
              '38.96.156.163', '38.96.156.165', '38.96.156.169', '38.96.156.186', '38.96.156.188',
              '38.96.156.190',
              '38.96.156.192', '38.96.156.194', '38.96.156.199', '38.96.156.218', '38.96.156.236',
              '38.96.156.240',
              '38.96.156.252', '38.96.156.28', '38.96.156.32', '38.96.156.35', '38.96.156.56', '38.96.156.57',
              '38.96.156.6', '38.96.156.67', '38.96.156.77', '38.96.156.80', '38.96.156.83', '38.96.156.84',
              '38.96.156.89', '38.96.156.92', '45.238.157.100', '45.238.157.104', '45.238.157.106',
              '45.238.157.110', '45.238.157.116', '45.238.157.118', '45.238.157.119', '45.238.157.12',
              '45.238.157.123', '45.238.157.132', '45.238.157.14', '45.238.157.149', '45.238.157.15',
              '45.238.157.183', '45.238.157.186', '45.238.157.189', '45.238.157.2', '45.238.157.212',
              '45.238.157.214', '45.238.157.217', '45.238.157.22', '45.238.157.228', '45.238.157.23',
              '45.238.157.247', '45.238.157.43', '45.238.157.48', '45.238.157.51', '45.238.157.52',
              '45.238.157.53',
              '45.238.157.56', '45.238.157.61', '45.238.157.65', '45.238.157.72', '45.238.157.79',
              '45.238.157.8',
              '45.238.159.103', '45.238.159.107', '45.238.159.110', '45.238.159.114', '45.238.159.116',
              '45.238.159.123', '45.238.159.126', '45.238.159.144', '45.238.159.148', '45.238.159.15',
              '45.238.159.156', '45.238.159.165', '45.238.159.167', '45.238.159.183', '45.238.159.20',
              '45.238.159.208', '45.238.159.217', '45.238.159.220', '45.238.159.23', '45.238.159.230',
              '45.238.159.235', '45.238.159.237', '45.238.159.238', '45.238.159.24', '45.238.159.249',
              '45.238.159.251', '45.238.159.32', '45.238.159.34', '45.238.159.51', '45.238.159.6',
              '45.238.159.66',
              '45.238.159.77', '45.238.159.79', '45.238.159.82', '45.238.159.91']
            p_auth = str("csimonra:h19VA2xZ")
            p_host = random.choice(p1)
            p_port = "29842"
            proxy = {
                'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
            }

            proxy_iplum = ["lum-customer-c_127755f5-zone-us_zone-ip-205.237.95.118","lum-customer-c_127755f5-zone-us_zone-ip-205.237.94.12","lum-customer-c_127755f5-zone-us_zone-ip-205.237.93.15","lum-customer-c_127755f5-zone-us_zone-ip-168.151.138.24","lum-customer-c_127755f5-zone-us_zone-ip-66.56.81.81","lum-customer-c_127755f5-zone-us_zone-ip-168.151.101.195","lum-customer-c_127755f5-zone-us_zone-ip-168.151.185.214","lum-customer-c_127755f5-zone-us_zone-ip-168.151.166.124","lum-customer-c_127755f5-zone-us_zone-ip-216.19.221.179","lum-customer-c_127755f5-zone-us_zone-ip-168.151.174.51","lum-customer-c_127755f5-zone-us_zone-ip-74.85.208.20","lum-customer-c_127755f5-zone-us_zone-ip-168.151.114.58","lum-customer-c_127755f5-zone-us_zone-ip-168.151.119.236","lum-customer-c_127755f5-zone-us_zone-ip-198.240.89.44","lum-customer-c_127755f5-zone-us_zone-ip-168.151.199.211","lum-customer-c_127755f5-zone-us_zone-ip-199.244.60.208","lum-customer-c_127755f5-zone-us_zone-ip-168.151.120.29","lum-customer-c_127755f5-zone-us_zone-ip-74.85.210.146","lum-customer-c_127755f5-zone-us_zone-ip-168.151.124.31","lum-customer-c_127755f5-zone-us_zone-ip-168.151.115.29","lum-customer-c_127755f5-zone-us_zone-ip-198.240.101.9","lum-customer-c_127755f5-zone-us_zone-ip-168.151.126.198","lum-customer-c_127755f5-zone-us_zone-ip-67.213.122.166","lum-customer-c_127755f5-zone-us_zone-ip-216.19.200.134","lum-customer-c_127755f5-zone-us_zone-ip-168.151.240.143","lum-customer-c_127755f5-zone-us_zone-ip-216.19.199.1","lum-customer-c_127755f5-zone-us_zone-ip-119.13.196.121","lum-customer-c_127755f5-zone-us_zone-ip-91.92.218.14","lum-customer-c_127755f5-zone-us_zone-ip-168.151.113.199","lum-customer-c_127755f5-zone-us_zone-ip-168.151.131.157","lum-customer-c_127755f5-zone-us_zone-ip-46.232.209.111","lum-customer-c_127755f5-zone-us_zone-ip-119.13.217.58","lum-customer-c_127755f5-zone-us_zone-ip-119.13.193.211","lum-customer-c_127755f5-zone-us_zone-ip-180.149.17.222","lum-customer-c_127755f5-zone-us_zone-ip-203.78.175.112","lum-customer-c_127755f5-zone-us_zone-ip-188.119.117.166","lum-customer-c_127755f5-zone-us_zone-ip-119.12.184.31","lum-customer-c_127755f5-zone-us_zone-ip-94.176.59.134","lum-customer-c_127755f5-zone-us_zone-ip-119.13.200.246","lum-customer-c_127755f5-zone-us_zone-ip-119.12.203.25","lum-customer-c_127755f5-zone-us_zone-ip-180.149.25.80","lum-customer-c_127755f5-zone-us_zone-ip-188.211.24.139","lum-customer-c_127755f5-zone-us_zone-ip-91.192.215.74","lum-customer-c_127755f5-zone-us_zone-ip-119.12.202.151","lum-customer-c_127755f5-zone-us_zone-ip-180.149.6.40","lum-customer-c_127755f5-zone-us_zone-ip-180.149.26.196","lum-customer-c_127755f5-zone-us_zone-ip-91.92.217.112","lum-customer-c_127755f5-zone-us_zone-ip-193.200.104.140","lum-customer-c_127755f5-zone-us_zone-ip-94.176.54.25","lum-customer-c_127755f5-zone-us_zone-ip-180.149.2.89","lum-customer-c_127755f5-zone-us_zone-ip-206.204.38.62","lum-customer-c_127755f5-zone-us_zone-ip-213.188.83.143","lum-customer-c_127755f5-zone-us_zone-ip-94.176.53.18","lum-customer-c_127755f5-zone-us_zone-ip-91.245.235.235","lum-customer-c_127755f5-zone-us_zone-ip-94.176.85.212","lum-customer-c_127755f5-zone-us_zone-ip-78.138.40.246","lum-customer-c_127755f5-zone-us_zone-ip-185.246.173.58","lum-customer-c_127755f5-zone-us_zone-ip-119.12.198.227","lum-customer-c_127755f5-zone-us_zone-ip-94.176.57.142","lum-customer-c_127755f5-zone-us_zone-ip-94.176.60.18","lum-customer-c_127755f5-zone-us_zone-ip-89.38.132.172","lum-customer-c_127755f5-zone-us_zone-ip-213.188.76.146","lum-customer-c_127755f5-zone-us_zone-ip-119.12.182.89","lum-customer-c_127755f5-zone-us_zone-ip-94.176.51.212","lum-customer-c_127755f5-zone-us_zone-ip-208.86.196.158","lum-customer-c_127755f5-zone-us_zone-ip-168.151.179.15","lum-customer-c_127755f5-zone-us_zone-ip-119.12.188.46","lum-customer-c_127755f5-zone-us_zone-ip-213.188.75.89","lum-customer-c_127755f5-zone-us_zone-ip-119.12.190.164","lum-customer-c_127755f5-zone-us_zone-ip-213.188.68.39","lum-customer-c_127755f5-zone-us_zone-ip-89.40.81.29","lum-customer-c_127755f5-zone-us_zone-ip-185.223.56.108","lum-customer-c_127755f5-zone-us_zone-ip-161.129.160.90","lum-customer-c_127755f5-zone-us_zone-ip-213.188.88.130","lum-customer-c_127755f5-zone-us_zone-ip-152.39.153.185","lum-customer-c_127755f5-zone-us_zone-ip-119.13.205.38","lum-customer-c_127755f5-zone-us_zone-ip-203.78.174.235","lum-customer-c_127755f5-zone-us_zone-ip-206.204.49.106","lum-customer-c_127755f5-zone-us_zone-ip-94.176.61.225","lum-customer-c_127755f5-zone-us_zone-ip-180.149.23.229","lum-customer-c_127755f5-zone-us_zone-ip-95.215.38.210","lum-customer-c_127755f5-zone-us_zone-ip-209.95.161.115","lum-customer-c_127755f5-zone-us_zone-ip-162.43.236.118","lum-customer-c_127755f5-zone-us_zone-ip-45.142.97.199","lum-customer-c_127755f5-zone-us_zone-ip-152.39.160.7","lum-customer-c_127755f5-zone-us_zone-ip-119.13.221.75","lum-customer-c_127755f5-zone-us_zone-ip-139.5.105.86","lum-customer-c_127755f5-zone-us_zone-ip-180.149.0.122","lum-customer-c_127755f5-zone-us_zone-ip-206.204.5.224","lum-customer-c_127755f5-zone-us_zone-ip-203.109.62.114","lum-customer-c_127755f5-zone-us_zone-ip-139.5.107.122","lum-customer-c_127755f5-zone-us_zone-ip-84.39.228.157","lum-customer-c_127755f5-zone-us_zone-ip-203.109.63.54","lum-customer-c_127755f5-zone-us_zone-ip-95.215.37.169","lum-customer-c_127755f5-zone-us_zone-ip-162.43.235.85","lum-customer-c_127755f5-zone-us_zone-ip-152.39.214.9","lum-customer-c_127755f5-zone-us_zone-ip-110.238.215.203","lum-customer-c_127755f5-zone-us_zone-ip-162.43.229.39","lum-customer-c_127755f5-zone-us_zone-ip-185.10.4.83","lum-customer-c_127755f5-zone-us_zone-ip-216.194.92.180"]

            port = '22225'
            rand_ips = "zproxy.lum-superproxy.io"
            rndusername = random.choice(proxy_iplum)
            usern_passw = rndusername + ':dngrv4oofa9a'
            proxy1 = {'https': "https://{}@{}:{}/".format(usern_passw, rand_ips, port)}

            response = requests.get(pdp_url, headers=headers, proxies=proxy)
            #general.write_file('deich.html', response.text, 'a', encoding='UTF-8')
            if response.status_code == 200:
                break
        details_dict = {
            'SKU_ID': sku_id,
            'Website': website_name,
            'Country': country,
            'RPC': '-',
            'MPC': '-',
            'Product_ID': '-',
            'Product_URL': pdp_url,
            'Product_Name': '-',
            'Category_Path': '-',
            'Specification': '-',
            'Description': '-',
            'Currency': '-',
            'List_Price': '-',
            'Promo_Price': '-',
            'Discount': '-',
            'Brand': '-',
            'Rating_Count': '-',
            'Review_Count': '-',
            'Image_URLs': '-',
            'Variant': '-',
            'Variant_ID': '-',
            'Colour_of_Variant': '-',
            'Colour_Grouping': '-',
            'Seller_Name': '-',
            'Stock_Count': '-',
            'Stock_Condition': '-',
            'Stock_Message': '-',
            'Sustainability_Badge': '-',
            'Reason_Code': '-',
            'Crawling_TimeStamp': datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ"),
            'Cache_Page_Link': '-',
            'Extra1': '-',
            'Extra2': '-',
            'Extra3': '-',
            'Extra4': '-',
            'Extra5': '-'
        }
        if response.status_code == 200:

            response123 = HtmlResponse(pdp_url, body=response.content)
            tree = html.fromstring(response.content)
            soup = BeautifulSoup(response.content, 'lxml')

            details_dict['SKU_ID'] = sku_id

            product_id = tree.xpath(
                "//m-product-details[@id='product-details']/section/details/dl/dt/span[contains(@i18ntext,'Artikelnummer')]//following::dd[1]/text()")

            details_dict['Product_ID'] = product_id[0]

            details_dict['RPC'] = product_id[0]

            mpc = tree.xpath('//*[contains(text(),"Artikelnummer: ")]/../following-sibling::dd/text()')
            if len(mpc) > 0:
                details_dict['MPC'] = mpc[-1]
            else:
                details_dict['MPC'] = 'Not Available'

            details_dict['Product_URL'] = pdp_url
            details_dict['Product_Name'] = general.clean(tree.xpath('//div[@class="product"]//h1[@class="product-name"]/text()')[0])
            details_dict['Category_Path'] = 'Not Available'

            f_specs = list(
                zip(tree.xpath('//m-product-details[@id="product-details"]/section/details/dl/dt/span/text()'),
                    tree.xpath('//m-product-details[@id="product-details"]/section/details/dl/dd/text()')))

            s_specs = list(zip(tree.xpath(
                '//m-product-details[@id="product-details"]/section/details/dl/m-product-classification/dt/text()'),
                tree.xpath(
                    '//m-product-details[@id="product-details"]/section/details/dl/m-product-classification/dd/text()')))

            all_specs = '|'.join(i[0] + i[1] for i in f_specs) + '|' + '|'.join(i[0] + i[1] for i in s_specs)
            details_dict['Specification'] = general.clean(all_specs)

            try:
                disc_json = json.loads(json.dumps(tree.xpath('//script[@id="product-schema"]/text()')[0]))
                jsd = json.loads(disc_json)
                try:
                    data = jsd['description']
                    xyz = str(data).split('<li>')
                    desc1 = '|'.join(xyz).replace('</li>', '').replace('\n', '').replace('\t', '').replace('</ul>',
                                                                                                           '').strip().replace(
                        '<ul>', '').strip().strip('|').strip()
                    details_dict['Description'] = general.clean(desc1)
                except Exception as e:
                    details_dict['Description'] = '-'
            except:
                details_dict['Description'] = '-'

            prices = tree.xpath('//div[@class="order-lane"]/section/m-product-price/strong/p/del/text()')
            if len(prices) > 0:
                details_dict['List_Price'] = \
                    tree.xpath('//div[@class="order-lane"]/section/m-product-price/strong/p[1]/text()')[
                        0].replace('€', '')
                # details_dict['Promo_Price'] =prices
                details_dict['Promo_Price'] = prices[0].replace('€', '')

                details_dict['Discount'] = tree.xpath('//section[@class="interferer"]/p/text()')[
                    0].replace('-', '')
            else:
                pric = tree.xpath('//div[@class="order-lane"]/section/m-product-price/strong/p[1]/text()')[
                        0].replace('€', '')
                details_dict['List_Price'] = pric.strip()
                details_dict['Promo_Price'] = pric.strip()
                details_dict['Discount'] = "-"
            details_dict['Currency'] = 'EURO'

            try:
                details_dict['Brand'] = tree.xpath('//div[@class="product"]/div[@class="order-lane"]/section/strong[@class="brand"]/a/text()')[0]
                details_dict['Brand']=general.clean(details_dict['Brand'])
            except:
                try:
                    details_dict['Brand'] = tree.xpath('//strong[@class="brand"]/a/text()')[0]
                    details_dict['Brand']=general.clean(details_dict['Brand'])
                except:
                    details_dict['Brand'] = '-'

            try:
                details_dict["Rating_Count"] = jsd["aggregateRating"]["ratingValue"]
                details_dict["Review_Count"] = jsd["aggregateRating"]["reviewCount"]
            except:
                details_dict["Rating_Count"] = '-'
                details_dict["Review_Count"] = '-'
                # print('error..')
            try:
                images = tree.xpath('//div[@class="product-slider"]/div/div/div/img/@src')
                details_dict['Image_URLs'] = '|'.join(i for i in images)
            except:
                details_dict['Image_URLs'] = '-'

            try:
                details_dict['Colour_of_Variant'] = jsd["color"]
                details_dict['Colour_of_Variant']= general.clean(details_dict['Colour_of_Variant'])
            except:
                details_dict['Colour_of_Variant'] = '-'
            try:
                colors = jsd["isSimilarTo"]
                # onlycolors=[i.split(" ")[-1] for i in colors]
                # print(colors)
                details_dict['Colour_Grouping'] = '|'.join(
                    ele for ele in [i['name'].split(' ')[-1] for i in colors])
                details_dict['Colour_Grouping'] = general.clean(details_dict['Colour_Grouping'])
            except:
                details_dict['Colour_Grouping'] = '-'
            sname = tree.xpath('//div[@class="partner-hint-container"]/section/a/span/text()')
            if sname:
                sname = sname[0]
            else:
                sname = ''


            details_dict['Seller_Name'] = sname
            details_dict['Stock_Count'] = 'Not Available'
            # details_dict['Stock_Condition'] = "jsd['availability']"
            details_dict['Sustainability_Badge'] = 'Not Available'
            details_dict['Reason_Code'] = 'Success-PF'
            details_dict['Stock_Message'] = 'Not Available'
            details_dict['Extra1'] = '-'
            details_dict['Extra2'] = '-'
            details_dict['Extra3'] = '-'
            details_dict['Extra4'] = '-'
            details_dict['Extra5'] = '-'
            try:
                if 'xlink:href="#icon-sustainable"></use>' in str(response.text):
                    sus = 'Yes'
                else:
                    sus = 'No'
            except Exception as e:
                sus = 'No'
            details_dict['Sustainability_Badge'] = sus
            now = datetime.now()
            date_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")
            details_dict['Crawling_TimeStamp'] = date_time
            date_time1 = str(now.strftime("%m%d%Y"))

            try:
                datazone = datetime.now()
                f_date = datazone.strftime("%d_%m_%Y")
                # f_date = '03-04_2022'
                cpid = str(sku_id) + '_' + str(f_date)
                ASS_folder = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Deichmann_DE\\PDP"
                sos_date_wise_folder = ASS_folder + f"\\{str(f_date)}"
                if os.path.exists(sos_date_wise_folder):
                    pass
                else:
                    os.mkdir(sos_date_wise_folder)
                sos_filename = sos_date_wise_folder + "\\" + cpid + ".html"
                sos_filename = sos_filename.replace("+", "_").replace("-", "_")
                page_path = sos_filename
                page_path = page_path.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\',
                                              'https:////ecxus440.eclerx.com//cachepages//').replace('\\',
                                                                                                     '//').replace(
                    '//', '/')
                # cptext = response.text + response1.text
                cptext = response.text
                print(page_path)
                if os.path.exists(sos_filename):
                    # with open(sos_filename, 'wb', encoding='utf-8') as f:
                    with open(sos_filename, 'wb') as f:
                        f.write(response123.body)
                else:
                    with open(sos_filename, 'wb') as f:
                        # f.write(response.text)
                        f.write(response123.body)
            except Exception as e:
                page_location = ''
                print(e)
            details_dict['Cache_Page_Link'] = page_path
            all_uk_size = general.xpath(tree, '//ul[@class="list-grid"]//li/div/text()', mode='set')

            chksize = general.xpath(tree, '//ul[@class="size-system"]/li', mode='set_tc')
            if 'Herstellergrößen|EU|UK' in chksize:
                value1 = response.text
                variantraw = response.text.split('[/')[4]
                variantraw = variantraw.replace('&q;','"')
                varraw1 = general.midtext(variantraw,'"variantMatrix',',"variantType"')
                value1 = value1.replace('&q;', '"')
                varraw1 = general.midtext(value1, 'mosaicproduct', '[/')
                varraw2 = general.midtext(value1, '"variantMatrix":[{"articleNumber', ',"variantType"')
                if varraw2:
                    varraw3 = '{"variantMatrix":[{"articleNumber' + varraw2 + '}'
                    varjson = json.loads(varraw3)
                    vmatrix = varjson.get("variantMatrix")[0]['elements']
                    for varid in vmatrix:
                        vname = varid.get("sizeMap")[0]["name"]
                        locationsize = varid.get("sizeMap")[0]["sizeSystem"]["name"]
                        try:
                            pp = varid.get('variantOption')['priceData']['value']
                        except:
                            pp=''

                        try:
                            lp = varid.get('variantOption')['allPrices'][0]['value']
                        except:
                            lp=''
                        # if pp:
                        #     details_dict['Promo_Price'] = pp
                        #     details_dict['List_Price'] = lp
                        details_dict['Seller_Name'] = 'Adidas'
                        if 'eu' in locationsize.lower():
                            details_dict['Seller_Name'] = '-'
                        details_dict['Website'] = 'DEICHMANN'
                        details_dict['Country'] = 'DE'
                        details_dict['Variant'] = vname
                        stockurl = varid.get('variantOption')['url']
                        ppid = stockurl.split('/')[-1]
                        startsarch = '],"productCode":"' + ppid
                        stokraw = general.midtext(value1,startsarch,'},{')
                        stock = general.midtext(stokraw,'stockLevelStatus":"','"')
                        if stock == 'outOfStock':
                            stock = 'Out Of Stock'
                        else:
                            stock = 'InStock'
                        details_dict['Stock_Condition'] = stock
                        details_dict['Variant_ID'] = '"' + varid.get("sku") +'"'
                        details_dict['Extra5'] = locationsize
                        all_product_data.append(details_dict)
                        store_data_in_csv(dict=[details_dict])
            elif 'EU|UK' in chksize:
                value1 = response.text
                variantraw = response.text.split('[/')[4]
                variantraw = variantraw.replace('&q;','"')
                varraw1 = general.midtext(variantraw,'"variantMatrix',',"variantType"')
                value1 = value1.replace('&q;', '"')
                varraw1 = general.midtext(value1, 'mosaicproduct', '[/')
                varraw2 = general.midtext(value1, '"variantMatrix":[{"articleNumber', ',"variantType"')
                if varraw2:
                    varraw3 = '{"variantMatrix":[{"articleNumber' + varraw2 + '}'
                    varjson = json.loads(varraw3)
                    vmatrix = varjson.get("variantMatrix")[0]['elements']
                    for varid in vmatrix:
                        vname = varid.get("sizeMap")[0]["name"]
                        locationsize = varid.get("sizeMap")[0]["sizeSystem"]["name"]
                        try:
                            pp = varid.get('variantOption')['priceData']['value']
                        except:
                            pp=''

                        try:
                            lp = varid.get('variantOption')['allPrices'][0]['value']
                        except:
                            lp=''
                        # if pp:
                        #     details_dict['Promo_Price'] = pp
                        #     details_dict['List_Price'] = lp

                        details_dict['Website'] = 'DEICHMANN'
                        details_dict['Country'] = 'DE'
                        details_dict['Variant'] = vname
                        stockurl = varid.get('variantOption')['url']
                        ppid = stockurl.split('/')[-1]
                        startsarch = '],"productCode":"' + ppid
                        stokraw = general.midtext(value1,startsarch,'},{')
                        stock = general.midtext(stokraw,'stockLevelStatus":"','"')
                        if stock == 'outOfStock':
                            stock = 'Out Of Stock'
                        else:
                            stock = 'InStock'
                        details_dict['Stock_Condition'] = stock
                        details_dict['Variant_ID'] = '"' + varid.get("sku") +'"'
                        details_dict['Extra5'] = locationsize
                        all_product_data.append(details_dict)
                        store_data_in_csv(dict=[details_dict])

            # try:
            else:
                chkv = 0
                value1 = response.text
                value1 = value1.replace('&q;', '"')
                varraw2 = general.midtext(value1, '"variantMatrix":[{"classifications', ',"variantType"')
                if varraw2 =='':
                    chkv = 1
                    varraw2 = general.midtext(value1, '"variantMatrix":[{"articleNumber', ',"variantType"')
                if varraw2:
                    if chkv ==1:
                        varraw3 = '{"variantMatrix":[{"articleNumber' + varraw2 + '}'
                    else:
                        varraw3 = '{"variantMatrix":[{"classifications' + varraw2 + '}'
                    varjson = json.loads(varraw3)
                    vmatrix = varjson.get("variantMatrix")[0]['elements']
                    for varid in vmatrix:
                        vname = varid.get("sizeMap")[0]["name"]
                        locationsize = varid.get("sizeMap")[0]["sizeSystem"]["name"]
                        try:
                            pp = varid.get('variantOption')['priceData']['value']
                        except:
                            pp = ''

                        try:
                            lp = varid.get('variantOption')['allPrices'][0]['value']
                        except:
                            lp = ''
                        # if pp:
                        #     details_dict['Promo_Price'] = pp
                        #     details_dict['List_Price'] = lp

                        details_dict['Website'] = 'DEICHMANN'
                        details_dict['Country'] = 'DE'
                        details_dict['Variant'] = vname
                        stockurl = varid.get('variantOption')['url']
                        ppid = stockurl.split('/')[-1]
                        startsarch = '],"productCode":"' + ppid
                        stokraw = general.midtext(value1, startsarch, '},{')
                        stock = general.midtext(stokraw, 'stockLevelStatus":"', '"')
                        if stock == 'outOfStock':
                            stock = 'Out Of Stock'
                        else:
                            stock = 'InStock'
                        details_dict['Stock_Condition'] = stock
                        details_dict['Variant_ID'] = '"' + varid.get("sku") + '"'
                        details_dict['Extra5'] = locationsize
                        all_product_data.append(details_dict)
                        store_data_in_csv(dict=[details_dict])
                else:
                    details_dict['Variant'] = '-'
                    details_dict['Stock_Condition'] = 'In Stock'
                    details_dict['Variant_ID'] = '-'
                    all_product_data.append(details_dict)
                    store_data_in_csv(dict=[details_dict])
                    print(details_dict)

            #     last_avail_sizes1 = []
            #     varsize = general.xpath(tree,'//div[@class="content"]/nav/div/ul[2]/li',mode='set')
            #     aa = tree.xpath('//div[@class="content"]/nav/div/ul[2]/li/text()')
            #     for w in varsize:
            #         var = general.xpath(w,'.//span',mode='tc')
            #         # var = var.replace('Versand durch adidas','').replace('Leider ausverkauft','')
            #         print(var)
            #         last_avail_sizes1.append(var)
            #
            #     availabele_sizes = tree.xpath('//div[@class="content"]/nav/div/ul[2]/li/span/text()',encodings='UTF-16')
            #     last_avail_sizes = []
            #     for a in availabele_sizes:
            #         if '€' in a:
            #             pass
            #         else:
            #             last_avail_sizes.append(a)
            #
            #     if 'Herstellergrößen|EU|UK' in chksize:
            #         last_avail_sizes = last_avail_sizes1
            #     elif 'EU|UK' in chksize:
            #         last_avail_sizes = []
            #         varsize = general.xpath(tree, '//div[@class="content"]/nav/div/ul[2]/li[@class="active-element"]', mode='set')
            #         aa = tree.xpath('//div[@class="content"]/nav/div/ul[2]/li/text()')
            #         for w in varsize:
            #             var = general.xpath(w, './/span', mode='tc')
            #             # var = var.replace('Versand durch adidas','').replace('Leider ausverkauft','')
            #             print(var)
            #             last_avail_sizes.append(var)
            #     else:
            #         last_avail_sizes = []
            #         varsize = general.xpath(tree, '//div[@class="content"]/nav/div/ul[2]/li[@class="active-element"]', mode='set')
            #         aa = tree.xpath('//div[@class="content"]/nav/div/ul[2]/li/text()')
            #         for w in varsize:
            #             var = general.xpath(w, './/span', mode='tc')
            #             # var = var.replace('Versand durch adidas','').replace('Leider ausverkauft','')
            #             print(var)
            #             last_avail_sizes.append(var)
            #
            #     for item in last_avail_sizes:
            #         details_dict['Website'] = 'DEICHMANN'
            #         details_dict['Country'] = 'DE'
            #
            #         details_dict['Variant']=general.clean(details_dict['Variant'])
            #         if 'Leider ausverkauft' in item:
            #             stock = 'Out Of Stock'
            #         else:
            #             stock = 'In Stock'
            #         details_dict['Stock_Condition'] = stock
            #         details_dict['Variant'] = item.replace('Versand durch adidas','').replace('Leider ausverkauft','').strip()
            #         details_dict['Variant_ID'] = '-'
            #         all_product_data.append(details_dict)
            #         store_data_in_csv(dict=[details_dict])

            # except Exception as e:
            #     print("available size not found")


            # try:
            #     not_availabele_sizes = tree.xpath(
            #         '//div[@class="content"]/nav/div/ul[2]/li/div/section/a/span/text()')#,encodings='UTF-16'
            #
            #     for item in not_availabele_sizes:
            #         details_dict['Variant'] = item.strip()
            #         details_dict['Variant']=general.clean(details_dict['Variant'])
            #         details_dict['Stock_Condition'] = 'Out of Stock'
            #         details_dict['Variant_ID'] = '-'
            #         details_dict['Seller_Name'] = 'Adidas'
            #         all_product_data.append(details_dict)
            #         store_data_in_csv(dict=[details_dict])
            #         print(details_dict)
            #
            # except:
            #     print('size not found')

            # if not_availabele_sizes ==[] and availabele_sizes ==[]:
            # if availabele_sizes == []:
            #     details_dict['Variant'] = '-'
            #     # details_dict['Variant'] = general.clean(details_dict['Variant'])
            #     details_dict['Stock_Condition'] = 'In Stock'
            #     details_dict['Variant_ID'] = '-'
            #     all_product_data.append(details_dict)
            #     store_data_in_csv(dict=[details_dict])
            #     print(details_dict)
        else:
            print('Something went wrong...')
            # all_product_data.append(details_dict)
            # details_dict['Reason_Code'] = 'Success-PNF'
            # store_data_in_csv(dict=[details_dict])

    except:
        print('Request Failed...')
        print(pdp_url)
        details_dict = {
            'SKU_ID': sku_id,
            'Website': website_name,
            'Country': country,
            'RPC': '-',
            'MPC': '-',
            'Product_ID': '-',
            'Product_URL': pdp_url,
            'Product_Name': '-',
            'Category_Path': '-',
            'Specification': '-',
            'Description': '-',
            'Currency': '-',
            'List_Price': '-',
            'Promo_Price': '-',
            'Discount': '-',
            'Brand': '-',
            'Rating_Count': '-',
            'Review_Count': '-',
            'Image_URLs': '-',
            'Variant': '-',
            'Variant_ID': '-',
            'Colour_of_Variant': '-',
            'Colour_Grouping': '-',
            'Seller_Name': '-',
            'Stock_Count': '-',
            'Stock_Condition': '-',
            'Stock_Message': '-',
            'Sustainability_Badge': '-',
            'Reason_Code': 'Success-PNF',
            'Crawling_TimeStamp': datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ"),
            'Cache_Page_Link': '-',
            'Extra1': '-',
            'Extra2': '-',
            'Extra3': '-',
            'Extra4': '-',
            'Extra5': '-'
        }
        all_product_data.append(details_dict)
        store_data_in_csv(dict=[details_dict])
        #store_data_in_csv(dict=[details_dict])

