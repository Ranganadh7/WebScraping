import random
from lxml import html
import time
import general
import xlrd
import requests
import os
from lxml import html
import time
import json
from datetime import datetime, timedelta
import datetime
import openpyxl
import datetime
import pandas as pd
from datetime import datetime

headers ={
# 'Referer': 'https://en.aboutyou.de/',
'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="96", "Google Chrome";v="96"',
'sec-ch-ua-mobile': '?0',
'sec-ch-ua-platform': "Windows",
'sec-fetch-dest': 'empty',
'sec-fetch-mode': 'cors',
'sec-fetch-site': 'cross-site',
'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36',
'x-grpc-web': '1'}

from csv import DictWriter
def store_data_by_product_id(dict):
    df = pd.DataFrame.from_dict(dict)
    print(df)
    # df.to_csv("output-1.csv", mode="a+", index=False, header=False)
    df.to_csv("output-1_0104.csv", encoding="utf-8-sig",mode="a+", index=False, header=False)
    """
    fieldnames = ['INPUT_PLATFORMID', 'INPUT_ADIDASID', 'Website Name', 'Country','Product_ID','Product_URL','Product_Name','Category_Path','Specification','Description','Currency','List_Price','Promo_Price','Discount','Brand','Rating_Count','Review_Count','Image_URLs','Variant','Variant_ID','Colour_of_Variant','Colour_Grouping','Seller_Name','Stock_Count','Stock_Condition','Stock_Message','Sustainability_Badge','Reason_Code','Crawling_TimeStamp','Cache_Page_Link','Extra1','Extra2','Extra3','Extra4','Extra5']
    with open('ABOUTYOU_02.csv', 'a+', encoding='UTF-16', newline='') as file:
        writer = DictWriter(file, fieldnames=fieldnames)
        # writer.writeheader()
        writer.writerows(dict)
        file.close()
    """
file_path=r'input-1.xlsx'
# file_path='E:\\Ranganadh.S\\About-You\\pdp-update\\today-task\\input-rehit.xlsx'
def extract_xl():
    workbook= openpyxl.load_workbook(file_path)
    sheet_obj= workbook.active
    all_rows=sheet_obj.max_row

    inputdata=[]
    for i in range(1,all_rows):
        limit=sheet_obj.cell(row=i+1,column=2)
        urls= sheet_obj.cell(row=i+1,column=1)

        u_data=[urls.value,limit.value]

        inputdata.append(u_data)
    print(inputdata)
    return inputdata



def by_product_id(prod_id_data):
    # product_id_url = 'https://www.aboutyou.de/p/puma/sportschuh-cali-sport-soft-wns-{}?is_s=none&is_h=sl'.format(prod_id_data[0])
    try:
        product_id_url=str(prod_id_data[0])
        print(product_id_url)
        sku = str(prod_id_data[1])
        for _ in range(10):
            p1 = ['154.28.67.106', '154.28.67.111', '154.28.67.116', '154.28.67.117', '154.28.67.125', '154.28.67.131',
                  '154.28.67.133', '154.28.67.142', '154.28.67.156', '154.28.67.163', '154.28.67.173', '154.28.67.18',
                  '154.28.67.182', '154.28.67.184', '154.28.67.20', '154.28.67.200', '154.28.67.210', '154.28.67.218',
                  '154.28.67.222', '154.28.67.223', '154.28.67.231', '154.28.67.240', '154.28.67.243', '154.28.67.253',
                  '154.28.67.39', '154.28.67.4', '154.28.67.49', '154.28.67.5', '154.28.67.61', '154.28.67.80',
                  '154.28.67.81', '154.28.67.87', '154.28.67.88', '154.28.67.96', '154.28.67.99', '154.7.230.100',
                  '154.7.230.101', '154.7.230.103', '154.7.230.107', '154.7.230.109', '154.7.230.130', '154.7.230.132',
                  '154.7.230.14', '154.7.230.140', '154.7.230.147', '154.7.230.151', '154.7.230.156', '154.7.230.163',
                  '154.7.230.170', '154.7.230.18', '154.7.230.183', '154.7.230.188', '154.7.230.189', '154.7.230.19',
                  '154.7.230.190', '154.7.230.198', '154.7.230.204', '154.7.230.209', '154.7.230.235', '154.7.230.238',
                  '154.7.230.246', '154.7.230.29', '154.7.230.41', '154.7.230.42', '154.7.230.51', '154.7.230.55',
                  '154.7.230.60', '154.7.230.61', '154.7.230.74', '154.7.230.82', '154.7.230.89', '23.131.8.112',
                  '23.131.8.115', '23.131.8.117', '23.131.8.12', '23.131.8.121', '23.131.8.124', '23.131.8.150',
                  '23.131.8.161', '23.131.8.166', '23.131.8.171', '23.131.8.173', '23.131.8.176', '23.131.8.177',
                  '23.131.8.181', '23.131.8.19', '23.131.8.192', '23.131.8.194', '23.131.8.199', '23.131.8.202',
                  '23.131.8.203', '23.131.8.204', '23.131.8.207', '23.131.8.209', '23.131.8.213', '23.131.8.216',
                  '23.131.8.225', '23.131.8.228', '23.131.8.231', '23.131.8.238', '23.131.8.254', '23.131.8.36',
                  '23.131.8.5', '23.131.8.76', '23.131.8.93', '23.131.8.95', '23.131.8.99', '23.131.88.105',
                  '23.131.88.12', '23.131.88.137', '23.131.88.139', '23.131.88.140', '23.131.88.145', '23.131.88.150',
                  '23.131.88.151', '23.131.88.153', '23.131.88.154', '23.131.88.156', '23.131.88.165', '23.131.88.18',
                  '23.131.88.191', '23.131.88.192', '23.131.88.194', '23.131.88.198', '23.131.88.202', '23.131.88.206',
                  '23.131.88.220', '23.131.88.223', '23.131.88.228', '23.131.88.233', '23.131.88.24', '23.131.88.242',
                  '23.131.88.244', '23.131.88.47', '23.131.88.63', '23.131.88.67', '23.131.88.73', '23.131.88.80',
                  '23.131.88.81', '23.131.88.82', '23.131.88.88', '23.131.88.97', '23.170.144.149', '23.170.144.209',
                  '23.170.144.212', '23.170.144.242', '23.170.144.83', '23.170.145.117', '23.170.145.167',
                  '23.170.145.182', '23.170.145.19', '23.170.145.203', '23.226.17.101', '23.226.17.109',
                  '23.226.17.112', '23.226.17.113', '23.226.17.115', '23.226.17.123', '23.226.17.129', '23.226.17.143',
                  '23.226.17.148', '23.226.17.165', '23.226.17.186', '23.226.17.199', '23.226.17.201', '23.226.17.207',
                  '23.226.17.210', '23.226.17.219', '23.226.17.220', '23.226.17.222', '23.226.17.229', '23.226.17.250',
                  '23.226.17.254', '23.226.17.26', '23.226.17.33', '23.226.17.4', '23.226.17.49', '23.226.17.5',
                  '23.226.17.55', '23.226.17.66', '23.226.17.7', '23.226.17.72', '23.226.17.78', '23.226.17.8',
                  '23.226.17.86', '23.226.17.90', '23.226.17.93', '23.230.177.105', '23.230.177.110', '23.230.177.113',
                  '23.230.177.121', '23.230.177.130', '23.230.177.14', '23.230.177.143', '23.230.177.15',
                  '23.230.177.150', '23.230.177.154', '23.230.177.165', '23.230.177.173', '23.230.177.191',
                  '23.230.177.196', '23.230.177.203', '23.230.177.206', '23.230.177.208', '23.230.177.217',
                  '23.230.177.220', '23.230.177.221', '23.230.177.224', '23.230.177.228', '23.230.177.231',
                  '23.230.177.235', '23.230.177.237', '23.230.177.241', '23.230.177.27', '23.230.177.38',
                  '23.230.177.52', '23.230.177.61', '23.230.177.67', '23.230.177.72', '23.230.177.80', '23.230.177.88',
                  '23.230.177.94', '23.230.177.99', '23.230.197.103', '23.230.197.106', '23.230.197.109',
                  '23.230.197.11', '23.230.197.12', '23.230.197.122', '23.230.197.124', '23.230.197.146',
                  '23.230.197.155', '23.230.197.156', '23.230.197.174', '23.230.197.179', '23.230.197.181',
                  '23.230.197.196', '23.230.197.2', '23.230.197.201', '23.230.197.207', '23.230.197.208',
                  '23.230.197.225', '23.230.197.227', '23.230.197.233', '23.230.197.236', '23.230.197.239',
                  '23.230.197.240', '23.230.197.244', '23.230.197.251', '23.230.197.50', '23.230.197.52',
                  '23.230.197.54', '23.230.197.60', '23.230.197.71', '23.230.197.80', '23.230.197.81', '23.230.197.84',
                  '23.230.197.97', '23.230.74.102', '23.230.74.110', '23.230.74.116', '23.230.74.125', '23.230.74.133',
                  '23.230.74.135', '23.230.74.14', '23.230.74.141', '23.230.74.149', '23.230.74.15', '23.230.74.157',
                  '23.230.74.16', '23.230.74.170', '23.230.74.172', '23.230.74.174', '23.230.74.183', '23.230.74.187',
                  '23.230.74.19', '23.230.74.198', '23.230.74.208', '23.230.74.212', '23.230.74.215', '23.230.74.23',
                  '23.230.74.230', '23.230.74.231', '23.230.74.252', '23.230.74.30', '23.230.74.41', '23.230.74.57',
                  '23.230.74.58', '23.230.74.59', '23.230.74.6', '23.230.74.75', '23.230.74.81', '23.230.74.88',
                  '23.230.74.91', '23.27.222.108', '23.27.222.109', '23.27.222.134', '23.27.222.138', '23.27.222.159',
                  '23.27.222.161', '23.27.222.164', '23.27.222.166', '23.27.222.178', '23.27.222.19', '23.27.222.195',
                  '23.27.222.201', '23.27.222.202', '23.27.222.203', '23.27.222.208', '23.27.222.21', '23.27.222.211',
                  '23.27.222.218', '23.27.222.223', '23.27.222.228', '23.27.222.234', '23.27.222.236', '23.27.222.242',
                  '23.27.222.251', '23.27.222.253', '23.27.222.34', '23.27.222.61', '23.27.222.62', '23.27.222.69',
                  '23.27.222.70', '23.27.222.72', '23.27.222.73', '23.27.222.74', '23.27.222.81', '23.27.222.93',
                  '38.131.131.110', '38.131.131.114', '38.131.131.123', '38.131.131.125', '38.131.131.137',
                  '38.131.131.142', '38.131.131.145', '38.131.131.147', '38.131.131.15', '38.131.131.154',
                  '38.131.131.16', '38.131.131.17', '38.131.131.173', '38.131.131.18', '38.131.131.193',
                  '38.131.131.204', '38.131.131.207', '38.131.131.227', '38.131.131.229', '38.131.131.233',
                  '38.131.131.238', '38.131.131.246', '38.131.131.248', '38.131.131.250', '38.131.131.31',
                  '38.131.131.36', '38.131.131.50', '38.131.131.58', '38.131.131.64', '38.131.131.70', '38.131.131.71',
                  '38.131.131.74', '38.131.131.83', '38.131.131.94', '38.131.131.99', '38.75.75.104', '38.75.75.111',
                  '38.75.75.112', '38.75.75.119', '38.75.75.120', '38.75.75.123', '38.75.75.127', '38.75.75.139',
                  '38.75.75.14', '38.75.75.143', '38.75.75.155', '38.75.75.156', '38.75.75.158', '38.75.75.170',
                  '38.75.75.179', '38.75.75.188', '38.75.75.2', '38.75.75.201', '38.75.75.231', '38.75.75.232',
                  '38.75.75.241', '38.75.75.246', '38.75.75.251', '38.75.75.26', '38.75.75.29', '38.75.75.4',
                  '38.75.75.44', '38.75.75.49', '38.75.75.56', '38.75.75.58', '38.75.75.62', '38.75.75.72',
                  '38.75.75.76', '38.75.75.79', '38.75.75.88', '38.96.156.108', '38.96.156.112', '38.96.156.128',
                  '38.96.156.131', '38.96.156.14', '38.96.156.142', '38.96.156.143', '38.96.156.149', '38.96.156.16',
                  '38.96.156.163', '38.96.156.165', '38.96.156.169', '38.96.156.186', '38.96.156.188', '38.96.156.190',
                  '38.96.156.192', '38.96.156.194', '38.96.156.199', '38.96.156.218', '38.96.156.236', '38.96.156.240',
                  '38.96.156.252', '38.96.156.28', '38.96.156.32', '38.96.156.35', '38.96.156.56', '38.96.156.57',
                  '38.96.156.6', '38.96.156.67', '38.96.156.77', '38.96.156.80', '38.96.156.83', '38.96.156.84',
                  '38.96.156.89', '38.96.156.92', '45.238.157.100', '45.238.157.104', '45.238.157.106',
                  '45.238.157.110', '45.238.157.116', '45.238.157.118', '45.238.157.119', '45.238.157.12',
                  '45.238.157.123', '45.238.157.132', '45.238.157.14', '45.238.157.149', '45.238.157.15',
                  '45.238.157.183', '45.238.157.186', '45.238.157.189', '45.238.157.2', '45.238.157.212',
                  '45.238.157.214', '45.238.157.217', '45.238.157.22', '45.238.157.228', '45.238.157.23',
                  '45.238.157.247', '45.238.157.43', '45.238.157.48', '45.238.157.51', '45.238.157.52', '45.238.157.53',
                  '45.238.157.56', '45.238.157.61', '45.238.157.65', '45.238.157.72', '45.238.157.79', '45.238.157.8',
                  '45.238.159.103', '45.238.159.107', '45.238.159.110', '45.238.159.114', '45.238.159.116',
                  '45.238.159.123', '45.238.159.126', '45.238.159.144', '45.238.159.148', '45.238.159.15',
                  '45.238.159.156', '45.238.159.165', '45.238.159.167', '45.238.159.183', '45.238.159.20',
                  '45.238.159.208', '45.238.159.217', '45.238.159.220', '45.238.159.23', '45.238.159.230',
                  '45.238.159.235', '45.238.159.237', '45.238.159.238', '45.238.159.24', '45.238.159.249',
                  '45.238.159.251', '45.238.159.32', '45.238.159.34', '45.238.159.51', '45.238.159.6', '45.238.159.66',
                  '45.238.159.77', '45.238.159.79', '45.238.159.82', '45.238.159.91']

            p_auth = str("csimonra:h19VA2xZ")
            p_host = random.choice(p1)
            p_port = "29842"
            proxy = {
                'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
            }
            header = {
                'authority': 'www.zalando.de',
                'method': 'GET',
                'accept-language': 'de-US,en;q=0.9',
                'cookie': 'bm_sz=069A88C760C08379B172D2DBCDEBD42B~YAAQVYpRaKQopgR9AQAAvQpECQ03kVHgTXiCd1GbcrCjcxBlNHPXVX282CkSqkAH//H+cs51HE6w7uYSYAHRJ6wmuTCtQPpIzURHd7qoQJgSOyO2Dhw9mOSBLYqatbOQzWHpsQk1EeOBCTFLEyZ3RgNyb78dhH7CArjNuWC1Je+c8MqofhH2DApJK4y5VuGuoAFW4lqRBOuOCn2LqgYE4uKimrEPQL16spOf26MFeVP67junsaaP9xcklQTh7Nc85eiSHaTwllWBnMcTDQmvenp9+Fm9XaI0RRjGHTcPvQtfp/Dbi9m8Jn8WGtiCfbsWKSfyXuempeqsrtsJJWzkkMp3KrPNWyYNV/dnjwuinvx8QA0pokBt9Z8ih3/f510IY14Ibw/Ise0CupxNqapb~4274485~3753014; frsx=AAAAAOLx0CRkDYZWFa1ckXV0kpQU5BSA4wz2nUql40qRjMoj5UoHE7fQ4DMzBTasAPxf6Nvy5guPrJa8Fq0nTcTtUpz7mWWh40z-inM_pXODdyHFseZ88cPh-7BD9ttGRCTgD4FyR8nl_WvX3FrEoeE=; Zalando-Client-Id=36582f25-e91a-4d88-ab49-46566cef2192; ncx=f; ak_bmsc=FC3464107F8B6F47B865B5E3178594A1~000000000000000000000000000000~YAAQHopRaEHi6Ah9AQAA9xhECQ3YM3UsO/ilGETvpoGVURTUVMFezuyLuNsyPIeKh50n6G7dByCWT0j4/+hnscRWObt+E3cVOsuMpibnzoEuLDGZnBCzmEcmROUSBpWvFOI45KLLcz/OxXEJEpK68fp9pr0B339KMFurCHsqN8JNLyBhe2s1qhCQHhPJ/XmrojyIPbMYTYjMo/LhHnMSD2ltPlBOBxOC1leIRziqwcSD4lJ98qjphY9Z/YLGZqn9dYn+w76JklHyFBVCOOAkMvhaKosi8bC//JnrfpJviQFgDUobJyux0fItfUCxfCcCKDby8K9c/Hzd4FxzSwJPXnqDJJoSGJNRCMdJoBL9YE1CFFKyXMzdiVWKcEsdpOcTD5dQalAEGrQteJ1Q3RV9V+Iq/bmGxoS5UwDvnfqa2Ldl4DMua6Jk/xxuMqnCCtl6iTJDmyHGCzNjyD0XeDTA5MtI0g8jg9CJ64PgmGmsLqyBInfzgCGnWDlF; _gcl_au=1.1.1341373456.1636538052; _ga=GA1.2.380652769.1636538053; _gid=GA1.2.1007503823.1636538053; _gat_zalga=1; bm_sv=0D04788DC6CB39B5D18D9F039783D418~KxLmMXYkJMJEFsadY5E5tWQxEFCov92v8oT33wJK9ghOj+ZxtemU+5soAP8tRdV22HWAL81btxwGtKKrLKxMgFDH4r0Pu1+3dWr5M3hGn/sQSHYPiXXumZxA8ZR6pnBW3BLje2dUK9Sf2U9YLvn3RwGhUdl+i2eIVTDJHKN53ss=; language-preference=de; _abck=802962DA192798FEE29D6176A9193159~-1~YAAQXMYcuEXiE/98AQAAl2RGCQbMs8vomqDAzPEXShw6pYC8MeIUgJ/YTOml5IznyPFyuEaW1ZYFpjPDa0M/6Vuz9xKwmtDiQ0ku+UmqSiC07R5Shd6sMFFGfz2P3bQCl7XDJcM5+GlgMQL924nm7517pzA1eSzeBBFbUwYK2r7OreZj1Ph3iPBrgoB/jn2mOeOzB2RSemE1Gdqgyic5Iprt0sx1XsynIgiIUS6fQj3Cf/HTf1bJ6T0QmA+mGCJcgBeinMJQQ6ZZnKISziHLgyQ4ii380xtNd0H7G6r616n6x/v+7PY16Jvuq29Ile1+ftAe7VbHH+XxF3u0BLdSdm3DEUOT4eh8n3fDPp4YmcsQECh/9yW5v0VNk1UJP1zC1lovZN5GUaI8ZAZ1FR+htgv8Xb9zVnk33FoCun3gSPvVpGjAKJVwU0nwpjlarADprniC8ivMOCXrSh90dpHyAySjFnC6WvR9q6U7Jk92vP+hcMyW/wcpAK+V8ZAtcw==~-1~-1~-1',
                'referer': 'https://www.zalando.de/',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36'
            }
            # response = requests.get(url, proxies=proxy, headers=header)
            geturl_hit = requests.get(product_id_url,proxies=proxy, headers=headers, timeout=30)
            if geturl_hit.status_code == 200:
                saving_data = geturl_hit.text
                break


        print(geturl_hit.status_code)
        import datetime
        time = datetime.datetime.now()
        # d1 = time.strftime("%Y-%m-%dT%H:%M:%SZ")
        # d2 = str(d1).replace('-04-01','-03-31')
        details_dict = {'INPUT_PLATFORMID': prod_id_data[1],
                        #'INPUT_ADIDASID': 'Not Available',
                        'Website Name': 'ABOUTYOU',
                        'Country': 'DE',
                        'RPC':'',
                        'MPC':'',
                        'Product_ID': '',
                        'Product_URL': product_id_url,
                        'Product_Name': '',
                        'Category_Path': '',
                        'Specification': '',
                        'Description': '',
                        'Currency': 'EURO',
                        'List_Price': '',
                        'Promo_Price': '',
                        'Discount': '',
                        'Brand': '',
                        'Rating_Count': '',
                        'Review_Count': '',
                        'Image_URLs': '',
                        'Variant': '',
                        'Variant_ID': '',
                        'Colour_of_Variant': '',
                        'Colour_Grouping': '',
                        'Seller_Name': '',
                        'Stock_Count': '',
                        'Stock_Condition': '',
                        'Stock_Message': '',
                        'Sustainability_Badge': '',
                        'Reason_Code': '',
                        'Crawling_TimeStamp': time.strftime("%Y-%m-%dT%H:%M:%SZ"),
                        'Cache_Page_Link': '',
                        'Extra1': '',
                        'Extra2': '',
                        'Extra3': '',
                        'Extra4': '',
                        'Extra5': ''
                        }
        if geturl_hit.status_code == 200:
            datazone = datetime.datetime.now()
            f_date = datazone.strftime("%d_%m_%Y")
            # f_date='28_04_2022'
            strdate = datazone.day
            strm = datazone.month
            stry = datazone.year
            pageid = sku
            cpid = pageid + '_' + str(strdate) + '_' + str(strm) + '_' + str(stry)
            # cpid=pageid+'_'+'28_04_2022'
            # cpid1 = pageid + '_31_03_2022'
            # ASS_folder = f"E:\ADIDAS_SavePages\Jdsports_uk\ASS"

            ASS_folder = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Aboutyou_de\\PDP"
            sos_date_wise_folder = ASS_folder + f"\\{f_date}"
            if os.path.exists(sos_date_wise_folder):
                pass
            else:
                os.mkdir(sos_date_wise_folder)
            sos_filename = sos_date_wise_folder + "\\" + cpid + ".html"
            sos_filename = sos_filename.replace("+", "_").replace("-", "_")
            page_path = sos_filename.replace('/', '')
            print(page_path)
            page_path = page_path.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\',
                                          'https:////ecxus440.eclerx.com//cachepages//').replace('\\',
                                                                                                 '//').replace(
                '//', '/')
            print(page_path)
            # Cache_Page_Link = page_path
            details_dict['Cache_Page_Link'] = page_path
            if os.path.exists(sos_filename):
                with open(sos_filename, 'w', encoding='utf-8') as f:
                    f.write(saving_data)
            else:
                with open(sos_filename, 'w', encoding='utf-8') as f:
                    f.write(saving_data)

            # # 31032022  ------------------------------------------------------------
            # cpid1 = pageid + '_31_03_2022'
            # # ASS_folder = f"E:\ADIDAS_SavePages\Jdsports_uk\ASS"
            # f_date1 = '31_03_2022'
            # ASS_folder1 = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Aboutyou_de\\PDP"
            # sos_date_wise_folder1 = ASS_folder1 + f"\\{f_date1}"
            # if os.path.exists(sos_date_wise_folder1):
            #     pass
            # else:
            #     os.mkdir(sos_date_wise_folder1)
            # sos_filename1 = sos_date_wise_folder1 + "\\" + cpid1 + ".html"
            # sos_filename1 = sos_filename1.replace("+", "_").replace("-", "_")
            # page_path1 = sos_filename1.replace('/', '')
            # print(page_path1)
            # page_path1 = page_path1.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\',
            #                               'https:////ecxus440.eclerx.com//cachepages//').replace('\\',
            #                                                                                      '//').replace(
            #     '//', '/')
            # print(page_path1)
            # # Cache_Page_Link = page_path
            # details_dict['Extra1'] = page_path1
            # if os.path.exists(sos_filename1):
            #     with open(sos_filename1, 'w', encoding='utf-8') as f:
            #         f.write(saving_data)
            # else:
            #     with open(sos_filename1, 'w', encoding='utf-8') as f:
            #         f.write(saving_data)
            # # ---------------------------------------------------------------------------------------------

            tree2 = html.fromstring(geturl_hit.content)

            data = tree2.xpath('//script[@data-tadarida-initial-state="true"]/text()')
            # data=tree2.xpath('//main[@id="react-root"]/script/text()')
            # print(data)
            all_json_data = json.loads(data[0])
            # print(all_json_data)


           # details_dict['INPUT_PLATFORMID'] = prod_id_data[0].split('-')[-1]

            #details_dict['Product_ID'] = 'no'
            rpc = product_id_url[-7:]
            details_dict['RPC'] = product_id_url[-7:]
            details_dict['MPC'] = "-"
            details_dict['Product_ID'] = rpc
            details_dict['Product_URL'] = 'https://www.aboutyou.de'+all_json_data[2][1]["data"]["priceAndBrandSection"]["product"]["path"]
            # details_dict['Product_URL']=product_id_url
            details_dict['Product_Name'] = all_json_data[2][1]["data"]["priceAndBrandSection"]["product"]["name"]
            #cat_path = tree2.xpath('//ul[@class="sc-gqnvwf-1 kLNcem"]/li/a/text()')
            #cat_path=tree2.xpath('//ul[@class="sc-gqnvwf-1 fQHkse"]/li/a/text()')
            cat_path=tree2.xpath('//ul[@class="sc-gqnvwf-1 hOEZmQ"]/li/a/text()')
            r_path = 'aboutyou.de'
            for val in cat_path:
                new = r_path + '>' + val
                r_path = new
            # print(r_path)
            details_dict['Category_Path'] = r_path
            # details_dict['Category_Path'] = ''
            try:
                mt = all_json_data[2][1]["data"]["productDetailsSection"]["articleDetails"]["lanes"]
                print(mt)
                # jsonrawspec = '{"data":' + mt + '}'
                # jsospec = json.loads(jsonrawspec)
                spectext = ''
                for data in mt:
                    spec1 = ''
                    spec2 = ''
                    spec3 = ''

                    try:
                        if data["label"] == 'Material & Pflege':
                            specs = data["type"]["materialLane"]["bulletPoints"]
                            spec1 = '|'.join(i for i in specs)
                            spec1 = 'Material & Pflege - ' + spec1
                        else:
                            spec1 = ''

                        if data["label"] == 'Größe & Fit':
                            specs = data["type"]["sizeLane"]["bulletPoints"]
                            spec2 = '|'.join(i for i in specs)
                            spec2 = 'Größe & Fit - ' + spec2
                        else:
                            spec2 = ''

                        if data["label"] == 'Funktionalität':
                            specs = data["type"]["regularLane"]["items"]
                            spec3 = '|'.join(i for i in specs)
                            spec3 = 'Funktionalität - ' + spec3
                        else:
                            spec3 = ''
                        spec0 =''
                        if spec3:
                            spec0 = spec3
                        if spec2:
                            spec0 = spec2
                        if spec1:
                            spec0 = spec1
                    except:
                        spec0 = ''

                    if spec0:
                        if spectext == '':
                            spectext = spec0
                        else:
                            spectext = spectext + "|" + spec0
                spectext=general.clean(spectext)
                details_dict['Specification'] = spectext
                print(spectext)

                #time.sleep(200000000000000000000)
            except Exception as e:
                print(e)
                details_dict['Specification'] = '-'
            try:
                chekdata = all_json_data[2][1]["data"]["productDetailsSection"]["articleDetails"]
                lebal = all_json_data[2][1]["data"]["productDetailsSection"]["articleDetails"]["lanes"][0]["label"]
                disc = all_json_data[2][1]["data"]["productDetailsSection"]["articleDetails"]["lanes"][0]["type"]["bulletPointLane"]["bulletPoints"][0::1]
                desc1 = '|'.join(ele for ele in disc)
                desc1=general.clean(desc1)
                details_dict['Description'] = lebal + ' - ' + desc1
            except:
                details_dict['Description'] = '-'


            details_dict['Currency']='EURO'
            try:
                #price = tree2.xpath('//div[@class="sc-18q4lz4-2 izorSa sc-pslaq3-4 eigJiG"]/span/text()')
                price=tree2.xpath('//div[@class="sc-18q4lz4-2 jdvTpQ sc-pslaq3-4 cxlXew"]/span/text()')
                if len(price) > 1:

                    #value_MRP = tree2.xpath('//div[@class="sc-18q4lz4-2 izorSa sc-pslaq3-4 eigJiG"]/span[2]/text()')[0]
                    value_MRP = tree2.xpath('//div[@class="sc-18q4lz4-2 jdvTpQ sc-pslaq3-4 cxlXew"]/span[2]/text()')[0]
                    mrp = value_MRP.replace('EUR','').replace('ab','')
                    details_dict['List_Price']=mrp
                    promo= tree2.xpath('//div[@class="sc-18q4lz4-2 jdvTpQ sc-pslaq3-4 cxlXew"]/span[1]/text()')[0]
                    promo = general.clean(promo)
                    pro=promo.replace('EUR','').replace('ab','').strip()
                    details_dict['Promo_Price']=pro
                    dicso = tree2.xpath('//div[@class="sc-pslaq3-3 fVRfZP"]/div/text()')
                    if dicso:
                        dicso = dicso[0]

                    if '%' in dicso:
                        details_dict['Discount']=dicso
                    else:
                        cal=int((1-(float(pro.replace(',','.'))/float(mrp.replace(',','.'))))*100)
                        details_dict['Discount']='{}'.format(cal)+'%'
                        # print()


                else:

                    value_MRP = tree2.xpath('//div[@class="sc-18q4lz4-2 jdvTpQ sc-pslaq3-4 cxlXew"]/span/text()')[0]
                    if value_MRP:
                        promo = value_MRP
                    details_dict['List_Price'] = value_MRP.replace('EUR','').replace('ab','')
                    details_dict['Promo_Price'] = value_MRP.replace('EUR','').replace('ab','')
                    details_dict['Discount'] = '-'
            except:
                details_dict['List_Price'] = '-'
                details_dict['Promo_Price'] = '-'
                details_dict['Discount'] = '-'
                promo = ''


            # details_dict['List_Price']=
            # details_dict['Promo_Price']=
            # details_dict['Discount']=
            details_dict['Brand']=all_json_data[2][1]["data"]["priceAndBrandSection"]["brandLogo"]["brand"]["name"]

            details_dict["Rating_Count"] = 'Not Available'

            details_dict["Review_Count"]='Not Available'
            img=tree2.xpath('//ul[@class="sc-9kkfav-0 fcvgec"]/li/button/div/img/@srcset')
            single=[]
            for ele in img:
               single.append(ele.split(',')[-2])# just add the [-2] for single imgae
            details_dict["Image_URLs"] = '|'.join(ele for ele in single)
            # details_dict['Image_URLs']='|'.join(i for i in img)

            details_dict['Variant'] = 'not fetched'
            #try:
            #    details_dict['Variant_ID'] =tree2.xpath('//span[@class="sc-2qclq4-0 ePNAqF"]/text()[3]')
            #except:
            #    details_dict['Variant_ID']='-'

            details_dict['Variant_ID']=rpc

            try:
                details_dict['Colour_of_Variant']=all_json_data[2][1]["data"]["priceAndBrandSection"]["product"]["colors"][0]["name"]
            except:
                details_dict['Colour_of_Variant'] = '-'

            colors = tree2.xpath('//span[contains(@data-testid,"productColorOptionsOption")]/@name')
            if len(colors) == 0:
                details_dict['Colour_Grouping'] = '-'
            else:
                details_dict['Colour_Grouping'] = '|'.join(i for i in colors)

            details_dict['Seller_Name'] = 'Not Available'

            details_dict['Stock_Count'] = 'Not Available'
            details_dict['Stock_Condition'] = 'not fetched'
            details_dict['Stock_Message']='Not Available'
            if '"label":"Nachhaltigkeit"' in saving_data:
                details_dict['Sustainability_Badge'] = 'Yes'
            else:
                details_dict['Sustainability_Badge'] = 'No'
            details_dict['Reason_Code'] = 'Success-PF'

            if 'Deine Farbe ist ausverkauft<' in saving_data and not promo:

                check = 1
                details_dict['Variant'] = '-'
                details_dict['Stock_Condition'] = 'Out Of Stock'
                details_dict['Product_ID'] = '-'
                store_data_by_product_id(dict=[details_dict])
            elif 'Dieses Produkt ist ausverkauft' in saving_data and not promo:
                check = 1
                details_dict['Variant'] = '-'
                details_dict['Stock_Condition'] = 'Out Of Stock'
                details_dict['Product_ID'] = '-'
                store_data_by_product_id(dict=[details_dict])

            else:
                size=all_json_data[2][1]["data"]["productSelectionSection"]["product"]["sizes"]

                print('----------------------------------------------------')
                print(size)
                try:
                    try:
                        key=size[1]["shopSize"]["size"]['$case']
                    except:
                        try:
                            key = size[1]["vendorSize"]["size"]['$case']
                        except:
                            try:
                                key = size[0]["shopSize"]["size"]['$case']
                            except:
                                key = size[0]["vendorSize"]["size"]['$case']

                    try:
                        key1 = size[0]["shopSize"]["size"]['$case']
                    except:
                        key1 = size[0]["vendorSize"]["size"]['$case']
                    print(key)
                    variant_dict = '1'
                    if 'pants' in key and 'singleDimension' in key1:
                        for i in size:
                            variant_dict = 1
                            if i["quantity"] == 0:
                                try:
                                    var_dict = i["vendorSize"]["size"][key1]
                                except:
                                    var_dict = i["vendorSize"]["size"][key]

                                try:
                                    details_dict['Variant'] = var_dict['label']
                                except:
                                    try:
                                        details_dict['Variant'] = var_dict["firstDimension"]
                                    except:
                                        details_dict['Variant'] = var_dict["dimension"]

                                if variant_dict:
                                    details_dict['Stock_Condition'] = "Out of Stock"
                                    details_dict['Product_ID'] = i["articleNumber"]
                                    store_data_by_product_id(dict=[details_dict])
                                    print(details_dict)
                            else:
                                variant_dict = 1
                                try:
                                    var_dict = i["vendorSize"]["size"][key1]
                                except:
                                    var_dict = i["vendorSize"]["size"][key]
                                print(var_dict)
                                try:
                                    details_dict['Variant'] = var_dict['label']
                                except:
                                    try:
                                        details_dict['Variant'] = var_dict["firstDimension"]
                                    except:
                                        details_dict['Variant'] = var_dict["dimension"]
                                if variant_dict:
                                    details_dict['Stock_Condition'] = "In Stock"
                                    details_dict['Product_ID'] = i["articleNumber"]
                                    print(details_dict)
                                    store_data_by_product_id(dict=[details_dict])
                                    print(details_dict)
                    elif 'pants' in key1 and 'singleDimension' in key:
                        for i in size:
                            variant_dict = 1
                            if i["quantity"] == 0:
                                try:
                                    var_dict = i["vendorSize"]["size"][key1]
                                except:
                                    variant_dict = ''

                                try:
                                    details_dict['Variant'] = var_dict['label']
                                except:
                                    try:
                                        details_dict['Variant'] = var_dict["firstDimension"]
                                    except:
                                        details_dict['Variant'] = var_dict["dimension"]

                                if variant_dict:
                                    details_dict['Stock_Condition'] = "Out of Stock"
                                    details_dict['Product_ID'] = i["articleNumber"]
                                    store_data_by_product_id(dict=[details_dict])
                                    print(details_dict)
                            else:
                                variant_dict = 1
                                try:
                                    var_dict = i["vendorSize"]["size"][key1]
                                except:
                                    var_dict = i["vendorSize"]["size"][key]
                                print(var_dict)
                                try:
                                    details_dict['Variant'] = var_dict['label']
                                except:
                                    try:
                                        details_dict['Variant'] = var_dict["firstDimension"]
                                    except:
                                        details_dict['Variant'] = var_dict["dimension"]
                                if variant_dict:
                                    details_dict['Stock_Condition'] = "In Stock"
                                    details_dict['Product_ID'] = i["articleNumber"]
                                    print(details_dict)
                                    store_data_by_product_id(dict=[details_dict])
                                    print(details_dict)
                    elif 'bra' in key or 'pants' in key:
                        for i in size:
                            if i["quantity"] == 0:
                                try:
                                    var_dict = i["vendorSize"]["size"][key]
                                except:
                                    var_dict = i["vendorSize"]["size"][key1]
                                try:
                                    details_dict['Variant'] = var_dict['label']
                                except:
                                    details_dict['Variant'] = var_dict["firstDimension"]
                                details_dict['Stock_Condition'] = "Out of Stock"
                                details_dict['Product_ID'] = i["articleNumber"]
                                store_data_by_product_id(dict=[details_dict])
                                print(details_dict)
                            else:
                                var_dict = i["shopSize"]["size"][key]
                                print(var_dict)
                                try:
                                    details_dict['Variant'] = var_dict['label']
                                except:
                                    try:
                                        details_dict['Variant'] = var_dict["firstDimension"]
                                    except:
                                        details_dict['Variant'] = var_dict["firstDimension"]
                                details_dict['Stock_Condition'] = "In Stock"
                                details_dict['Product_ID'] = i["articleNumber"]
                                print(details_dict)
                                store_data_by_product_id(dict=[details_dict])
                                print(details_dict)
                    else:
                        for i in size:
                            if i["quantity"]==0:
                                try:
                                    if 'singleDimension' in key:
                                        var_dict = i["vendorSize"]["size"][key]
                                    else:
                                        var_dict=i["shopSize"]["size"][key]
                                    print(var_dict)
                                except:
                                    if 'singleDimension' in key:
                                        var_dict = i["vendorSize"]["size"][key1]
                                    else:
                                        var_dict=i["shopSize"]["size"][key1]
                                    print(var_dict)
                                try:
                                    details_dict['Variant']=var_dict['dimension']
                                except:
                                    details_dict['Variant'] = var_dict["firstDimension"]

                                details_dict['Stock_Condition']="Out of Stock"
                                details_dict['Product_ID'] = i["articleNumber"]
                                store_data_by_product_id(dict=[details_dict])
                                print(details_dict)
                            else:
                                try:
                                    if 'singleDimension' in key:
                                        var_dict = i["vendorSize"]["size"][key]
                                    else:
                                        var_dict = i["shopSize"]["size"][key]
                                    print(var_dict)
                                except:
                                    if 'singleDimension' in key:
                                        var_dict = i["vendorSize"]["size"][key1]
                                    else:
                                        var_dict = i["shopSize"]["size"][key1]
                                try:
                                    details_dict['Variant'] = var_dict['dimension']
                                except:
                                    try:
                                        details_dict['Variant'] = var_dict["firstDimension"]
                                    except:
                                        details_dict['Variant'] = var_dict["firstDimension"]
                                details_dict['Stock_Condition'] = "In Stock "
                                details_dict['Product_ID'] =  i["articleNumber"]
                                store_data_by_product_id(dict=[details_dict])
                                print(details_dict)
                except Exception as e:
                    details_dict['Variant'] = '-'
                    if promo:
                        details_dict['Stock_Condition'] = 'In Stock'
                    else:
                        details_dict['Stock_Condition'] = 'Out Of Stock'
                    details_dict['Product_ID'] = '-'
                    store_data_by_product_id(dict=[details_dict])

        else:
            details_dict['Reason_Code'] = 'Success-PNF'
            print(details_dict)
            store_data_by_product_id(dict=[details_dict])
    except Exception as e:
        print(e)
        """
        details_dict = {'INPUT_PLATFORMID': sku,
                        #'INPUT_ADIDASID': 'Not Available',
                        'Website Name': 'ABOUTYOU',
                        'Country': 'DE',
                        'RPC': '',
                        'MPC': '',
                        'Product_ID': '',

                        'Product_URL': product_id_url,
                        'Product_Name': '',
                        'Category_Path': '',
                        'Specification': '',
                        'Description': '',
                        'Currency': 'EURO',
                        'List_Price': '',
                        'Promo_Price': '',
                        'Discount': '',
                        'Brand': '',
                        'Rating_Count': '',
                        'Review_Count': '',
                        'Image_URLs': '',
                        'Variant': '',
                        'Variant_ID': '',
                        'Colour_of_Variant': '',
                        'Colour_Grouping': '',
                        'Seller_Name': '',
                        'Stock_Count': '',
                        'Stock_Condition': '',
                        'Stock_Message': '',
                        'Sustainability_Badge': '',
                        'Reason_Code': 'Blocked',
                        'Crawling_TimeStamp': '',
                        'Cache_Page_Link': '',
                        'Extra1': '',
                        'Extra2': '',
                        'Extra3': '',
                        'Extra4': '',
                        'Extra5': ''
                        }
        store_data_by_product_id(dict=[details_dict])
        """


for data in extract_xl():
    print('inside 1')
    by_product_id(prod_id_data=data)