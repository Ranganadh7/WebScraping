import random
import re
import general
from lxml import html
import time
# import xlrd
import requests
from lxml import html
import time
import json
from datetime import datetime
# import datetime
import openpyxl


headers={'authority': 'www.snipes.com',
'method': 'GET',
'path': '/',
'scheme': 'https',
'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
'accept-encoding': 'gzip, deflate, br',
'accept-language': 'en-US,en;q=0.9',
'cache-control': 'max-age=0',
'cookie': '_ga_TK0HQQCMWW=GS1.1.1642433622.2.0.1642433622.0; sid=h79PoYYY24Rrk3UVEzwnJ81EamwtW65233U; dwanonymous_f489fce3a63fde16454e4c1b6c85eee9=adSKD2bPGLeVSMQITWIgaupGZg; __cq_dnt=1; dw_dnt=1; dwsid=B7I9PzupcR7hzfB3GxaH701JGPKQ6gTsVSvvbYPmh4zAOvHuipXHXkYKx4hNtmH5UU0a-whkPq9AKXzZLLq7LQ==; __cf_bm=mSw6L_2IYdoYL8GiHWXgt4hQKVt92QdlI9KFX05eALI-1643357151-0-AfB/HDBaJtR6uxkGELaWJMuZurHL2ivFFJJul5vsUl/9vE90XDiIi3StpZewAIR3mGQJ/Z8imOt7+mQYOYByGxk=; test; OptanonAlertBoxClosed=2022-01-28T08:06:35.019Z; OptanonConsent=isIABGlobal=false&datestamp=Fri+Jan+28+2022+13%3A46%3A34+GMT%2B0530+(India+Standard+Time)&version=6.15.0&hosts=&consentId=4a9b15bc-0ac7-4e21-9489-3e5e4f6271fe&interactionCount=2&landingPath=NotLandingPage&groups=C0001%3A1%2CC0007%3A0%2CC0004%3A0&geolocation=%3B&AwaitingReconsent=false; customerCountry=in; hideLocalizationDialog=true',
'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
'sec-ch-ua-mobile': '?0',
'sec-ch-ua-platform': "Windows",
'sec-fetch-dest': 'document',
'sec-fetch-mode': 'navigate',
'sec-fetch-site': 'same-origin',
'sec-fetch-user': '?1',
'upgrade-insecure-requests': '1',
'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36'}
headers = {
    'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Google Chrome";v="99"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.84 Safari/537.36',
}
#
# file_path='C:\\Users\\Suraj.Kashid\\PycharmProjects\\adidas_1\\SNIPES_COM\\SNIPES_INPUTS\\PDP Input Template2.xlsx'
# file_path=r'PDP Input Template.xlsx'
file_path=r'PDP Input Template.xlsx'
def extract_xl():
    workbook= openpyxl.load_workbook(file_path)
    sheet_obj= workbook.active
    all_rows=sheet_obj.max_row
    inputdata=[]
    for i in range(1,all_rows):
        limit=sheet_obj.cell(row=i+1,column=3)
        urls= sheet_obj.cell(row=i+1,column=4)

        u_data=[urls.value,limit.value]

        inputdata.append(u_data)

    return inputdata
# extract_xl()

from csv import DictWriter

def store_data_by_product_id(dict):
    fieldnames = ['SKU_ID', 'Website', 'Country','RPC','MPC','Product_ID', 'Product_URL',
                  'Product_Name', 'Category_Path', 'Specification', 'Description', 'Currency', 'List_Price',
                  'Promo_Price', 'Discount', 'Brand', 'Rating_Count', 'Review_Count', 'Image_URLs', 'Variant',
                  'Variant_ID', 'Colour_of_Variant', 'Colour_Grouping', 'Seller_Name', 'Stock_Count', 'Stock_Condition',
                  'Stock_Message', 'Sustainability_Badge', 'Reason_Code', 'Crawling_TimeStamp', 'Cache_Page_Link',
                  'Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5']
    with open('08-05-snipes_pdp.csv', 'a+', encoding='utf-8-sig', newline='') as file:
        writer = DictWriter(file, fieldnames=fieldnames)
        # writer.writeheader()
        writer.writerows(dict)
        file.close()

def scraping_inputs(recursive_limit=2):
    # type_list = [1,14]
    # domain
    domain=''
    resp=requests.get('http://10.100.22.203/get_data',params={'type_id': [1],'domain_id':[16]},timeout=30)
    print(resp.status_code)
    if resp.status_code==200:
        print(resp.text)
        return resp.text
    elif recursive_limit != 0:
        new_n = recursive_limit - 1
        scraping_inputs(recursive_limit=new_n)
    elif resp.status_code == 408:
        print('TimeOut')
    else:
        print('Blocked')
# scraping_inputs()


#api-2
import requests
def status_update(status_dict,recursive_limit=2):
    dict_values=status_dict
    response=requests.post('http://10.100.22.203/post_data',json=[dict_values] )
    if response.status_code==200:
        print(response.text)
        return response.text
    elif recursive_limit != 0:
        new_n = recursive_limit - 1
        status_update(recursive_limit=new_n)
    elif response.status_code == 408:
        print('TimeOut')
    else:
        print('page_not_found')

import datetime
import os
def html_cache_page_saving(sku,saving_data):
    datazone = datetime.datetime.now()
    f_date = datazone.strftime("%d_%m_%Y")
    #f_date = '29_04_2022'
    strdate = datazone.day
    strm = datazone.month
    stry = datazone.year
    pageid = sku
    cpid = pageid + '_' + f_date
    #cpid = pageid + "_" + f_date
    # ASS_folder = f"E:\ADIDAS_SavePages\Jdsports_uk\ASS"

    ASS_folder = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Snipes_DE\\PDP"
    sos_date_wise_folder = ASS_folder + f"\\{f_date}"
    if os.path.exists(sos_date_wise_folder):
        pass
    else:
        os.mkdir(sos_date_wise_folder)
    sos_filename = sos_date_wise_folder + "\\" + cpid + ".html"
    sos_filename = sos_filename.replace("+", "_").replace("-", "_")
    page_path = sos_filename.replace('/', '')
    print(page_path)
    page_path = page_path.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\',
                                  'https:////ecxus440.eclerx.com//cachepages//').replace('\\',
                                                                                         '//').replace(
        '//', '/')
    print(page_path)
    if os.path.exists(sos_filename):
        with open(sos_filename, 'w', encoding='utf-8') as f:
            f.write(saving_data)
    else:
        with open(sos_filename, 'w', encoding='utf-8') as f:
            f.write(saving_data)
    return page_path


# from adidas_proxy import AdidasProxy
# ad_pr=AdidasProxy()
p1 = ['104.218.195.130','104.218.195.205','104.251.82.191','104.251.82.240','104.251.82.63','104.251.84.104','104.251.84.217','104.251.84.232','104.251.85.123','104.251.85.196','104.251.86.162','104.251.86.167','104.251.86.209','104.251.90.200','104.251.90.237','104.251.90.69','104.251.91.154','104.251.91.233','104.251.92.178','104.251.92.234','104.251.92.63','108.177.131.182','108.177.131.25','146.19.55.151','146.19.55.167','154.13.200.241','154.13.200.34','154.13.200.48','154.13.201.156','154.13.201.245','154.13.202.117','154.13.202.128','154.13.202.146','154.13.203.144','154.13.203.212','154.13.204.132','154.13.204.98','154.13.205.173','154.13.205.40','154.13.206.167','154.13.206.181','154.13.207.213','154.13.207.233','154.13.244.123','154.13.244.139','154.13.244.234','154.13.245.128','154.13.245.133','154.13.245.152','154.13.246.157','154.13.246.158','154.13.246.159','154.13.247.187','154.13.247.219','154.13.247.42','154.13.248.163','154.13.248.36','154.13.248.95','154.13.249.100','154.13.249.42','154.13.250.164','154.13.250.89','154.13.251.141','154.13.251.52','154.13.251.76','154.13.252.114','154.13.252.163','154.13.252.79','154.13.253.101','154.13.253.185','154.13.253.195','154.13.254.141','154.13.254.99','154.13.255.222','154.13.255.248','154.17.157.182','154.17.157.234','154.17.157.50','154.17.188.153','154.17.188.24','154.17.189.230','154.17.189.30','154.29.2.16','154.29.2.196','154.29.2.231','154.37.72.173','154.37.72.59','154.37.76.117','154.37.76.137','154.37.76.187','158.115.224.142','158.115.224.246','158.115.225.241','158.115.225.253','158.115.226.137','158.115.226.92','158.115.227.120','158.115.227.174','165.140.224.123','165.140.224.184','165.140.225.115','165.140.225.230','165.140.225.46','165.140.226.14','165.140.226.244','165.140.226.46','165.140.227.12','165.140.227.245','168.91.64.213','168.91.64.234','168.91.64.251','168.91.65.46','168.91.65.73','168.91.66.106','168.91.66.123','168.91.67.109','168.91.67.17','168.91.84.133','168.91.84.59','168.91.85.212','168.91.85.30','168.91.86.14','168.91.86.21','168.91.87.79','168.91.87.97','168.91.88.214','168.91.88.245','168.91.90.127','168.91.90.49','172.255.93.114','172.255.93.130','172.255.94.155','172.255.94.158','173.208.27.32','173.208.27.93','173.208.28.162','173.208.28.246','173.234.244.244','173.234.244.79','173.245.75.175','173.245.75.54','173.245.85.105','173.245.85.116','173.245.85.45','173.245.90.138','173.245.90.224','185.255.196.162','185.255.196.168','185.255.197.105','185.255.197.110','198.251.92.13','198.251.92.237','198.251.92.29','198.251.93.165','198.251.93.227','198.251.93.237','207.230.104.136','207.230.104.195','207.230.104.90','207.230.105.118','207.230.105.205','207.230.105.84','207.230.106.19','207.230.106.198','207.230.106.204','207.230.107.92','207.230.107.95','213.109.148.122','213.109.148.23','23.105.0.165','23.105.0.224','23.105.0.63','23.105.142.171','23.105.142.57','23.105.143.213','23.105.143.73','23.105.144.123','23.105.144.96','23.105.145.215','23.105.145.242','23.105.146.181','23.105.146.245','23.105.147.152','23.105.147.192','23.105.147.203','23.105.150.11','23.105.150.199','23.105.151.138','23.105.151.3','23.105.3.42','23.105.3.68','23.105.4.172','23.105.4.231','23.106.16.106','23.106.16.203','23.106.18.234','23.106.18.44','23.106.20.181','23.106.20.233','23.106.22.125','23.106.22.147','23.106.24.173','23.106.24.41','23.106.26.65','23.106.26.84','23.106.27.13','23.106.27.139','23.106.27.144','23.106.28.230','23.106.28.237','23.106.30.117','23.106.30.126','23.110.166.102','23.110.166.26','23.110.166.76','23.110.169.100','23.110.169.162','23.110.173.171','23.110.173.225','23.129.136.120','23.129.136.245','23.129.40.19','23.129.40.44','23.129.40.76','23.129.56.191','23.129.56.237','23.161.3.146','23.161.3.67','23.170.144.104','23.170.144.108','23.170.144.19','23.170.145.103','23.170.145.252','23.170.145.51','23.175.176.21','23.175.176.24','23.175.177.176','23.175.177.183','23.175.177.8','23.176.49.110','23.176.49.183','23.177.240.144','23.177.240.217','23.177.240.90','23.184.144.105','23.184.144.124','23.184.144.231','23.185.112.167','23.185.112.229','23.185.144.110','23.185.144.171','23.185.144.197','23.185.80.164','23.185.80.4','23.185.80.6','23.186.48.210','23.186.48.248','23.226.16.211','23.226.16.243','23.226.17.178','23.226.17.240','23.226.18.106','23.226.18.193','23.226.19.212','23.226.19.87','23.226.20.187','23.226.20.90','23.226.21.13','23.226.21.22','23.226.22.216','23.226.22.53','23.226.23.178','23.226.23.250','23.226.24.190','23.226.24.6','23.226.24.64','23.226.25.107','23.226.25.68','23.226.26.158','23.226.26.202','23.226.26.220','23.226.27.246','23.226.27.94','23.226.28.159','23.226.28.194','23.226.28.231','23.226.29.126','23.226.29.99','23.226.30.191','23.226.30.235','23.226.31.131','23.226.31.169','23.226.31.193','23.247.172.197','23.247.172.214','23.247.172.51','23.247.173.202','23.247.173.6','23.247.174.196','23.247.174.211','23.247.174.81','23.247.175.156','23.247.175.215','23.247.175.218','23.27.9.103','23.27.9.228','23.82.105.11','23.82.105.194','23.82.105.45','23.82.109.165','23.82.109.242','23.82.184.118','23.82.184.227','23.82.184.80','23.82.186.178','23.82.186.223','23.82.40.136','23.82.40.202','23.82.40.42','23.82.41.119','23.82.41.40','23.82.41.6','23.82.44.209','23.82.44.48','23.82.80.145','23.82.80.68','23.82.81.171','23.82.81.79','45.146.117.234','45.146.117.253','45.146.118.204','45.146.118.228','45.146.119.223','45.146.119.244','45.154.141.33','45.154.141.50','45.154.142.21','45.154.142.231','45.154.142.42','45.224.228.187','45.224.228.87','45.224.228.94','45.224.230.211','45.224.230.228','45.224.231.141','45.224.231.68','45.237.84.117','45.237.84.33','45.237.86.170','45.237.86.178','45.238.157.141','45.238.157.198','45.238.157.225','45.238.159.115','45.238.159.59','45.238.159.8','45.59.128.144','45.59.128.198','45.59.128.236','45.59.129.177','45.59.129.217','45.59.130.16','45.59.130.218','45.59.131.107','45.59.131.140','45.59.131.209','45.59.180.245','45.59.180.58','45.59.181.38','45.59.181.71','45.59.181.80','45.59.182.209','45.59.182.217','45.59.183.171','45.59.183.214','45.59.183.67','45.71.19.128','45.71.19.159','52.128.0.45','52.128.0.98','52.128.1.105','52.128.1.124','52.128.10.164','52.128.10.20','52.128.11.123','52.128.11.71','52.128.12.46','52.128.12.70','52.128.13.125','52.128.13.207','52.128.14.107','52.128.14.115','52.128.14.30','52.128.15.114','52.128.15.92','52.128.196.173','52.128.196.240','52.128.196.70','52.128.197.105','52.128.197.17','52.128.198.105','52.128.198.72','52.128.198.76','52.128.199.206','52.128.199.237','52.128.2.17','52.128.2.58','52.128.200.127','52.128.200.182','52.128.200.79','52.128.201.194','52.128.201.3','52.128.201.56','52.128.202.108','52.128.202.189','52.128.202.242','52.128.203.21','52.128.203.230','52.128.204.144','52.128.204.91','52.128.205.148','52.128.205.204','52.128.206.110','52.128.206.219','52.128.206.96','52.128.207.116','52.128.207.237','52.128.208.176','52.128.208.60','52.128.209.109','52.128.209.120','52.128.210.159','52.128.210.80','52.128.211.121','52.128.211.155','52.128.216.224','52.128.216.41','52.128.217.208','52.128.217.87','52.128.218.165','52.128.218.65','52.128.219.165','52.128.219.177','52.128.219.44','52.128.220.40','52.128.220.48','52.128.221.181','52.128.221.230','52.128.222.161','52.128.222.38','52.128.223.201','52.128.223.219','52.128.3.86','52.128.3.90','52.128.4.106','52.128.4.227','52.128.5.35','52.128.5.52','52.128.6.111','52.128.6.149','52.128.6.93','52.128.7.247','52.128.7.81','52.128.8.195','52.128.8.33','52.128.9.12','52.128.9.177','62.3.61.119','62.3.61.2','62.3.61.40','88.218.172.175','88.218.172.203','88.218.172.83','88.218.173.65','88.218.173.94','88.218.173.98','88.218.174.16','88.218.174.44','88.218.175.235','88.218.175.5','88.218.175.57','95.164.224.233','95.164.224.38','95.164.225.124','95.164.225.21','95.164.226.220','95.164.226.221','95.164.226.7','95.164.227.152','95.164.227.199','95.164.227.206','95.164.236.212','95.164.236.249','95.164.236.58','95.164.237.111','95.164.237.251','95.164.238.183','95.164.238.206','95.164.239.113','95.164.239.253']
import pandas as pd
# lacoste-powercourt-2-0-menshoes
dom='http://www.snipes.com'
products_df = pd.DataFrame()
def by_product_id(prod_id_data,n):
    # new=
    # products_df = pd.DataFrame()
    # product_id_url='https://www.snipes.com/search?q={}&lang=de_DE'.format(str(prod_id_data[0]))
    product_id_url=prod_id_data[0]
    print(product_id_url)
    # pro_id=product_id_url.split('-')[-1]
    # print(product_id_url)
    # print(pro_id)
    skuId=prod_id_data[1]

    import datetime
    time = datetime.datetime.now()
    details_dict = {'SKU_ID': skuId,
                    'Website': 'SNIPES',
                    'Country': 'DE',
                    'RPC': '',
                    'MPC': '',
                    'Product_ID': '',
                    'Product_URL': product_id_url,
                    'Product_Name': '',
                    'Category_Path': '',
                    'Specification': '',
                    'Description': '',
                    'Currency': 'EURO',
                    'List_Price': '',
                    'Promo_Price': '',
                    'Discount': 'Not Available',
                    'Brand': '',
                    'Rating_Count': 'Not Available',
                    'Review_Count': 'Not Available',
                    'Image_URLs': '',
                    'Variant': '',
                    'Variant_ID': '',
                    'Colour_of_Variant': '',
                    'Colour_Grouping': '',
                    'Seller_Name': 'Not Available',
                    'Stock_Count': 'Not Available',
                    'Stock_Condition': '-',
                    'Stock_Message': 'Not Available',
                    'Sustainability_Badge': 'Not Available',
                    'Reason_Code': '',
                    'Crawling_TimeStamp': time.strftime("%Y-%m-%dT%H:%M:%SZ"),
                    'Cache_Page_Link': '',
                    'Extra1': '',
                    'Extra2': '',
                    'Extra3': '',
                    'Extra4': '',
                    'Extra5': ''
                    }

    try:
        mpp_proxy = ['154.28.67.106', '154.28.67.111', '154.28.67.116', '154.28.67.117', '154.28.67.125',
                     '154.28.67.131', '154.28.67.133', '154.28.67.142', '154.28.67.156', '154.28.67.163',
                     '154.28.67.173', '154.28.67.18', '154.28.67.182', '154.28.67.184', '154.28.67.20', '154.28.67.200',
                     '154.28.67.210', '154.28.67.218', '154.28.67.222', '154.28.67.223', '154.28.67.231',
                     '154.28.67.240', '154.28.67.243', '154.28.67.253', '154.28.67.39', '154.28.67.4', '154.28.67.49',
                     '154.28.67.5', '154.28.67.61', '154.28.67.80', '154.28.67.81', '154.28.67.87', '154.28.67.88',
                     '154.28.67.96', '154.28.67.99', '154.7.230.100', '154.7.230.101', '154.7.230.103', '154.7.230.107',
                     '154.7.230.109', '154.7.230.130', '154.7.230.132', '154.7.230.14', '154.7.230.140',
                     '154.7.230.147', '154.7.230.151', '154.7.230.156', '154.7.230.163', '154.7.230.170',
                     '154.7.230.18', '154.7.230.183', '154.7.230.188', '154.7.230.189', '154.7.230.19', '154.7.230.190',
                     '154.7.230.198', '154.7.230.204', '154.7.230.209', '154.7.230.235', '154.7.230.238',
                     '154.7.230.246', '154.7.230.29', '154.7.230.41', '154.7.230.42', '154.7.230.51', '154.7.230.55',
                     '154.7.230.60', '154.7.230.61', '154.7.230.74', '154.7.230.82', '154.7.230.89', '23.131.8.112',
                     '23.131.8.115', '23.131.8.117', '23.131.8.12', '23.131.8.121', '23.131.8.124', '23.131.8.150',
                     '23.131.8.161', '23.131.8.166', '23.131.8.171', '23.131.8.173', '23.131.8.176', '23.131.8.177',
                     '23.131.8.181', '23.131.8.19', '23.131.8.192', '23.131.8.194', '23.131.8.199', '23.131.8.202',
                     '23.131.8.203', '23.131.8.204', '23.131.8.207', '23.131.8.209', '23.131.8.213', '23.131.8.216',
                     '23.131.8.225', '23.131.8.228', '23.131.8.231', '23.131.8.238', '23.131.8.254', '23.131.8.36',
                     '23.131.8.5', '23.131.8.76', '23.131.8.93', '23.131.8.95', '23.131.8.99', '23.131.88.105',
                     '23.131.88.12', '23.131.88.137', '23.131.88.139', '23.131.88.140', '23.131.88.145',
                     '23.131.88.150', '23.131.88.151', '23.131.88.153', '23.131.88.154', '23.131.88.156',
                     '23.131.88.165', '23.131.88.18', '23.131.88.191', '23.131.88.192', '23.131.88.194',
                     '23.131.88.198', '23.131.88.202', '23.131.88.206', '23.131.88.220', '23.131.88.223',
                     '23.131.88.228', '23.131.88.233', '23.131.88.24', '23.131.88.242', '23.131.88.244', '23.131.88.47',
                     '23.131.88.63', '23.131.88.67', '23.131.88.73', '23.131.88.80', '23.131.88.81', '23.131.88.82',
                     '23.131.88.88', '23.131.88.97', '23.170.144.149', '23.170.144.209', '23.170.144.212',
                     '23.170.144.242', '23.170.144.83', '23.170.145.117', '23.170.145.167', '23.170.145.182',
                     '23.170.145.19', '23.170.145.203', '23.226.17.101', '23.226.17.109', '23.226.17.112',
                     '23.226.17.113', '23.226.17.115', '23.226.17.123', '23.226.17.129', '23.226.17.143',
                     '23.226.17.148', '23.226.17.165', '23.226.17.186', '23.226.17.199', '23.226.17.201',
                     '23.226.17.207', '23.226.17.210', '23.226.17.219', '23.226.17.220', '23.226.17.222',
                     '23.226.17.229', '23.226.17.250', '23.226.17.254', '23.226.17.26', '23.226.17.33', '23.226.17.4',
                     '23.226.17.49', '23.226.17.5', '23.226.17.55', '23.226.17.66', '23.226.17.7', '23.226.17.72',
                     '23.226.17.78', '23.226.17.8', '23.226.17.86', '23.226.17.90', '23.226.17.93', '23.230.177.105',
                     '23.230.177.110', '23.230.177.113', '23.230.177.121', '23.230.177.130', '23.230.177.14',
                     '23.230.177.143', '23.230.177.15', '23.230.177.150', '23.230.177.154', '23.230.177.165',
                     '23.230.177.173', '23.230.177.191', '23.230.177.196', '23.230.177.203', '23.230.177.206',
                     '23.230.177.208', '23.230.177.217', '23.230.177.220', '23.230.177.221', '23.230.177.224',
                     '23.230.177.228', '23.230.177.231', '23.230.177.235', '23.230.177.237', '23.230.177.241',
                     '23.230.177.27', '23.230.177.38', '23.230.177.52', '23.230.177.61', '23.230.177.67',
                     '23.230.177.72', '23.230.177.80', '23.230.177.88', '23.230.177.94', '23.230.177.99',
                     '23.230.197.103', '23.230.197.106', '23.230.197.109', '23.230.197.11', '23.230.197.12',
                     '23.230.197.122', '23.230.197.124', '23.230.197.146', '23.230.197.155', '23.230.197.156',
                     '23.230.197.174', '23.230.197.179', '23.230.197.181', '23.230.197.196', '23.230.197.2',
                     '23.230.197.201', '23.230.197.207', '23.230.197.208', '23.230.197.225', '23.230.197.227',
                     '23.230.197.233', '23.230.197.236', '23.230.197.239', '23.230.197.240', '23.230.197.244',
                     '23.230.197.251', '23.230.197.50', '23.230.197.52', '23.230.197.54', '23.230.197.60',
                     '23.230.197.71', '23.230.197.80', '23.230.197.81', '23.230.197.84', '23.230.197.97',
                     '23.230.74.102', '23.230.74.110', '23.230.74.116', '23.230.74.125', '23.230.74.133',
                     '23.230.74.135', '23.230.74.14', '23.230.74.141', '23.230.74.149', '23.230.74.15', '23.230.74.157',
                     '23.230.74.16', '23.230.74.170', '23.230.74.172', '23.230.74.174', '23.230.74.183',
                     '23.230.74.187', '23.230.74.19', '23.230.74.198', '23.230.74.208', '23.230.74.212',
                     '23.230.74.215', '23.230.74.23', '23.230.74.230', '23.230.74.231', '23.230.74.252', '23.230.74.30',
                     '23.230.74.41', '23.230.74.57', '23.230.74.58', '23.230.74.59', '23.230.74.6', '23.230.74.75',
                     '23.230.74.81', '23.230.74.88', '23.230.74.91', '23.27.222.108', '23.27.222.109', '23.27.222.134',
                     '23.27.222.138', '23.27.222.159', '23.27.222.161', '23.27.222.164', '23.27.222.166',
                     '23.27.222.178', '23.27.222.19', '23.27.222.195', '23.27.222.201', '23.27.222.202',
                     '23.27.222.203', '23.27.222.208', '23.27.222.21', '23.27.222.211', '23.27.222.218',
                     '23.27.222.223', '23.27.222.228', '23.27.222.234', '23.27.222.236', '23.27.222.242',
                     '23.27.222.251', '23.27.222.253', '23.27.222.34', '23.27.222.61', '23.27.222.62', '23.27.222.69',
                     '23.27.222.70', '23.27.222.72', '23.27.222.73', '23.27.222.74', '23.27.222.81', '23.27.222.93',
                     '38.131.131.110', '38.131.131.114', '38.131.131.123', '38.131.131.125', '38.131.131.137',
                     '38.131.131.142', '38.131.131.145', '38.131.131.147', '38.131.131.15', '38.131.131.154',
                     '38.131.131.16', '38.131.131.17', '38.131.131.173', '38.131.131.18', '38.131.131.193',
                     '38.131.131.204', '38.131.131.207', '38.131.131.227', '38.131.131.229', '38.131.131.233',
                     '38.131.131.238', '38.131.131.246', '38.131.131.248', '38.131.131.250', '38.131.131.31',
                     '38.131.131.36', '38.131.131.50', '38.131.131.58', '38.131.131.64', '38.131.131.70',
                     '38.131.131.71', '38.131.131.74', '38.131.131.83', '38.131.131.94', '38.131.131.99',
                     '38.75.75.104', '38.75.75.111', '38.75.75.112', '38.75.75.119', '38.75.75.120', '38.75.75.123',
                     '38.75.75.127', '38.75.75.139', '38.75.75.14', '38.75.75.143', '38.75.75.155', '38.75.75.156',
                     '38.75.75.158', '38.75.75.170', '38.75.75.179', '38.75.75.188', '38.75.75.2', '38.75.75.201',
                     '38.75.75.231', '38.75.75.232', '38.75.75.241', '38.75.75.246', '38.75.75.251', '38.75.75.26',
                     '38.75.75.29', '38.75.75.4', '38.75.75.44', '38.75.75.49', '38.75.75.56', '38.75.75.58',
                     '38.75.75.62', '38.75.75.72', '38.75.75.76', '38.75.75.79', '38.75.75.88', '38.96.156.108',
                     '38.96.156.112', '38.96.156.128', '38.96.156.131', '38.96.156.14', '38.96.156.142',
                     '38.96.156.143', '38.96.156.149', '38.96.156.16', '38.96.156.163', '38.96.156.165',
                     '38.96.156.169', '38.96.156.186', '38.96.156.188', '38.96.156.190', '38.96.156.192',
                     '38.96.156.194', '38.96.156.199', '38.96.156.218', '38.96.156.236', '38.96.156.240',
                     '38.96.156.252', '38.96.156.28', '38.96.156.32', '38.96.156.35', '38.96.156.56', '38.96.156.57',
                     '38.96.156.6', '38.96.156.67', '38.96.156.77', '38.96.156.80', '38.96.156.83', '38.96.156.84',
                     '38.96.156.89', '38.96.156.92', '45.238.157.100', '45.238.157.104', '45.238.157.106',
                     '45.238.157.110', '45.238.157.116', '45.238.157.118', '45.238.157.119', '45.238.157.12',
                     '45.238.157.123', '45.238.157.132', '45.238.157.14', '45.238.157.149', '45.238.157.15',
                     '45.238.157.183', '45.238.157.186', '45.238.157.189', '45.238.157.2', '45.238.157.212',
                     '45.238.157.214', '45.238.157.217', '45.238.157.22', '45.238.157.228', '45.238.157.23',
                     '45.238.157.247', '45.238.157.43', '45.238.157.48', '45.238.157.51', '45.238.157.52',
                     '45.238.157.53', '45.238.157.56', '45.238.157.61', '45.238.157.65', '45.238.157.72',
                     '45.238.157.79', '45.238.157.8', '45.238.159.103', '45.238.159.107', '45.238.159.110',
                     '45.238.159.114', '45.238.159.116', '45.238.159.123', '45.238.159.126', '45.238.159.144',
                     '45.238.159.148', '45.238.159.15', '45.238.159.156', '45.238.159.165', '45.238.159.167',
                     '45.238.159.183', '45.238.159.20', '45.238.159.208', '45.238.159.217', '45.238.159.220',
                     '45.238.159.23', '45.238.159.230', '45.238.159.235', '45.238.159.237', '45.238.159.238',
                     '45.238.159.24', '45.238.159.249', '45.238.159.251', '45.238.159.32', '45.238.159.34',
                     '45.238.159.51', '45.238.159.6', '45.238.159.66', '45.238.159.77', '45.238.159.79',
                     '45.238.159.82', '45.238.159.91']

        p_auth = str("csimonra:h19VA2xZ")
        p_host = random.choice(mpp_proxy)
        p_port = "29842"
        proxy = {
            'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
            'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
        }
        headers = {
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Google Chrome";v="99"',
            'sec-ch-ua-mobile': '?0',
            'baqend-speed-kit-retry': '0',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.84 Safari/537.36',
            'baqend-speed-kit-request': 'dynamic',
            'cache-control': 'no-cache',
            'Referer': product_id_url,
            'sec-ch-ua-platform': '"Windows"',
        }
        #product_id_url = 'https://www.snipes.com/p/adidas_originals-forum_mid_-ftwr_white%2Fwonder_mauve%2Fftwr_white-00013802029039.html'
        response = requests.get(product_id_url, headers=headers,proxies=proxy ,timeout=30)
        print(response.status_code)
        general.write_file('snipes.html', response.text, 'a', encoding='utf-8')
        cache_data = response.text
        cache_page_l = html_cache_page_saving(sku=skuId, saving_data=cache_data)

        if response.status_code==200:
            tree=html.fromstring(response.content)

            try:
                # print(product_id_url)
                # tree1 = html.fromstring(response.text)
                # p = general.xpath(tree1, '//script[@type="application/ld+json"]')
                # print(p)
                #p = tree1.xpath('//ul[@class="b-breadcrumb js-breadcrumbs"]//li')
                prod_name = general.xpath(tree, '//div[@class="js-target"]/text()[2]')
                prod_name=general.clean(prod_name)
                print(prod_name)
                details_dict['Product_Name'] = prod_name
                #general.write_file('snipe.html', response.text, 'a', encoding='utf-8')
                brand = general.xpath(tree, '//a[@class="b-pdp-brand"]/text()')
                brand = brand
                print(brand)
                details_dict['Brand'] = brand

                cat_path = tree.xpath('//div[@class="l-row h-hide-sm h-hide-md"]/div/div/div/ul/li/a/text()')
                cat_values = [i for i in cat_path if i != '\n' and i != '\nZurück\n']
                final_word = tree.xpath('//div[@class="l-row h-hide-sm h-hide-md"]/div/div/div/ul/li/span/text()')[
                    0].replace('\n', '')
                prod_name = cat_values[-1]
                print(prod_name)

                details_dict['Category_Path'] = '>'.join(i.replace('\n', '') for i in cat_values) + '>' + final_word
                varient_col = tree.xpath('//div[@class="b-pdp-product-info-main"]/div/h1/div/span/text()')[0].replace(
                    '\n', '')
                details_dict['Colour_of_Variant'] = varient_col

                print(varient_col)
                try:
                    price = tree.xpath('//div[@class="b-pdp-product-info"]/div/div[@class="b-pdp-product-info-main-right"]/div[@class="b-pdp-prices"]/div/div/div/div[@class="b-price-section"]/span[2]/span/text()')
                    if len(price) ==0:
                        value_MRP = tree.xpath(
                            '//div[@class="b-pdp-product-info"]/div/div[@class="b-pdp-product-info-main-right"]/div[@class="b-pdp-prices"]/div/div/div/div[@class="b-price-section"]/span[1]/span/text()')[
                            0].replace('€', '').replace('\n', '')
                        details_dict['List_Price'] = value_MRP
                        details_dict['Promo_Price'] = value_MRP
                        details_dict['Discount'] = '-'
                        print(value_MRP)

                    else:
                        value_MRP = tree.xpath(
                            '//div[@class="b-pdp-product-info"]/div/div[@class="b-pdp-product-info-main-right"]/div[@class="b-pdp-prices"]/div/div/div/div[@class="b-price-section"]/span[1]/span/text()')[
                            0].replace('€', '').replace('\n', '')
                        details_dict['List_Price'] = value_MRP
                        promo = tree.xpath(
                            '//div[@class="b-pdp-product-info"]/div/div[@class="b-pdp-product-info-main-right"]/div[@class="b-pdp-prices"]/div/div/div/div[@class="b-price-section"]/span[2]/span/text()')[
                            0].replace('€', '').replace('\n', '')
                        details_dict['Promo_Price'] = promo
                        disco = tree.xpath(
                            '//div[@class="b-pdp-product-info"]/div/div[@class="b-pdp-product-info-main-right"]/div[@class="b-pdp-prices"]/div/div/div/div[@class="b-price-discount"]/span/text()')[
                            0].replace('-', '').replace('\n', '')
                        details_dict['Discount'] = disco
                        print(value_MRP, promo, disco)


                except:
                    details_dict['List_Price'] = '-'
                    details_dict['Promo_Price'] = '-'
                    details_dict['Discount'] = '-'

                try:
                    img_url=tree.xpath('//div[@class="b-pdp-carousel-item"]/picture/img/@src')
                    print(img_url)
                    details_dict['Image_URLs']='|'.join(i for i in img_url)
                except:
                    details_dict['Image_URLs'] = '-'


                try:

                    col_grp=tree.xpath('//div[@class="b-pdp-color-carousel-item"]/a/@data-value')
                    print(col_grp)
                    if len(col_grp)>1:
                        details_dict['Colour_Grouping']='|'.join(i for i in col_grp)
                    else:
                        details_dict['Colour_Grouping']='-'
                except:
                    details_dict['Colour_Grouping']='-'

                try:
                    disc=tree.xpath('//div[@class="b-details-content"]/p/text()')
                    if disc[0]!='\xa0':
                        if len(disc)>1:
                            details_dict['Description']=' adidas '.join(i for i in disc)
                        else:
                            details_dict['Description']=disc[0]
                except:
                    details_dict['Description']='-'


                # print(disc)
                try:

                    specs_tree=tree.xpath('//tr[@class="b-fact-attr js-fact-attr"]')
                    # print(specs_tree)
                    s_list=[]
                    for specs in specs_tree:
                        key=specs.xpath('.//td[@class="b-fact-label"]/text()')

                        value=specs.xpath('.//td[@class="b-fact-value js-fact-value"]/text()')
                        s_list.append(key[0]+value[0])
                    # print(len(specs_tree))
                    print(s_list)
                    details_dict['Specification']='|'.join(i for i in s_list)

                    prod_id=''.join(ele.split(':')[-1] for ele in s_list if 'Art.-Nr' in ele)  # if this logic fails directly takke input as prod id
                    details_dict['Product_ID']=prod_id
                    details_dict['RPC']=prod_id

                    # details_dict['INPUT_PLATFORMID']=prod_id
                    details_dict['MPC']=''.join(ele.split(':')[-1] for ele in s_list if 'Hersteller Nr' in ele)
                except:
                    pass

                details_dict['Cache_Page_Link']=cache_page_l

                try:
                    varient_tree= tree.xpath('//div[@class="b-swatch-value-wrapper"]')
                    if len(varient_tree)>0:
                        for vari in varient_tree:
                            details_dict['Variant']=vari.xpath('.//a/span/@data-attr-value')[0]

                            stock_data=vari.xpath('.//a/span/@class')[0]
                            # print(stock_data)

                            if 'orderable' in stock_data :
                                details_dict['Stock_Condition']='In Stock'
                            else:
                                details_dict['Stock_Condition']='Out of Stock'

                            v_id=vari.xpath('.//a/@data-variant-id')[0]
                            # print(v_id)

                            details_dict['Variant_ID']='"{}"'.format(v_id) #.replace('"','')

                            details_dict['Reason_Code']='Success-PF'

                            store_data_by_product_id(dict=[details_dict])
                            print(details_dict)

                    else:
                        details_dict['Variant_ID'] = '-'
                        details_dict['Stock_Condition'] = 'In Stock'
                        details_dict['Variant'] = 'one size'
                        details_dict['Reason_Code'] = 'Success-PF'
                        print(details_dict)
                        store_data_by_product_id(dict=[details_dict])

                except:
                    details_dict['Variant_ID'] ='-'
                    details_dict['Stock_Condition'] = 'In Stock'
                    details_dict['Variant'] ='one size'
                    details_dict['Reason_Code'] = 'Success-PF'
                    print(details_dict)
                    store_data_by_product_id(dict=[details_dict])

            except:
                if n != 0:
                    n = n - 1
                    by_product_id(prod_id_data, n)
                else:
                    details_dict['Reason_Code'] = 'Success-PNF'
        elif n!=0:
            n = n - 1
            by_product_id(prod_id_data,n)
            print(prod_id_data)
            #store_data_by_product_id(dict=[details_dict])

        elif response.status_code==403:
            details_dict['Reason_Code']='Blocked'
            print(details_dict)
            #store_data_by_product_id(dict=[details_dict])

        else:
            details_dict['Reason_Code']= 'Success-PNF'
            print(details_dict)
            #store_data_by_product_id(dict=[details_dict])

    except:
        if n != 0:
            n = n - 1
            by_product_id(prod_id_data, n)
        else:
            details_dict['Reason_Code'] = 'TimeOut'
            print(details_dict)
            #store_data_by_product_id(dict=[details_dict])


count=0
for data in extract_xl():
    print(data)
    count+=1
    products_df=by_product_id(prod_id_data=data,n=6)
    print(count)


