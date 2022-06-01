
import re
import random
import general
from lxml import html
import time
# import xlrd
import requests
from lxml import html
import time
import json
from datetime import datetime, timedelta
import datetime
import openpyxl
p1 = ['154.28.67.106','154.28.67.111','154.28.67.116','154.28.67.117','154.28.67.125','154.28.67.131','154.28.67.133','154.28.67.142','154.28.67.156','154.28.67.163','154.28.67.173','154.28.67.18','154.28.67.182','154.28.67.184','154.28.67.20','154.28.67.200','154.28.67.210','154.28.67.218','154.28.67.222','154.28.67.223','154.28.67.231','154.28.67.240','154.28.67.243','154.28.67.253','154.28.67.39','154.28.67.4','154.28.67.49','154.28.67.5','154.28.67.61','154.28.67.80','154.28.67.81','154.28.67.87','154.28.67.88','154.28.67.96','154.28.67.99','154.7.230.100','154.7.230.101','154.7.230.103','154.7.230.107','154.7.230.109','154.7.230.130','154.7.230.132','154.7.230.14','154.7.230.140','154.7.230.147','154.7.230.151','154.7.230.156','154.7.230.163','154.7.230.170','154.7.230.18','154.7.230.183','154.7.230.188','154.7.230.189','154.7.230.19','154.7.230.190','154.7.230.198','154.7.230.204','154.7.230.209','154.7.230.235','154.7.230.238','154.7.230.246','154.7.230.29','154.7.230.41','154.7.230.42','154.7.230.51','154.7.230.55','154.7.230.60','154.7.230.61','154.7.230.74','154.7.230.82','154.7.230.89','23.131.8.112','23.131.8.115','23.131.8.117','23.131.8.12','23.131.8.121','23.131.8.124','23.131.8.150','23.131.8.161','23.131.8.166','23.131.8.171','23.131.8.173','23.131.8.176','23.131.8.177','23.131.8.181','23.131.8.19','23.131.8.192','23.131.8.194','23.131.8.199','23.131.8.202','23.131.8.203','23.131.8.204','23.131.8.207','23.131.8.209','23.131.8.213','23.131.8.216','23.131.8.225','23.131.8.228','23.131.8.231','23.131.8.238','23.131.8.254','23.131.8.36','23.131.8.5','23.131.8.76','23.131.8.93','23.131.8.95','23.131.8.99','23.131.88.105','23.131.88.12','23.131.88.137','23.131.88.139','23.131.88.140','23.131.88.145','23.131.88.150','23.131.88.151','23.131.88.153','23.131.88.154','23.131.88.156','23.131.88.165','23.131.88.18','23.131.88.191','23.131.88.192','23.131.88.194','23.131.88.198','23.131.88.202','23.131.88.206','23.131.88.220','23.131.88.223','23.131.88.228','23.131.88.233','23.131.88.24','23.131.88.242','23.131.88.244','23.131.88.47','23.131.88.63','23.131.88.67','23.131.88.73','23.131.88.80','23.131.88.81','23.131.88.82','23.131.88.88','23.131.88.97','23.170.144.149','23.170.144.209','23.170.144.212','23.170.144.242','23.170.144.83','23.170.145.117','23.170.145.167','23.170.145.182','23.170.145.19','23.170.145.203','23.226.17.101','23.226.17.109','23.226.17.112','23.226.17.113','23.226.17.115','23.226.17.123','23.226.17.129','23.226.17.143','23.226.17.148','23.226.17.165','23.226.17.186','23.226.17.199','23.226.17.201','23.226.17.207','23.226.17.210','23.226.17.219','23.226.17.220','23.226.17.222','23.226.17.229','23.226.17.250','23.226.17.254','23.226.17.26','23.226.17.33','23.226.17.4','23.226.17.49','23.226.17.5','23.226.17.55','23.226.17.66','23.226.17.7','23.226.17.72','23.226.17.78','23.226.17.8','23.226.17.86','23.226.17.90','23.226.17.93','23.230.177.105','23.230.177.110','23.230.177.113','23.230.177.121','23.230.177.130','23.230.177.14','23.230.177.143','23.230.177.15','23.230.177.150','23.230.177.154','23.230.177.165','23.230.177.173','23.230.177.191','23.230.177.196','23.230.177.203','23.230.177.206','23.230.177.208','23.230.177.217','23.230.177.220','23.230.177.221','23.230.177.224','23.230.177.228','23.230.177.231','23.230.177.235','23.230.177.237','23.230.177.241','23.230.177.27','23.230.177.38','23.230.177.52','23.230.177.61','23.230.177.67','23.230.177.72','23.230.177.80','23.230.177.88','23.230.177.94','23.230.177.99','23.230.197.103','23.230.197.106','23.230.197.109','23.230.197.11','23.230.197.12','23.230.197.122','23.230.197.124','23.230.197.146','23.230.197.155','23.230.197.156','23.230.197.174','23.230.197.179','23.230.197.181','23.230.197.196','23.230.197.2','23.230.197.201','23.230.197.207','23.230.197.208','23.230.197.225','23.230.197.227','23.230.197.233','23.230.197.236','23.230.197.239','23.230.197.240','23.230.197.244','23.230.197.251','23.230.197.50','23.230.197.52','23.230.197.54','23.230.197.60','23.230.197.71','23.230.197.80','23.230.197.81','23.230.197.84','23.230.197.97','23.230.74.102','23.230.74.110','23.230.74.116','23.230.74.125','23.230.74.133','23.230.74.135','23.230.74.14','23.230.74.141','23.230.74.149','23.230.74.15','23.230.74.157','23.230.74.16','23.230.74.170','23.230.74.172','23.230.74.174','23.230.74.183','23.230.74.187','23.230.74.19','23.230.74.198','23.230.74.208','23.230.74.212','23.230.74.215','23.230.74.23','23.230.74.230','23.230.74.231','23.230.74.252','23.230.74.30','23.230.74.41','23.230.74.57','23.230.74.58','23.230.74.59','23.230.74.6','23.230.74.75','23.230.74.81','23.230.74.88','23.230.74.91','23.27.222.108','23.27.222.109','23.27.222.134','23.27.222.138','23.27.222.159','23.27.222.161','23.27.222.164','23.27.222.166','23.27.222.178','23.27.222.19','23.27.222.195','23.27.222.201','23.27.222.202','23.27.222.203','23.27.222.208','23.27.222.21','23.27.222.211','23.27.222.218','23.27.222.223','23.27.222.228','23.27.222.234','23.27.222.236','23.27.222.242','23.27.222.251','23.27.222.253','23.27.222.34','23.27.222.61','23.27.222.62','23.27.222.69','23.27.222.70','23.27.222.72','23.27.222.73','23.27.222.74','23.27.222.81','23.27.222.93','38.131.131.110','38.131.131.114','38.131.131.123','38.131.131.125','38.131.131.137','38.131.131.142','38.131.131.145','38.131.131.147','38.131.131.15','38.131.131.154','38.131.131.16','38.131.131.17','38.131.131.173','38.131.131.18','38.131.131.193','38.131.131.204','38.131.131.207','38.131.131.227','38.131.131.229','38.131.131.233','38.131.131.238','38.131.131.246','38.131.131.248','38.131.131.250','38.131.131.31','38.131.131.36','38.131.131.50','38.131.131.58','38.131.131.64','38.131.131.70','38.131.131.71','38.131.131.74','38.131.131.83','38.131.131.94','38.131.131.99','38.75.75.104','38.75.75.111','38.75.75.112','38.75.75.119','38.75.75.120','38.75.75.123','38.75.75.127','38.75.75.139','38.75.75.14','38.75.75.143','38.75.75.155','38.75.75.156','38.75.75.158','38.75.75.170','38.75.75.179','38.75.75.188','38.75.75.2','38.75.75.201','38.75.75.231','38.75.75.232','38.75.75.241','38.75.75.246','38.75.75.251','38.75.75.26','38.75.75.29','38.75.75.4','38.75.75.44','38.75.75.49','38.75.75.56','38.75.75.58','38.75.75.62','38.75.75.72','38.75.75.76','38.75.75.79','38.75.75.88','38.96.156.108','38.96.156.112','38.96.156.128','38.96.156.131','38.96.156.14','38.96.156.142','38.96.156.143','38.96.156.149','38.96.156.16','38.96.156.163','38.96.156.165','38.96.156.169','38.96.156.186','38.96.156.188','38.96.156.190','38.96.156.192','38.96.156.194','38.96.156.199','38.96.156.218','38.96.156.236','38.96.156.240','38.96.156.252','38.96.156.28','38.96.156.32','38.96.156.35','38.96.156.56','38.96.156.57','38.96.156.6','38.96.156.67','38.96.156.77','38.96.156.80','38.96.156.83','38.96.156.84','38.96.156.89','38.96.156.92','45.238.157.100','45.238.157.104','45.238.157.106','45.238.157.110','45.238.157.116','45.238.157.118','45.238.157.119','45.238.157.12','45.238.157.123','45.238.157.132','45.238.157.14','45.238.157.149','45.238.157.15','45.238.157.183','45.238.157.186','45.238.157.189','45.238.157.2','45.238.157.212','45.238.157.214','45.238.157.217','45.238.157.22','45.238.157.228','45.238.157.23','45.238.157.247','45.238.157.43','45.238.157.48','45.238.157.51','45.238.157.52','45.238.157.53','45.238.157.56','45.238.157.61','45.238.157.65','45.238.157.72','45.238.157.79','45.238.157.8','45.238.159.103','45.238.159.107','45.238.159.110','45.238.159.114','45.238.159.116','45.238.159.123','45.238.159.126','45.238.159.144','45.238.159.148','45.238.159.15','45.238.159.156','45.238.159.165','45.238.159.167','45.238.159.183','45.238.159.20','45.238.159.208','45.238.159.217','45.238.159.220','45.238.159.23','45.238.159.230','45.238.159.235','45.238.159.237','45.238.159.238','45.238.159.24','45.238.159.249','45.238.159.251','45.238.159.32','45.238.159.34','45.238.159.51','45.238.159.6','45.238.159.66','45.238.159.77','45.238.159.79','45.238.159.82','45.238.159.91']
header={'authority': 'www.asos.com',
'method': 'GET',
'path': '/es/hombre/',
'scheme': 'https',
'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
'accept-encoding': 'gzip, deflate, br',
'accept-language': 'en-US,en;q=0.9',
'cookie': 'geocountry=IN; bm_sz=AD29DD6D020324D241B36866A347D202~YAAQfqUFF0lJF2R/AQAA0D26cw9E6WvDikamWlNzORbhhxTIZ0ag6b78JtGPIY36lbYE9fXoeLIgMBnAsAPDFw0SeOtkwCYmQaDrJKgMC6SpwiEgR5n7O6MVdvBVKOPVVLuhp6SFtm3KZCOytLIIX+uElaXRk/bQP9yM02E1zFGQibFz8v8g6mGjSDd8Hb+VySx31SRe//QQJ8zQ2Bd9ejaRn6xGrkCFLcD3SFzU8PzZGW9TliMgxzYX4iU6lu6uz4kEoOKmF9PeZAlp0ygp+aVbdsf6HQEQG+jLmjbK3ZR3~3683398~3356976; s_ecid=MCMID%7C72534221667583640275321337621776600116; AMCVS_C0137F6A52DEAFCC0A490D4C%40AdobeOrg=1; browseCountry=ES; storeCode=ES; browseSizeSchema=ES; browseLanguage=es-ES; browseCurrency=EUR; currency=19; floor=1000; plp_columsCount=fourColumns; featuresId=a05a9a8b-3090-4a3c-bb68-e14331fad640; asos-b-sdv629=8o580s6-31; asos=PreferredSite=&currencyid=19&currencylabel=EUR&topcatid=1000&customerguid=25512f5dc75d4d59b662ec0ba5d3b7ad; asos-perx=25512f5dc75d4d59b662ec0ba5d3b7ad||ce4eeeb0be3440c58261f431b3f09fa2; s_cc=true; AMCV_C0137F6A52DEAFCC0A490D4C%40AdobeOrg=-1303530583%7CMCMID%7C72534221667583640275321337621776600116%7CMCAID%7CNONE%7CMCOPTOUT-1646921319s%7CNONE%7CMCAAMLH-1647518935%7C6%7CMCAAMB-1647518935%7Cj8Odv6LonN4r3an7LhD3WZrU1bUpAkFkkiY1ncBR96t2PTI%7CMCCIDH%7C0%7CvVersion%7C3.3.0; _cs_c=0; _gcl_au=1.1.929088223.1646914137; _gid=GA1.2.397163914.1646914137; __gads=ID=1b9cedda098806d1:T=1646914175:S=ALNI_MbSTpODp0J8UKy4QhjWgZM5A2TIcQ; optimizelyEndUserId=oeu1646914190330r0.3828388980611097; stc-welcome-message=resolvedDeliveryCountry=IN&cappedPageCount=2; _abck=FAE5B40CC94C2D32B0A26BAD453340C8~0~YAAQfqUFF6ZMF2R/AQAAUEy8cwf8P8/HxGQ0I5FeYvzF6J8qL/Ue1CuTdsLwxZaxue79LHEze9DIhsQwaYVDSjQisgqthyi1cn3DxflSj+z8XPa97P2CZuPYCaP/g11yBuaaOgN1yqEEA/JdaLp/iJokWjMGkUikq16tzucNkRLF+fv/SqTcB4mZkBKT7nWPtDcXRRAwJG4p1wIoP60AMgwPVjs4S7favm1NWgCbnxC41jwDfh0CXnavtFcKqm1JD6XQuskfiDrO08wnYlw6CKykBLMg5EDs9HgmIAUqGbhaUHuu/EoH+MhwWOHXyQ27qLurXe9N+IwCzAQBQ71zgTAxdD3N8OkwU8BvwGqH2hSnXNXhp9tCViB6B3YgD0N+/yP3XG8IV3JlY2hTXjPappyZwnJ4tQ==~-1~-1~-1; fita.sid.asos=RbWjpMyMKeQLip3AXo4JT4qrCjeF1QZ8; _s_fpv=true; siteChromeVersion=au=11&com=11&de=11&dk=11&es=11&fr=11&it=11&nl=11&pl=11&roe=11&row=11&ru=11&se=11&us=11; keyStoreDataversion=8o580s6-31; bm_mi=733246C390554DFE2ABC0E7502D004A7~1rYU3V3iF9X6mcsmdgaHV5CboPfY+62BruE7d768fBPzpLnVFN9yVcCMcuEW4oBdYb2+yzGGPSIAXJPQ6JOA3bEqHKrQ8rjpOznvUkS9snfxvx4uTAEfPwGb5052lVEfig8PY6pNj813cxQ22OR/xAz1EXSj3isV7OmSPX/SQS/ZrBzn2LalNznrYQBo7zy6OTBSLw00vjZ1CwYtVhApZM2YyaU6LcPAnt7TYipqsADYKi3Re39rXfhbks9WE3aegPukqtV5TyPjeZhffy1yaE8JcZEP4UfWiyHPquwtTEa2CNxrxzqW5PeNNkX0Aawu; ak_bmsc=44C7C8290045E5DE41D92F9A72A767C6~000000000000000000000000000000~YAAQFVx6XJJxOmR/AQAA2vP7cw+pJV6mXzXTXg+YICUWmHIP2qRfmaXg9BSXjCZgI9HtqDUdzBj2ZsjIqriV7jEcflSjtDOhqH4RjYw9ecHps9P6/+WzfDa+VSY7zkDxdcRNdbcTrbouczlV7Kbz8I69NtUjEEaatTKdPfqdLZY6tuZb2NxQk5aeyqA7lH+ssBvDh6ioqtcKKOGrXNfLXDJYuyXwDt9uFNWqt48e8hA2iEe2hFI7etV82ZgW9rfMqmkFf6MPBH9wmXEE7FP0vsTvyNULcQ/FcKjleJm9gZlcczzcdb8n9S9ONbZrgLxu8m7BxwDkeF7FX4rw3Oq+7sOLtrcsbzA+WLWtvBtODB+vZRn0U1WSMy7kKOa7sUvkkWaVmw4RbbOGSv5v7jsGpmViHYnPH7RT6kgNzge4i0cPN5a2WHiEZ0Iw9Jv22Jul3NvZBMWwumAN; bt_recUser=0; s_pers=%20s_vnum%3D1648751400909%2526vn%253D2%7C1648751400909%3B%20eVar225%3D5%7C1646920229859%3B%20visitCount%3D2%7C1646920229863%3B%20s_invisit%3Dtrue%7C1646920231077%3B%20s_nr%3D1646918431081-Repeat%7C1678454431081%3B%20gpv_p10%3Ddesktop%2520es%257Cfloor%2520page%257Cwomen%7C1646920231083%3B%20gpv_p6%3D%2520%7C1646920231086%3B%20gpv_e47%3Dwomen%257Chome%7C1646920231088%3B; OptanonConsent=isIABGlobal=false&datestamp=Thu+Mar+10+2022+18%3A50%3A33+GMT%2B0530+(India+Standard+Time)&version=6.30.0&hosts=&landingPath=NotLandingPage&groups=C0001%3A1%2CC0003%3A1%2CC0004%3A1&AwaitingReconsent=false; _cs_id=26e6855f-d34e-a65a-ca05-5e378390c07a.1646914135.2.1646918433.1646917858.1628755191.1681078135783; _cs_s=5.0.0.1646920233998; _gat=1; _ga_54TNE49WS4=GS1.1.1646918262.2.1.1646918434.55; _ga=GA1.1.1591537829.1646914137; _ga_1JR0QCFRSY=GS1.1.1646918262.2.1.1646918434.0; s_sq=asoscomprod%3D%2526c.%2526a.%2526activitymap.%2526page%253Ddesktop%252520es%25257Cfloor%252520page%25257Cwomen%2526link%253DHOMBRE%2526region%253Dchrome-sticky-header%2526pageIDType%253D1%2526.activitymap%2526.a%2526.c; RT="z=1&dm=asos.com&si=7f57c6e4-e640-4f12-a5aa-0be66c7bb4a0&ss=l0l0mqq8&sl=4&tt=f4u&bcn=%2F%2F02179915.akstat.io%2F&ld=2xs0&nu=20sxaohj&cl=446i&ul=4473"',
'referer': 'https://www.asos.com/es/mujer/',
'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Google Chrome";v="99"',
'sec-ch-ua-mobile': '?0',
'sec-ch-ua-platform': "Windows",
'sec-fetch-dest': 'document',
'sec-fetch-mode': 'navigate',
'sec-fetch-site': 'same-origin',
'sec-fetch-user': '?1',
'upgrade-insecure-requests': '1',
'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'}


file_path='ass-1.xlsx'
# file_path='Pending-1.xlsx'
def extract_xl():
    workbook= openpyxl.load_workbook(file_path)
    sheet_obj= workbook.active
    all_rows=sheet_obj.max_row
    inputdata=[]
    for i in range(1,all_rows):
        limit=sheet_obj.cell(row=i+1,column=1)
        urls= sheet_obj.cell(row=i+1,column=2)

        u_data=[urls.value,limit.value]

        inputdata.append(u_data)
    # print(inputdata)
    return inputdata

from csv import DictWriter
def store_data_by_product_id(dict):
    fieldnames = ['SKU_ID', 'Website', 'Country','RPC','MPC','Product_ID', 'Product_URL',
                  'Product_Name', 'Category_Path', 'Specification', 'Description', 'Currency', 'List_Price',
                  'Promo_Price', 'Discount', 'Brand', 'Rating_Count', 'Review_Count', 'Image_URLs', 'Variant',
                  'Variant_ID', 'Colour_of_Variant', 'Colour_Grouping', 'Seller_Name', 'Stock_Count', 'Stock_Condition',
                  'Stock_Message', 'Sustainability_Badge', 'Reason_Code', 'Crawling_TimeStamp', 'Cache_Page_Link',
                  'Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5']
    with open('output1-2503.csv', 'a+', encoding='UTF-8-sig', newline='') as file:
        writer = DictWriter(file, fieldnames=fieldnames)
        # writer.writeheader()
        writer.writerows(dict)
        file.close()

def scraping_inputs(recursive_limit=2):
    try:
        resp=requests.get('http://10.100.22.203/get_data',params={'type_id': [2],'domain_id':[16]},timeout=30)
        print(resp.status_code)
        if resp.status_code==200:
            print(resp.text)
            return resp.text
        elif recursive_limit != 0:
            new_n = recursive_limit - 1
            scraping_inputs(recursive_limit=new_n)
        elif resp.status_code == 408:
            print('TimeOut')
        else:
            print('Blocked')
    except:
        print('data not came')
# scraping_inputs()


#api-2
import requests
def status_update(status_dict,recursive_limit=2):
    try:
        dict_values=status_dict
        response=requests.post('http://10.100.22.203/post_data',json=[dict_values] )
        if response.status_code==200:
            print(response.text)
            return response.text
        elif recursive_limit != 0:
            new_n = recursive_limit - 1
            status_update(status_dict,recursive_limit=new_n)
        elif response.status_code == 408:
            print('TimeOut')
        else:
            print('page_not_found')
    except:
        print('status not updated')

import json
import requests
def save_data_in_database_son(data,n=2):
    try:
        # http://10.100.22.203/insert_son
        resp=requests.post('http://10.100.22.203/insert_son',json=[data],timeout=20 )
        print(resp.status_code)
        if resp.status_code==200:
            return resp.text
        elif n != 0:
            new_n = n - 1
            save_data_in_database_son(data,n=new_n)
        elif resp.status_code == 408:
            print('TimeOut')
        else:
            print('Blocked')
    except:
        print('not saved data in database')



import datetime
import os
def html_cache_page_saving(sku,saving_data):
    datazone = datetime.datetime.now()
    f_date = datazone.strftime("%d_%m_%Y")
    strdate = datazone.day
    strm = datazone.month
    stry = datazone.year
    pageid = sku
    cpid = pageid + '_' + str(strdate) + '_' + str(strm) + '_' + str(stry)

    ASS_folder = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\ASOS_ES\\PDP"
    sos_date_wise_folder = ASS_folder + f"\\{f_date}"
    if os.path.exists(sos_date_wise_folder):
        pass
    else:
        os.mkdir(sos_date_wise_folder)
    sos_filename = sos_date_wise_folder + "\\" + cpid + ".html"
    sos_filename = sos_filename.replace("+", "_").replace("-", "_")
    page_path = sos_filename.replace('/', '')
    print(page_path)
    page_path = page_path.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\',
                                  'https:////ecxus440.eclerx.com//cachepages//').replace('\\',
                                                                                         '//').replace(
        '//', '/')
    print(page_path)
    if os.path.exists(sos_filename):
        with open(sos_filename, 'w', encoding='utf-8') as f:
            f.write(saving_data)
    else:
        with open(sos_filename, 'w', encoding='utf-8') as f:
            f.write(saving_data)
    return page_path


def listToString(s):
    # initialize an empty string
    str1 = " "

    # return string
    return (str1.join(s))


def join_string1(list_string):
    # Join the string based on '-' delimiter
    string = '|'.join(list_string)

    return string


from adidas_proxy import AdidasProxy
ad_pr=AdidasProxy()

dom='http://www.snipes.com'
def by_product_id(prod_id_data,n):
    product_id_url=prod_id_data[0]
    print(product_id_url)
    # pro_id=product_id_url.split('-')[-1]
    # print(product_id_url)
    INPUT_url = prod_id_data[0]
    # INPUT_url = 'https://www.asos.com/es/reebok/shorts-negros-con-diseno-2-en-1-workout-ready-de-reebok-training/prd/200957937?clr=negro&colourWayId=200957938&SearchQuery=reebok'
    print(INPUT_url)
    SKUID = prod_id_data[1]

    skuId=prod_id_data[1]

    import datetime
    time = datetime.datetime.now()
    details_dict = {'SKU_ID': skuId,
                    'Website': 'ASOS',
                    'Country': 'ES',
                    'RPC': '-',
                    'MPC': 'Not Available',
                    'Product_ID': '-',
                    'Product_URL': product_id_url,
                    'Product_Name': '-',
                    'Category_Path': '-',
                    'Specification': '-',
                    'Description': '-',
                    'Currency': 'EURO',
                    'List_Price': '-',
                    'Promo_Price': '-',
                    'Discount': '-',
                    'Brand': '-',
                    'Rating_Count': '-',
                    'Review_Count': '-',
                    'Image_URLs': '-',
                    'Variant': '-',
                    'Variant_ID': 'Not Available',
                    'Colour_of_Variant': '-',
                    'Colour_Grouping':'Not Available',
                    'Seller_Name': 'Not Available',
                    'Stock_Count': 'Not Available',
                    'Stock_Condition': '-',
                    'Stock_Message': 'Not Available',
                    'Sustainability_Badge': 'Not Available',
                    'Reason_Code': '-',
                    'Crawling_TimeStamp': time.strftime("%Y-%m-%dT%H:%M:%SZ"),
                    'Cache_Page_Link': '-',
                    'Extra1': '-',
                    'Extra2': '-',
                    'Extra3': '-',
                    'Extra4': '-',
                    'Extra5': '-'
                    }
    status_dict = {"SubRequestId": 2,
                   "RequestId": 3,
                   "RequestRunId": 5,
                   "StatusId": 0}
    try:
        # response = requests.get(product_id_url, headers=header,proxies=ad_pr.random_proxy(domain=dom) ,timeout=30)
        # print(response.status_code)
        #
        # cache_data = response.text
        # cache_page_l = html_cache_page_saving(sku=skuId, saving_data=cache_data)
        for _ in range(18):
            p_auth = str("csimonra:h19VA2xZ")
            p_host = random.choice(p1)
            p_port = "29842"
            proxy = {
                'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
            }
            response = requests.get(INPUT_url, headers=header,proxies=proxy, timeout=30)
            if response.status_code == 200:
                tree = html.fromstring(response.content)
                break
        cache_data = response.text

        html_file_path = html_cache_page_saving(sku=str(prod_id_data[1]),saving_data=cache_data)

        if response.status_code == 200:
            reason_Code = "SucessPF"
            print('status code', response.status_code)


            p_data = ''.join(re.findall(r'window.asos.pdp.config.product = (.*?);', response.text))
            if not p_data:
                print(INPUT_url, 'not found')

            try:
                product_json = json.loads(p_data)
                # print(product_json)
                details_dict['Cache_Page_Link']=html_file_path

                Product_name = product_json['name']

                details_dict['Product_Name']=general.clean(Product_name)

                product_brand = product_json['brandName']
                details_dict['Brand']=general.clean(product_brand)

                cat = tree.xpath('//*[@aria-label="breadcrumbs"]//li//text()')
                if cat == []:
                    cat = tree.xpath('//*[@id="chrome-breadcrumb"]//li//text()')
                Category_Path = "".join([i.strip() for i in cat if i.strip()]).strip()
                print(Category_Path)
                details_dict['Category_Path']=general.clean(Category_Path)

                image = []
                img = len(product_json['images'])
                for im in range(0, img):
                    imgs = product_json['images'][im]['url']
                    image.append(imgs)
                Images = '|'.join(image)
                # print(Images)
                details_dict['Image_URLs']=Images

                try:
                    color_of_variant = product_json['variants'][0]['colour']
                    print(color_of_variant)
                    details_dict['Colour_of_Variant']=general.clean(color_of_variant)
                except:
                    try:
                        color_of_variant = product_json['products'][0]['variants'][0]['colour']
                        print(color_of_variant)
                        details_dict['Colour_of_Variant'] = general.clean(color_of_variant)
                    except Exception as e:
                        details_dict['Colour_of_Variant'] ='-'

                try:
                    Description = '|'.join(tree.xpath('//*[@class="product-description"]//ul/li//text()'))
                    print(Description)
                    details_dict['Description']=general.clean(Description)
                except:
                    details_dict['Description']='-'



                ####        specification
                # Specifications = '|'.join(tree.xpath('//*[@class="about-me"]/p/text()'))
                #
                # spe = Specifications if Specifications else '|'.join(tree.xpath('//*[@class="about-me"]/text()[preceding-sibling::br]'))

                specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/text()')
                xyz = listToString(specifications)
                if ":" in xyz:
                    for abc in specifications:
                        if ':' in abc:
                            specifications = abc.strip(' ').strip('\n').strip('.')
                else:
                    specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/p/text()')
                    xyz = listToString(specifications)
                    if ":" in xyz:
                        for abc in specifications:
                            if ':' in abc:
                                specifications = abc.strip(' ').strip('\n').strip('.')
                    else:
                        specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/div/text()')
                        xyz = listToString(specifications)
                        if ":" in xyz:
                            for abc in specifications:
                                if ":" in abc:
                                    specifications = abc.strip(' ').strip('\n').strip('.')
                        else:
                            specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/p/span/text()')
                            xyz = listToString(specifications)
                            if ':' in xyz:
                                for abc in specifications:
                                    if ":" in abc:
                                        specifications = abc.strip(' ').strip('\n').strip('.')
                            else:
                                specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/div/div/text()')
                                xyz = listToString(specifications)
                                if ':' in xyz:
                                    for abc in specifications:
                                        if ':' in abc:
                                            specifications = abc.strip(' ').strip('\n').strip('.')
                                else:
                                    specifications = tree.xpath(
                                        '//div[@class="col"]/div[@class="about-me"]/span/text()')
                                    xyz = listToString(specifications)
                                    if ':' in xyz:
                                        for abc in specifications:
                                            if ':' in abc:
                                                specifications = abc.strip(' ').strip('\n').strip('.')
                                    else:
                                        specifications = tree.xpath(
                                            '//div[@class="col"]/div[@class="about-me"]/div/span/text()')
                                        xyz = listToString(specifications)
                                        if ':' in xyz:
                                            for abc in specifications:
                                                if ':' in abc:
                                                    specifications = abc.strip(' ').strip('\n').strip('.')

                if len(specifications) == 0:
                    specifications = '-'
                    details_dict['Specification'] = specifications

                if specifications != '-':
                    specifications_value = specifications.split('| ')
                    specifications_bool = specifications_value[-1][0].isdigit()
                    if specifications_bool == True:
                        specifications_value[-2] = specifications_value[-2] + ", " + specifications_value[-1]
                        specifications_value.pop()
                        specifications = specifications_value
                        specifications = join_string1(specifications)

                specifications = general.clean(specifications)

                details_dict['Specification']=specifications

                try:
                    Product_id = product_json['productCode']
                except Exception as e:
                    Product_id = product_json['id']

                details_dict['Product_ID']=Product_id
                details_dict['RPC']=Product_id

                product_url = tree.xpath('//*[@rel="canonical"]/@href')[0]

                try:
                    riview_data = ''.join(re.findall(r'window.asos.pdp.config.ratings = (.*?);', response.text))
                    review_json = json.loads(riview_data)

                    try:
                        Ratings = review_json['averageOverallRating']
                        details_dict['Rating_Count']=Ratings
                    except:
                        Ratings = '-'
                        details_dict['Rating_Count'] = Ratings
                    try:
                        Review = review_json['totalReviewCount']
                        details_dict['Review_Count']=Review
                    except:
                        Review = '-'
                        details_dict['Review_Count'] = Review
                except Exception as e:
                    riview_data = ''.join(re.findall(r'window.asos.pdp.config.ratings = (.*?);', response.text))
                    try:
                        rr = re.findall(r'totalReviewCount":(.*?),"averageOverallRating":(.*?),"averageOverallStar',riview_data)
                        details_dict['Rating_Count']=rr[0][-1]
                        details_dict['Review_Count']=rr[0][0]
                    except Exception as e:
                        details_dict['Rating_Count'] = '-'
                        details_dict['Review_Count'] = '-'
                # print(Ratings, Review)

                details_dict['Reason_Code']='Success-PF'
                try:
                    product_id_specific = product_json['images'][0]['productId']
                    print(product_id_specific)
                    price_link = f'https://www.asos.com/api/product/catalogue/v3/stockprice?productIds={product_id_specific}&store=ES&currency=EUR&keyStoreDataversion=hgk0y12-29'

                    response = requests.request("GET", price_link, headers=header,proxies=proxy,timeout=30)
                    data = response.text
                    price_json = json.loads(data)
                    print(price_json)
                    List_price = price_json[0]['productPrice']['previous']['text']
                    Promo_price = price_json[0]['productPrice']['current']['text']

                    a = List_price.replace('€', '')
                    details_dict['List_Price']=a

                    b = Promo_price.replace('€', '')
                    details_dict['Promo_Price']=b
                    print(a, b)
                except:
                    details_dict['List_Price']='-'
                    details_dict['Promo_Price']='-'

                try:
                    disurlfr = f'https://www.asos.com/api/product/catalogue/v3/stockprice?productIds={product_id_specific}&store=DE&currency=EUR&keyStoreDataversion=hgk0y12-29'
                    response = requests.request("GET", disurlfr, headers=header,proxies=proxy,timeout=30)
                    data = response.text
                    dis_json = json.loads(data)
                    # print(dis_json)
                    z = dis_json[0]['discountPercentage']

                    dis = (str(z) + '%')
                    details_dict['Discount']=dis

                except:
                    details_dict['Discount']='-'

                try:
                    variant = []
                    var = len(product_json['variants'])
                    for v in range(0, var):
                        vars = product_json['variants'][v]['size']
                        variant.append(vars)
                    variants = '|'.join(variant)
                    varientSize = variants.split('|')
                    print(variant)

                    variant = []
                    ss = product_json['variants']
                    stockvar = []
                    var = len(product_json['variants'])
                    if var >0:
                        for v in range(0, var):
                            vars = product_json['variants'][v]['size']
                            stock = product_json['variants'][v]['isInStock']
                            vid = product_json['variants'][v]['variantId']
                            if stock == True:
                                details_dict['Variant']=general.clean(vars)
                                details_dict['Variant_ID'] = vid
                                details_dict['Stock_Condition']="In Stock"
                                print(details_dict)
                                store_data_by_product_id(dict=[details_dict])

                            else:
                                details_dict['Variant'] = general.clean(vars)
                                details_dict['Variant_ID'] = vid
                                details_dict['Stock_Condition'] = "Out of Stock"
                                print(details_dict)
                                store_data_by_product_id(dict=[details_dict])

                    else:
                        details_dict['Variant'] = '-'
                        details_dict['Stock_Condition'] = 'In Stock'
                        details_dict['Variant_ID'] = '-'
                        store_data_by_product_id(dict=[details_dict])
                except:
                    details_dict['Variant'] = '-'
                    details_dict['Stock_Condition'] = '-'
                    details_dict['Variant_ID'] = '-'
                    store_data_by_product_id(dict=[details_dict])
            except Exception as e:
                if n != 0:
                    n = n - 1
                    by_product_id(prod_id_data, n)
                else:
                    print('this2222')
                    details_dict['Reason_Code'] = 'Success-PNF'
                    print(details_dict)
                    store_data_by_product_id(dict=[details_dict])
        elif n!=0:
            n = n - 1

            by_product_id(prod_id_data,n)
            print(prod_id_data)
            # store_data_by_product_id(dict=[details_dict])

        elif response.status_code==408:
            details_dict['Reason_Code']='TimeOut'
            print(details_dict)
            store_data_by_product_id(dict=[details_dict])
        elif response.status_code==403:
            details_dict['Reason_Code']='Blocked'
            print(details_dict)
            store_data_by_product_id(dict=[details_dict])
        else:
            details_dict['Reason_Code']= 'Success-PNF'
            print(details_dict)
            store_data_by_product_id(dict=[details_dict])
    except:
        if n != 0:
            n = n - 1
            by_product_id(prod_id_data, n)
        else:
            details_dict['Reason_Code'] = 'TimeOut'
            print(details_dict)
            store_data_by_product_id(dict=[details_dict])

count=0
for data in extract_xl():
    print('inside 1')
    count+=1
    by_product_id(prod_id_data=data,n=6)
    print(count)
    # break
