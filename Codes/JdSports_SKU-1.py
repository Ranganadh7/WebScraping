import scrapy
import requests, re, json
from scrapy.cmdline import execute
import general
import os
import scrapy
current_path = os.path.dirname(os.path.abspath(__file__))
import requests, re, json
from scrapy.cmdline import execute
from scrapy.http import HtmlResponse
import json
import json
import lxml.html
import openpyxl
import pandas as pd
import lxml.html
import pandas as pd
import requests, threading, csv
import datetime
from lxml import html
import random
from urllib3.filepost import writer
import random


product_details = []

# excel_data = pd.read_csv('JDSportsSKU.csv')
# p_rpc = excel_data['Input_URL']#[7000:7100]

# file_location=r'E:\Nilesh\Adidas\PDP\Jdsport_uk_1\Input_PDP.xlsx'
# file_location=r'input-1.xlsx'
file_location=r'input-1.xlsx'

def get_xl_file(path):
    workbook=openpyxl.load_workbook(path)
    sheet_obj=workbook.active
    all_rows=sheet_obj.max_row

    urlsdata=[]
    for i in range(1, all_rows):
        # print(i)
        # Adidas_id = sheet_obj.cell(row=i + 1, column=3)
        zalando_id=sheet_obj.cell(row=i+1,column=1)
        urls=sheet_obj.cell(row=i+1,column=2)
        urls1 = sheet_obj.cell(row=i + 1, column=6)
        urls2 = sheet_obj.cell(row=i + 1, column=7)

        # print(Adidas_id)
        u_data=str(zalando_id.value)+'&*&'+str(urls.value)+'&*&'+str(urls1.value)+'&*&'+str(urls2.value)
        # print(u_data)
        urlsdata.append(u_data.split('&*&'))
    # print(urlsdata)
    return urlsdata
get_xl_file(path=file_location)

import csv
from csv import DictWriter
def store_data_in_csv(dict):
    # fieldnames=['INPUT','Website Name', 'Country', 'Brand', 'Product_Name','Sponsored_Tag', 'Image_URLS',
    #             'Rank', 'Page', 'Category_Path', 'Product_URL', 'Listing URL', 'Reason_code', 'Crawling_TimeStamp',
    #             'Cache_Page_Link', 'Extra1', 'Extra2', 'Extra3','Extra4','Extra5']
    fieldnames = ['Path_ID','Page_ID','Navigation Path','Website_Name','Country','Brand','RPC','Product_name','Sponsored_tag','Image_url','Rank','Page','Category_path','Product_url','Listing_url','Reason_code','Crawling_timestamp','Cache_page_link','Extra1','Extra2','Extra3','Extra4','Extra5']
    with open('ZALANDO_DE_SON.csv', 'a+' , encoding='UTF-8', newline='') as file:
        writer = DictWriter(file, fieldnames=fieldnames)
        # writer.writeheader()
        writer.writerows(dict)
        # file.close()

def scrap_data():
    # try:
    try:
        crawltime = datetime.datetime.now().strftime('%Y-%m-%dT%H:%M:%SZ')
    except Exception as e:
        crawltime = ''
    for i in get_xl_file(path=file_location):
        strrpc = ''
        try:
            website_Name = i[0]
            country = i[1]
            sku_ID = i[2]
            pdpurl = i[3]
            # pdpurl = 'https://www.jdsports.co.uk/product/grey-adidas-essential-slim-joggers/16238552/'
            product_code = ''
            headers = {
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36'
            }
            sku_url = pdpurl
            for _ in range(20):
                p1 = ['154.28.67.106', '154.28.67.111', '154.28.67.116', '154.28.67.117', '154.28.67.125',
                      '154.28.67.131',
                      '154.28.67.133', '154.28.67.142', '154.28.67.156', '154.28.67.163', '154.28.67.173',
                      '154.28.67.18',
                      '154.28.67.182', '154.28.67.184', '154.28.67.20', '154.28.67.200', '154.28.67.210',
                      '154.28.67.218',
                      '154.28.67.222', '154.28.67.223', '154.28.67.231', '154.28.67.240', '154.28.67.243',
                      '154.28.67.253',
                      '154.28.67.39', '154.28.67.4', '154.28.67.49', '154.28.67.5', '154.28.67.61', '154.28.67.80',
                      '154.28.67.81', '154.28.67.87', '154.28.67.88', '154.28.67.96', '154.28.67.99', '154.7.230.100',
                      '154.7.230.101', '154.7.230.103', '154.7.230.107', '154.7.230.109', '154.7.230.130',
                      '154.7.230.132',
                      '154.7.230.14', '154.7.230.140', '154.7.230.147', '154.7.230.151', '154.7.230.156',
                      '154.7.230.163',
                      '154.7.230.170', '154.7.230.18', '154.7.230.183', '154.7.230.188', '154.7.230.189',
                      '154.7.230.19',
                      '154.7.230.190', '154.7.230.198', '154.7.230.204', '154.7.230.209', '154.7.230.235',
                      '154.7.230.238',
                      '154.7.230.246', '154.7.230.29', '154.7.230.41', '154.7.230.42', '154.7.230.51', '154.7.230.55',
                      '154.7.230.60', '154.7.230.61', '154.7.230.74', '154.7.230.82', '154.7.230.89', '23.131.8.112',
                      '23.131.8.115', '23.131.8.117', '23.131.8.12', '23.131.8.121', '23.131.8.124', '23.131.8.150',
                      '23.131.8.161', '23.131.8.166', '23.131.8.171', '23.131.8.173', '23.131.8.176', '23.131.8.177',
                      '23.131.8.181', '23.131.8.19', '23.131.8.192', '23.131.8.194', '23.131.8.199', '23.131.8.202',
                      '23.131.8.203', '23.131.8.204', '23.131.8.207', '23.131.8.209', '23.131.8.213', '23.131.8.216',
                      '23.131.8.225', '23.131.8.228', '23.131.8.231', '23.131.8.238', '23.131.8.254', '23.131.8.36',
                      '23.131.8.5', '23.131.8.76', '23.131.8.93', '23.131.8.95', '23.131.8.99', '23.131.88.105',
                      '23.131.88.12', '23.131.88.137', '23.131.88.139', '23.131.88.140', '23.131.88.145',
                      '23.131.88.150',
                      '23.131.88.151', '23.131.88.153', '23.131.88.154', '23.131.88.156', '23.131.88.165',
                      '23.131.88.18',
                      '23.131.88.191', '23.131.88.192', '23.131.88.194', '23.131.88.198', '23.131.88.202',
                      '23.131.88.206',
                      '23.131.88.220', '23.131.88.223', '23.131.88.228', '23.131.88.233', '23.131.88.24',
                      '23.131.88.242',
                      '23.131.88.244', '23.131.88.47', '23.131.88.63', '23.131.88.67', '23.131.88.73', '23.131.88.80',
                      '23.131.88.81', '23.131.88.82', '23.131.88.88', '23.131.88.97', '23.170.144.149',
                      '23.170.144.209',
                      '23.170.144.212', '23.170.144.242', '23.170.144.83', '23.170.145.117', '23.170.145.167',
                      '23.170.145.182', '23.170.145.19', '23.170.145.203', '23.226.17.101', '23.226.17.109',
                      '23.226.17.112', '23.226.17.113', '23.226.17.115', '23.226.17.123', '23.226.17.129',
                      '23.226.17.143',
                      '23.226.17.148', '23.226.17.165', '23.226.17.186', '23.226.17.199', '23.226.17.201',
                      '23.226.17.207',
                      '23.226.17.210', '23.226.17.219', '23.226.17.220', '23.226.17.222', '23.226.17.229',
                      '23.226.17.250',
                      '23.226.17.254', '23.226.17.26', '23.226.17.33', '23.226.17.4', '23.226.17.49', '23.226.17.5',
                      '23.226.17.55', '23.226.17.66', '23.226.17.7', '23.226.17.72', '23.226.17.78', '23.226.17.8',
                      '23.226.17.86', '23.226.17.90', '23.226.17.93', '23.230.177.105', '23.230.177.110',
                      '23.230.177.113',
                      '23.230.177.121', '23.230.177.130', '23.230.177.14', '23.230.177.143', '23.230.177.15',
                      '23.230.177.150', '23.230.177.154', '23.230.177.165', '23.230.177.173', '23.230.177.191',
                      '23.230.177.196', '23.230.177.203', '23.230.177.206', '23.230.177.208', '23.230.177.217',
                      '23.230.177.220', '23.230.177.221', '23.230.177.224', '23.230.177.228', '23.230.177.231',
                      '23.230.177.235', '23.230.177.237', '23.230.177.241', '23.230.177.27', '23.230.177.38',
                      '23.230.177.52', '23.230.177.61', '23.230.177.67', '23.230.177.72', '23.230.177.80',
                      '23.230.177.88',
                      '23.230.177.94', '23.230.177.99', '23.230.197.103', '23.230.197.106', '23.230.197.109',
                      '23.230.197.11', '23.230.197.12', '23.230.197.122', '23.230.197.124', '23.230.197.146',
                      '23.230.197.155', '23.230.197.156', '23.230.197.174', '23.230.197.179', '23.230.197.181',
                      '23.230.197.196', '23.230.197.2', '23.230.197.201', '23.230.197.207', '23.230.197.208',
                      '23.230.197.225', '23.230.197.227', '23.230.197.233', '23.230.197.236', '23.230.197.239',
                      '23.230.197.240', '23.230.197.244', '23.230.197.251', '23.230.197.50', '23.230.197.52',
                      '23.230.197.54', '23.230.197.60', '23.230.197.71', '23.230.197.80', '23.230.197.81',
                      '23.230.197.84',
                      '23.230.197.97', '23.230.74.102', '23.230.74.110', '23.230.74.116', '23.230.74.125',
                      '23.230.74.133',
                      '23.230.74.135', '23.230.74.14', '23.230.74.141', '23.230.74.149', '23.230.74.15',
                      '23.230.74.157',
                      '23.230.74.16', '23.230.74.170', '23.230.74.172', '23.230.74.174', '23.230.74.183',
                      '23.230.74.187',
                      '23.230.74.19', '23.230.74.198', '23.230.74.208', '23.230.74.212', '23.230.74.215',
                      '23.230.74.23',
                      '23.230.74.230', '23.230.74.231', '23.230.74.252', '23.230.74.30', '23.230.74.41', '23.230.74.57',
                      '23.230.74.58', '23.230.74.59', '23.230.74.6', '23.230.74.75', '23.230.74.81', '23.230.74.88',
                      '23.230.74.91', '23.27.222.108', '23.27.222.109', '23.27.222.134', '23.27.222.138',
                      '23.27.222.159',
                      '23.27.222.161', '23.27.222.164', '23.27.222.166', '23.27.222.178', '23.27.222.19',
                      '23.27.222.195',
                      '23.27.222.201', '23.27.222.202', '23.27.222.203', '23.27.222.208', '23.27.222.21',
                      '23.27.222.211',
                      '23.27.222.218', '23.27.222.223', '23.27.222.228', '23.27.222.234', '23.27.222.236',
                      '23.27.222.242',
                      '23.27.222.251', '23.27.222.253', '23.27.222.34', '23.27.222.61', '23.27.222.62', '23.27.222.69',
                      '23.27.222.70', '23.27.222.72', '23.27.222.73', '23.27.222.74', '23.27.222.81', '23.27.222.93',
                      '38.131.131.110', '38.131.131.114', '38.131.131.123', '38.131.131.125', '38.131.131.137',
                      '38.131.131.142', '38.131.131.145', '38.131.131.147', '38.131.131.15', '38.131.131.154',
                      '38.131.131.16', '38.131.131.17', '38.131.131.173', '38.131.131.18', '38.131.131.193',
                      '38.131.131.204', '38.131.131.207', '38.131.131.227', '38.131.131.229', '38.131.131.233',
                      '38.131.131.238', '38.131.131.246', '38.131.131.248', '38.131.131.250', '38.131.131.31',
                      '38.131.131.36', '38.131.131.50', '38.131.131.58', '38.131.131.64', '38.131.131.70',
                      '38.131.131.71',
                      '38.131.131.74', '38.131.131.83', '38.131.131.94', '38.131.131.99', '38.75.75.104',
                      '38.75.75.111',
                      '38.75.75.112', '38.75.75.119', '38.75.75.120', '38.75.75.123', '38.75.75.127', '38.75.75.139',
                      '38.75.75.14', '38.75.75.143', '38.75.75.155', '38.75.75.156', '38.75.75.158', '38.75.75.170',
                      '38.75.75.179', '38.75.75.188', '38.75.75.2', '38.75.75.201', '38.75.75.231', '38.75.75.232',
                      '38.75.75.241', '38.75.75.246', '38.75.75.251', '38.75.75.26', '38.75.75.29', '38.75.75.4',
                      '38.75.75.44', '38.75.75.49', '38.75.75.56', '38.75.75.58', '38.75.75.62', '38.75.75.72',
                      '38.75.75.76', '38.75.75.79', '38.75.75.88', '38.96.156.108', '38.96.156.112', '38.96.156.128',
                      '38.96.156.131', '38.96.156.14', '38.96.156.142', '38.96.156.143', '38.96.156.149',
                      '38.96.156.16',
                      '38.96.156.163', '38.96.156.165', '38.96.156.169', '38.96.156.186', '38.96.156.188',
                      '38.96.156.190',
                      '38.96.156.192', '38.96.156.194', '38.96.156.199', '38.96.156.218', '38.96.156.236',
                      '38.96.156.240',
                      '38.96.156.252', '38.96.156.28', '38.96.156.32', '38.96.156.35', '38.96.156.56', '38.96.156.57',
                      '38.96.156.6', '38.96.156.67', '38.96.156.77', '38.96.156.80', '38.96.156.83', '38.96.156.84',
                      '38.96.156.89', '38.96.156.92', '45.238.157.100', '45.238.157.104', '45.238.157.106',
                      '45.238.157.110', '45.238.157.116', '45.238.157.118', '45.238.157.119', '45.238.157.12',
                      '45.238.157.123', '45.238.157.132', '45.238.157.14', '45.238.157.149', '45.238.157.15',
                      '45.238.157.183', '45.238.157.186', '45.238.157.189', '45.238.157.2', '45.238.157.212',
                      '45.238.157.214', '45.238.157.217', '45.238.157.22', '45.238.157.228', '45.238.157.23',
                      '45.238.157.247', '45.238.157.43', '45.238.157.48', '45.238.157.51', '45.238.157.52',
                      '45.238.157.53',
                      '45.238.157.56', '45.238.157.61', '45.238.157.65', '45.238.157.72', '45.238.157.79',
                      '45.238.157.8',
                      '45.238.159.103', '45.238.159.107', '45.238.159.110', '45.238.159.114', '45.238.159.116',
                      '45.238.159.123', '45.238.159.126', '45.238.159.144', '45.238.159.148', '45.238.159.15',
                      '45.238.159.156', '45.238.159.165', '45.238.159.167', '45.238.159.183', '45.238.159.20',
                      '45.238.159.208', '45.238.159.217', '45.238.159.220', '45.238.159.23', '45.238.159.230',
                      '45.238.159.235', '45.238.159.237', '45.238.159.238', '45.238.159.24', '45.238.159.249',
                      '45.238.159.251', '45.238.159.32', '45.238.159.34', '45.238.159.51', '45.238.159.6',
                      '45.238.159.66',
                      '45.238.159.77', '45.238.159.79', '45.238.159.82', '45.238.159.91']

                p_auth = str("csimonra:h19VA2xZ")

                p_host = random.choice(p1)
                p_port = "29842"
                proxy = {
                    'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                    'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
                }

                # res = requests.request("GET", son_url, headers=headers)
                res = requests.get(sku_url, headers=headers, proxies=proxy)
                if res.status_code == 200:
                    break
                elif res.status_code==404:
                    product_data = {

                        'SKU_ID': sku_ID,
                        'Website': website_Name,
                        'Country': country,
                        'RPC': strrpc,
                        'MPC': '-',
                        'Product_ID': product_code,
                        'Product_URL': pdpurl,
                        'Product_Name': '-',
                        'Category_Path': '-',
                        'Specification': '-',
                        'Description': '-',
                        'Currency': '-',
                        'List_Price': '-',
                        'Promo_Price': '-',
                        'Discount': '-',
                        'Brand': '-',
                        'Rating_Count': '-',
                        'Review_Count': '-',
                        'Image_URLs': '-',
                        'Variant': '-',
                        'Variant_ID': '-',
                        'Colour_of_Variant': '-',
                        'Colour_Grouping': '-',
                        'Seller_Name': '-',
                        'Stock_Count': '-',
                        'Stock_Condition': '-',
                        'Stock_Message': '-',
                        'Sustainability_Badge': '-',
                        'Reason_Code': 'Blocked',
                        'Crawling_TimeStamp': crawltime,
                        'Cache_Page_Link': '',
                        'Extra1': '-',
                        'Extra2': '-',
                        'Extra3': '-',
                        'Extra4': '-',
                        'Extra5': '-',

                    }
                    data_dict_copy = product_data.copy()
                    product_details.append(data_dict_copy)

                # res = requests.request("GET", sku_url, headers=headers)
            response = HtmlResponse(url=sku_url, body=res.content)

            # changelist  = re.findall('href=(.*?).css', response.text)
            # print(changelist)

            brand = ''.join(re.findall('brand: "(.*?)",', response.text, re.DOTALL))
            if brand == '':
                pass

            else:
                try:
                    rpc_id = re.findall('upc: "(.*?)"', response.text)[1]
                    if rpc_id == []:
                        rpc_id = ''
                except Exception as e:
                    rpc_id = ''
                    print(e)

                try:
                    datazone = datetime.datetime.now()
                    f_date = datazone.strftime("%d_%m_%Y")
                    # f_date = '05_05_2022'

                    strdate = datazone.day
                    strm = datazone.month
                    stry = datazone.year
                    cpid = sku_ID + '_' + str(strdate) + '_' + str(strm) + '_' + str(stry)
                    # cpid = sku_ID + '05_05_2022'
                    # ASS_folder = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Jdsports_uk\\PDP"
                    ASS_folder = f"\\\\10.100.20.40\\E$\\ADIDAS_SavePages\\Jdsports_uk\\PDP"
                    sos_date_wise_folder = ASS_folder + f"\\{f_date}"
                    if os.path.exists(sos_date_wise_folder):
                        pass
                    else:
                        os.mkdir(sos_date_wise_folder)
                    sos_filename = sos_date_wise_folder + "\\" + cpid + ".html"
                    sos_filename = sos_filename.replace("+", "_").replace("-", "_")
                    if os.path.exists(sos_filename):
                        with open(sos_filename, 'wb') as f:
                            f.write(response.body)
                            f.close()
                    else:
                        with open(sos_filename, 'wb') as f:
                            f.write(response.body)
                            f.close()
                    #https://ecxus440.eclerx.com/cachepages
                    page_location = sos_filename
                    page_location = page_location.replace('\\\\10.100.20.40\\E$\\ADIDAS_SavePages\\','https:////ecxus440.eclerx.com//cachepages//').replace('\\','//').replace('//','/')

                    # page_path = f'D:\\Prashant_Singh\\eclerx_tech\\page_save\\Jdsports\\SKU\\{rpc_id}.html'
                    # # page_location = page_path
                    # with open(page_path, 'wb') as f:
                    #     f.write(response.body)
                    #     f.close()
                except Exception as e:
                    page_location = ''
                    print(e)



                try:
                    p_name = ''.join(response.xpath('//h1[@itemprop="name"]/text()').extract()).strip()
                    p_name = general.clean(p_name)
                except Exception as e:
                    p_name = '-'
                    print(e)

                try:
                    brand = ''.join(re.findall('brand: "(.*?)",', response.text, re.DOTALL))
                    brand = general.clean(brand)
                except Exception as e:
                    brand = '-'
                    print(e)

                # try:
                #     category = ''.join(re.findall('"category": "(.*?)",', response.text, re.DOTALL)).replace(' / ','>')
                #     category = general.clean(category)
                #     if category == '':
                #         category = 'Home > ' + p_name
                #
                # except Exception as e:
                #     category = '-'
                #     print(e)

                try:
                    category_list = ['Home']
                    cat1 = response.xpath('//script[@type="application/ld+json"]/text()').getall()[-1]
                    categ = json.loads(cat1)
                    cat = categ['itemListElement']
                    for i in range(len(cat)):
                        cat_name = cat[i]['name']
                        category_list.append(cat_name)
                    category = '>'.join(category_list)
                    # cat1 = str(response.text).split(': "BreadcrumbList",')[-1]
                except Exception as e:
                    category ='-'
                    print(e)

                try:
                    promo_price = ''.join(response.xpath('//*[@data-e2e="product-price"]/span/text()').extract())
                    if promo_price == '':
                        promo_price = ''.join(response.xpath('//*[@data-e2e="product-price"]/@content').extract())
                    promo_price = promo_price.replace('£','')
                    promo_price = general.clean(promo_price)
                except Exception as e:
                    promo_price = ''
                    print(e)

                if promo_price == '':
                    promo_price = '-'

                list_price = ''
                try:
                    list_price = ''.join(response.xpath('//*[@class="itemPrices"]/span[@class="was"]/span/text()').extract())
                    list_price = list_price.replace('£', '')
                    list_price = general.clean(list_price)
                except Exception as e:
                    list_price = ''
                    print(e)
                discount_price = ''
                try:
                    discount_price = ''.join(response.xpath('//div[@class="itemPrices"]/span[@class="pri"]/span[@class="sav"]/text()').extract())
                    discount_price = general.clean(discount_price)
                except Exception as e:
                    discount_price = ''
                    print(e)

                if promo_price == '':
                    promo_price = list_price

                if list_price == '':
                    list_price = promo_price

                if discount_price == '':
                    discount_price = '-'

                try:
                    image = ''.join(re.findall('"image": (.*?)],',response.text)).replace('","','|').replace('["','').replace('"]','')
                    if image == '':
                        # imageraw = response.xpath('//div[@class="owl-thumbs"]/button/picture/img/@src')
                        image = '-'
                    image = general.clean(image)
                    image = image.replace('"','').strip()
                    # images = list(image)
                except Exception as e:
                    image = '-'
                    print(e)

                try:
                    product_url = ''.join(response.xpath('//*[@rel="canonical"]/@href').extract())
                    product_id_1 = product_url.split('/')[-2]
                except Exception as e:
                    product_url = ''
                    product_id_1 = ''
                    print(e)

                ###    Reviews and Rating Code here

                try:
                    rat_url = 'https://display.powerreviews.com/m/945969/l/all/product/'+str(product_id_1)+'/reviews?apikey=df8b31d7-b93d-4b8b-8f62-db57875c78fd&_noconfig=true&page_locale=en_GB'
                    # # rat_url = 'https://display.powerreviews.com/m/945969/l/all/product/16204230/reviews?apikey=df8b31d7-b93d-4b8b-8f62-db57875c78fd&_noconfig=true&page_locale=en_GB'
                    for _ in range(20):
                    #     p1 = ['104.218.195.130', '104.218.195.205', '104.251.82.191', '104.251.82.240', '104.251.82.63',
                    #           '104.251.84.104', '23.106.26.65', '23.106.26.84', '23.106.27.13', '23.106.27.139',
                    #           '23.106.27.144',
                    #           '23.106.28.230', '23.106.28.237', '23.106.30.117', '23.106.30.126', '23.110.166.102',
                    #           '23.110.166.26',
                    #           '23.110.166.76', '23.110.169.100', '23.110.169.162', '23.110.173.171', '23.110.173.225',
                    #           '23.129.136.120',
                    #           '23.129.136.245', '23.129.40.19', '23.129.40.44', '23.129.40.76', '23.129.56.191',
                    #           '23.129.56.237',
                    #           '23.161.3.146', '23.161.3.67', '23.170.144.104', '23.170.144.108', '23.170.144.19',
                    #           '23.170.145.103',
                    #           '23.170.145.252', '23.170.145.51', '23.175.176.21', '23.175.176.24', '23.175.177.176',
                    #           '23.175.177.183',
                    #           '23.175.177.8', '23.176.49.110', '23.176.49.183', '23.177.240.144', '23.177.240.217',
                    #           '23.177.240.90',
                    #           '23.184.144.105', '23.184.144.124', '23.184.144.231', '23.185.112.167', '23.185.112.229',
                    #           '23.185.144.110', '23.185.144.171', '23.185.144.197', '23.185.80.164', '23.185.80.4',
                    #           '23.185.80.6',
                    #           '23.186.48.210', '23.186.48.248', '23.226.16.211', '23.226.16.243', '23.226.17.178',
                    #           '23.226.17.240',
                    #           '23.226.18.106', '23.226.18.193', '23.226.19.212', '23.226.19.87', '23.226.20.187',
                    #           '23.226.20.90',
                    #           '23.226.21.13', '23.226.21.22', '23.226.22.216', '23.226.22.53', '23.226.23.178', '23.226.23.250',
                    #           '23.226.24.190', '23.226.24.6', '23.226.24.64', '23.226.25.107', '23.226.25.68', '23.226.26.158',
                    #           '23.226.26.202', '23.226.26.220', '23.226.27.246', '23.226.27.94', '23.226.28.159',
                    #           '23.226.28.194',
                    #           '23.226.28.231', '23.226.29.126', '23.226.29.99', '23.226.30.191', '23.226.30.235',
                    #           '23.226.31.131',
                    #           '23.226.31.169', '23.226.31.193', '23.247.172.197', '23.247.172.214', '23.247.172.51',
                    #           '23.247.173.202',
                    #           '23.247.173.6', '23.247.174.196', '23.247.174.211', '23.247.174.81', '23.247.175.156',
                    #           '23.247.175.215',
                    #           '23.247.175.218', '23.27.9.103', '23.27.9.228', '23.82.105.11', '23.82.105.194', '23.82.105.45',
                    #           '23.82.109.165', '23.82.109.242', '23.82.184.118', '23.82.184.227', '23.82.184.80',
                    #           '104.251.84.217',
                    #           '52.128.15.114', '52.128.15.92', '52.128.196.70', '23.105.0.63',
                    #           '45.59.181.80', '23.184.144.105', '52.128.198.105', '52.128.196.173'
                    #                                                               '154.13.247.219', '154.13.247.42']
                    #     p_auth = str("csimonra:h19VA2xZ")
                    #     p_host = random.choice(p1)
                    #     p_port = "29842"
                    #     proxy = {
                    #         'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                    #         'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
                    #     }
                        res1 = requests.request("GET", rat_url, headers=headers,proxies=proxy)
                        if res1.status_code == 200:
                            break
                    response1 = HtmlResponse(url=rat_url, body=res1.content)
                    sos_filename1 = sos_date_wise_folder + "\\" + cpid + "_review.html"
                    sos_filename1 = sos_filename1.replace("+", "_").replace("-", "_")
                    if os.path.exists(sos_filename1):
                        with open(sos_filename1, 'wb') as f:
                            f.write(response.body)
                            f.close()
                    else:
                        with open(sos_filename1, 'wb') as f:
                            f.write(response.body)
                            f.close()


                    rat_data = response1.text
                    rat_json = json.loads(rat_data)



                    try:
                        reviews = rat_json['results'][0]['rollup']['review_count']
                        rating = rat_json['results'][0]['rollup']['average_rating']
                        rating = round(rating,1)
                    except Exception as e:
                        reviews = '-'
                        rating = '-'
                        print(e)

                    try:
                        # reviews = rat_json['results'][0]['rollup']['review_count']
                        # rating = rat_json['results'][0]['rollup']['average_rating']
                        if rating == '':
                            rating = rat_json['results'][0]['rollup']['page_brand_score']['average_rating']
                    except Exception as e:
                        # reviews = ''
                        rating = '-'
                        print(e)
                except:
                    rating = '-'
                    reviews = '-'

                try:
                    color = ''.join(re.findall('"color": "(.*?)",',response.text))
                    # if image == ']':
                    #     image = ''
                except Exception as e:
                    color = '-'
                    print(e)


                try:
                    stock_details = ''.join(response.xpath('//*[@id="addToBasket"]/@title').extract())
                except Exception as e:
                    stock_details = ''
                    print(e)

                try:
                    description = ''.join(re.findall('"description": "(.*?)",',response.text))
                    description = description.replace('•','|')
                    description = general.clean(description)
                except Exception as e:
                    description = '-'
                    print(e)

                try:
                    specification = response.xpath('//*[contains(text(),"Care & Material")]//following-sibling::text()').extract()[0].strip()
                    specification= general.clean(specification)
                except Exception as e:
                    specification = '-'
                    print(e)

                try:
                    product_code = ''.join(response.xpath('//*[@class="product-code"]/text()').extract()).replace('Product Code:','').strip()
                    product_code = general.clean(product_code)
                except Exception as e:
                    product_code = ''
                    print(e)

                try:
                    color_variant = ''.join(response.xpath('//*[contains(text(),"Colour:")]//following-sibling::text()').extract()).strip()
                    color_variant = general.clean(color_variant)
                except Exception as e:
                    color_variant = '-'
                    print(e)

                try:
                    group_color = '|'.join(response.xpath('//*[@id="relThumbs"]//@title').extract())
                except Exception as e:
                    group_color = '-'
                    print(e)

                try:
                    stock_status = ''.join(response.xpath('//*[@name="twitter:data2"]/@content').extract())
                    if 'in stock' in stock_status.lower():
                        stock_status = 'InStock'
                    else:
                        stock_status = 'Out of Stock'
                except Exception as e:
                    stock_status = ''
                    print(e)


                try:
                    res1 = response.text.replace('\\','')
                    variant = re.findall('ame:"(.*?)",',res1)
                    # variant = general.clean(variant)
                    if variant == []:
                        variant = ''
                except Exception as e:
                    variant = ''
                    print(e)



                try:
                    variant_id =  ','.join(re.findall('upc: "(.*?)"',response.text))
                    #variant_id = general.clean(variant_id)
                    variant_id=str(variant_id)
                    if variant_id == []:
                        variant_id = '-'
                except Exception as e:
                    variant_id = ''
                    print(e)

                try:
                    crawltime = datetime.datetime.now().strftime('%Y-%m-%dT%H:%M:%SZ')
                except Exception as e:
                    crawltime = ''

                try:
                    strrpc = pdpurl.split('/')[-2]
                except Exception as e:
                    strrpc = ''

                # page_path = f'D:\\Eclerx_Fremwork_code\\page_save\\sku_page_save\\{product_code}.html'
                # with open(page_path, 'wb') as f:
                #     f.write(response.body)
                #     f.close()

                try:
                    cache_page_link = str(hash(product_url))[1:10]
                except Exception as e:
                    cache_page_link = ''
                    print(e)

                #seller_name --------------
                seller_name = '-'
                sell_url = pdpurl+'stock/'
                try:
                    for _ in range(5):
                        p1 = ['104.218.195.130', '104.218.195.205', '104.251.82.191', '104.251.82.240', '104.251.82.63',
                              '104.251.84.104', '23.106.26.65', '23.106.26.84', '23.106.27.13', '23.106.27.139',
                              '23.106.27.144',
                              '23.106.28.230', '23.106.28.237', '23.106.30.117', '23.106.30.126', '23.110.166.102',
                              '23.110.166.26',
                              '23.110.166.76', '23.110.169.100', '23.110.169.162', '23.110.173.171', '23.110.173.225',
                              '23.129.136.120',
                              '23.129.136.245', '23.129.40.19', '23.129.40.44', '23.129.40.76', '23.129.56.191',
                              '23.129.56.237',
                              '23.161.3.146', '23.161.3.67', '23.170.144.104', '23.170.144.108', '23.170.144.19',
                              '23.170.145.103',
                              '23.170.145.252', '23.170.145.51', '23.175.176.21', '23.175.176.24', '23.175.177.176',
                              '23.175.177.183',
                              '23.175.177.8', '23.176.49.110', '23.176.49.183', '23.177.240.144', '23.177.240.217',
                              '23.177.240.90',
                              '23.184.144.105', '23.184.144.124', '23.184.144.231', '23.185.112.167', '23.185.112.229',
                              '23.185.144.110', '23.185.144.171', '23.185.144.197', '23.185.80.164', '23.185.80.4',
                              '23.185.80.6',
                              '23.186.48.210', '23.186.48.248', '23.226.16.211', '23.226.16.243', '23.226.17.178',
                              '23.226.17.240',
                              '23.226.18.106', '23.226.18.193', '23.226.19.212', '23.226.19.87', '23.226.20.187',
                              '23.226.20.90',
                              '23.226.21.13', '23.226.21.22', '23.226.22.216', '23.226.22.53', '23.226.23.178',
                              '23.226.23.250',
                              '23.226.24.190', '23.226.24.6', '23.226.24.64', '23.226.25.107', '23.226.25.68',
                              '23.226.26.158',
                              '23.226.26.202', '23.226.26.220', '23.226.27.246', '23.226.27.94', '23.226.28.159',
                              '23.226.28.194',
                              '23.226.28.231', '23.226.29.126', '23.226.29.99', '23.226.30.191', '23.226.30.235',
                              '23.226.31.131',
                              '23.226.31.169', '23.226.31.193', '23.247.172.197', '23.247.172.214', '23.247.172.51',
                              '23.247.173.202',
                              '23.247.173.6', '23.247.174.196', '23.247.174.211', '23.247.174.81', '23.247.175.156',
                              '23.247.175.215',
                              '23.247.175.218', '23.27.9.103', '23.27.9.228', '23.82.105.11', '23.82.105.194',
                              '23.82.105.45',
                              '23.82.109.165', '23.82.109.242', '23.82.184.118', '23.82.184.227', '23.82.184.80',
                              '104.251.84.217',
                              '52.128.15.114', '52.128.15.92', '52.128.196.70', '23.105.0.63',
                              '45.59.181.80', '23.184.144.105', '52.128.198.105', '52.128.196.173'
                                                                                  '154.13.247.219', '154.13.247.42']
                        p_auth = str("csimonra:h19VA2xZ")
                        p_host = random.choice(p1)
                        p_port = "29842"
                        proxy = {
                            'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                            'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
                        }

                        # res = requests.request("GET", son_url, headers=headers)
                        try:
                            res_Sell = requests.get(sell_url, proxies=proxy)
                            if res_Sell.status_code == 200:

                                response123 = HtmlResponse(url=sku_url, body=res_Sell.text,encoding="utf-8")
                                # sname = response123.xpath('//img[@class="fulfilmentBrandImage"]/@scr').get('')
                                sname = re.findall(r'delivered direct from (.*?), any',str(response123.text))[0]
                                if 'reebok' in str(sname).lower():
                                    seller_name = 'reebok'
                                elif 'adidas' in str(sname).lower():
                                    seller_name = 'adidas'
                                else:
                                    seller_name = '-'
                                break
                        except:
                            pass
                except Exception as e:
                    seller_name = '-'
                for var in variant:
                    sarchid = 'name:"' + var +'",'
                    varraw = general.midtext(response.text,sarchid,'}')
                    vrid = general.midtext(varraw,'page_id_variant: "','"')
                    vrid=str(vrid)
                    vrid="'"+vrid+"'"
                    product_data = {

                        'SKU_ID': sku_ID,
                        'Website': website_Name,
                        'Country': country,
                        'RPC': strrpc,
                        'MPC': '-',
                        'Product_ID': product_code,
                        'Product_URL': pdpurl,
                        'Product_Name': p_name,
                        'Category_Path': category,
                        'Specification': specification,
                        'Description': description,
                        'Currency': 'GBP',
                        'List_Price': list_price,
                        'Promo_Price': promo_price,
                        'Discount': discount_price,
                        'Brand': brand,
                        'Rating_Count': rating,
                        'Review_Count': reviews,
                        'Image_URLs': image,
                        'Variant': var,
                        'Variant_ID': vrid,
                        'Colour_of_Variant': color_variant,
                        'Colour_Grouping': group_color,
                        'Seller_Name': seller_name,
                        'Stock_Count': '-',
                        'Stock_Condition': stock_status,
                        'Stock_Message': '-',
                        'Sustainability_Badge': '-',
                        'Reason_Code': 'Success-PF',
                        'Crawling_TimeStamp': crawltime,
                        'Cache_Page_Link': page_location,
                        'Extra1': '-',
                        'Extra2': '-',
                        'Extra3': '-',
                        'Extra4': '-',
                        'Extra5': '-',

                    }
                    data_dict_copy = product_data.copy()
                    product_details.append(data_dict_copy)
        except:
                product_data = {

                    'SKU_ID': sku_ID,
                    'Website': website_Name,
                    'Country': country,
                    'RPC': strrpc,
                    'MPC': '-',
                    'Product_ID': product_code,
                    'Product_URL': pdpurl,
                    'Product_Name': '-',
                    'Category_Path': '-',
                    'Specification': '-',
                    'Description': '-',
                    'Currency': '-',
                    'List_Price': '-',
                    'Promo_Price': '-',
                    'Discount': '-',
                    'Brand': '-',
                    'Rating_Count': '-',
                    'Review_Count': '-',
                    'Image_URLs': '-',
                    'Variant': '-',
                    'Variant_ID': '-',
                    'Colour_of_Variant': '-',
                    'Colour_Grouping': '-',
                    'Seller_Name': '-',
                    'Stock_Count': '-',
                    'Stock_Condition': '-',
                    'Stock_Message': '-',
                    'Sustainability_Badge': '-',
                    'Reason_Code': 'Blocked',
                    'Crawling_TimeStamp': crawltime,
                    'Cache_Page_Link': '',
                    'Extra1': '-',
                    'Extra2': '-',
                    'Extra3': '-',
                    'Extra4': '-',
                    'Extra5': '-',

                }
                data_dict_copy = product_data.copy()
                product_details.append(data_dict_copy)

scrap_data()
data_df = pd.DataFrame(product_details)
datazone = datetime.datetime.now()
f_date = datazone.strftime("%d_%m_%Y")
data_df = pd.DataFrame(product_details)
data_csv = data_df.to_csv(f"{f_date}-jdsports_SKU-1.csv",index=False)#, encoding="cp1252")


# with pd.ExcelWriter('07-JdSports-PDP.xlsx', options={'strings_to_urls': False}) as writer:
# # # with pd.ExcelWriter('jdsports_SKU_OUTPUT1_0604.xlsx', options={'strings_to_urls': False}) as writer:
#      data_df.to_excel(writer,index=False,freeze_panes=(1, 0))

print("\nData saved successfully!\n")
