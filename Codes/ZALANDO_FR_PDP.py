import random
import re
from lxml import html
import time
# import xlrd
import requests
from lxml import html
import time
import json
from datetime import datetime, timedelta
import datetime
import openpyxl
# import xmltodict
from datetime import date, timedelta, time, datetime
import datetime
import time
import os
import general

headers = {
    'accept-language': 'en-US''en;q=0.9',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36'}

start = time.time()

# file_path = 'C:\\Users\\Surabhi.Nanhe\\PycharmProjects\\ZalandoProject\\zalando.fr_input.xlsx'
# file_path = 'zalando_fr_input1.xlsx'
file_path = 'zalando_fr_input1.xlsx'


def extract_xl():
    workbook = openpyxl.load_workbook(file_path)
    sheet_obj = workbook.active
    all_rows = sheet_obj.max_row

    inputdata = []
    for i in range(1, all_rows):
        pdp_url = sheet_obj.cell(row=i + 1, column=4)
        # pdp_url = sheet_obj.cell(row=i + 1, column=4)
        sku_id = sheet_obj.cell(row=i + 1, column=3)
        country = sheet_obj.cell(row=i + 1, column=2)
        website_name = sheet_obj.cell(row=i + 1, column=1)

        u_data = [pdp_url.value, sku_id.value, country.value, website_name.value]
        # u_data=str(pdp_url.value)+'&*&'+str(sku_id.value)+'&*&'+str(country.value)+'&*&'+str(website_name.value)

        inputdata.append(u_data)
    # print(inputdata)
    return inputdata


# extract_xl()

from csv import DictWriter

datazone = datetime.datetime.now()
f_s_d = datazone.strftime("%d-%m-")
def store_data_by_product_id(dict):
    fieldnames = ['SKU_ID', 'Website', 'Country', 'RPC', 'MPC', 'Product_ID', 'Product_URL',
                  'Product_Name', 'Category_Path', 'Specification', 'Description', 'Currency', 'List_Price',
                  'Promo_Price', 'Discount', 'Brand', 'Rating_Count', 'Review_Count', 'Image_URLs', 'Variant',
                  'Variant_ID', 'Colour_of_Variant', 'Colour_Grouping', 'Seller_Name', 'Stock_Count', 'Stock_Condition',
                  'Stock_Message', 'Sustainability_Badge', 'Reason_Code', 'Crawling_TimeStamp', 'Cache_Page_Link',
                  'Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5']
    # with open('FR_ZALANDO_PDP1_0804.csv', 'a+', encoding='utf-8-sig', newline='') as file:
    with open(f'1-{f_s_d}FR_ZALANDO_PDP-1.csv', 'a+', encoding='utf-8-sig', newline='') as file:
        writer = DictWriter(file, fieldnames=fieldnames)
        # writer.writeheader()
        writer.writerows(dict)
        # file.close()


def html_cache_page_saving(sku, response):
    datazone = datetime.datetime.now()
    f_date = datazone.strftime("%d_%m_%Y")
    strdate = datazone.day
    strm = datazone.month
    stry = datazone.year
    pageid = sku
    cpid = pageid + '_' + str(strdate) + '_' + str(strm) + '_' + str(stry)
    # ASS_folder = f"E:\ADIDAS_SavePages\Jdsports_uk\ASS"

    ASS_folder = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Zalando_FR\\PDP"
    sos_date_wise_folder = ASS_folder + f"\\{f_date}"
    if os.path.exists(sos_date_wise_folder):
        pass
    else:
        os.mkdir(sos_date_wise_folder)
    sos_filename = sos_date_wise_folder + "\\" + cpid + ".html"
    sos_filename = sos_filename.replace("+", "_").replace("-", "_")
    page_path = sos_filename.replace('/', '')
    print(page_path)
    page_path = page_path.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\',
                                  'https:////ecxus440.eclerx.com//cachepages//').replace('\\',
                                                                                         '//').replace(
        '//', '/')
    print(page_path)
    if os.path.exists(sos_filename):
        with open(sos_filename, 'w', encoding='utf-8') as f:
            f.write(response)
            f.close()
    else:
        with open(sos_filename, 'w', encoding='utf-8') as f:
            f.write(response)
            f.close()
    # path = os.path.abspath(sos_filename)
    # print(path)
    return page_path


dom = 'https://www.zalando.fr'

p1 = ['154.28.67.106', '154.28.67.111', '154.28.67.116', '154.28.67.117', '154.28.67.125',
                      '154.28.67.131',
                      '154.28.67.133', '154.28.67.142', '154.28.67.156', '154.28.67.163', '154.28.67.173',
                      '154.28.67.18',
                      '154.28.67.182', '154.28.67.184', '154.28.67.20', '154.28.67.200', '154.28.67.210',
                      '154.28.67.218',
                      '154.28.67.222', '154.28.67.223', '154.28.67.231', '154.28.67.240', '154.28.67.243',
                      '154.28.67.253',
                      '154.28.67.39', '154.28.67.4', '154.28.67.49', '154.28.67.5', '154.28.67.61', '154.28.67.80',
                      '154.28.67.81', '154.28.67.87', '154.28.67.88', '154.28.67.96', '154.28.67.99', '154.7.230.100',
                      '154.7.230.101', '154.7.230.103', '154.7.230.107', '154.7.230.109', '154.7.230.130',
                      '154.7.230.132',
                      '154.7.230.14', '154.7.230.140', '154.7.230.147', '154.7.230.151', '154.7.230.156',
                      '154.7.230.163',
                      '154.7.230.170', '154.7.230.18', '154.7.230.183', '154.7.230.188', '154.7.230.189',
                      '154.7.230.19',
                      '154.7.230.190', '154.7.230.198', '154.7.230.204', '154.7.230.209', '154.7.230.235',
                      '154.7.230.238',
                      '154.7.230.246', '154.7.230.29', '154.7.230.41', '154.7.230.42', '154.7.230.51', '154.7.230.55',
                      '154.7.230.60', '154.7.230.61', '154.7.230.74', '154.7.230.82', '154.7.230.89', '23.131.8.112',
                      '23.131.8.115', '23.131.8.117', '23.131.8.12', '23.131.8.121', '23.131.8.124', '23.131.8.150',
                      '23.131.8.161', '23.131.8.166', '23.131.8.171', '23.131.8.173', '23.131.8.176', '23.131.8.177',
                      '23.131.8.181', '23.131.8.19', '23.131.8.192', '23.131.8.194', '23.131.8.199', '23.131.8.202',
                      '23.131.8.203', '23.131.8.204', '23.131.8.207', '23.131.8.209', '23.131.8.213', '23.131.8.216',
                      '23.131.8.225', '23.131.8.228', '23.131.8.231', '23.131.8.238', '23.131.8.254', '23.131.8.36',
                      '23.131.8.5', '23.131.8.76', '23.131.8.93', '23.131.8.95', '23.131.8.99', '23.131.88.105',
                      '23.131.88.12', '23.131.88.137', '23.131.88.139', '23.131.88.140', '23.131.88.145',
                      '23.131.88.150',
                      '23.131.88.151', '23.131.88.153', '23.131.88.154', '23.131.88.156', '23.131.88.165',
                      '23.131.88.18',
                      '23.131.88.191', '23.131.88.192', '23.131.88.194', '23.131.88.198', '23.131.88.202',
                      '23.131.88.206',
                      '23.131.88.220', '23.131.88.223', '23.131.88.228', '23.131.88.233', '23.131.88.24',
                      '23.131.88.242',
                      '23.131.88.244', '23.131.88.47', '23.131.88.63', '23.131.88.67', '23.131.88.73', '23.131.88.80',
                      '23.131.88.81', '23.131.88.82', '23.131.88.88', '23.131.88.97', '23.170.144.149',
                      '23.170.144.209',
                      '23.170.144.212', '23.170.144.242', '23.170.144.83', '23.170.145.117', '23.170.145.167',
                      '23.170.145.182', '23.170.145.19', '23.170.145.203', '23.226.17.101', '23.226.17.109',
                      '23.226.17.112', '23.226.17.113', '23.226.17.115', '23.226.17.123', '23.226.17.129',
                      '23.226.17.143',
                      '23.226.17.148', '23.226.17.165', '23.226.17.186', '23.226.17.199', '23.226.17.201',
                      '23.226.17.207',
                      '23.226.17.210', '23.226.17.219', '23.226.17.220', '23.226.17.222', '23.226.17.229',
                      '23.226.17.250',
                      '23.226.17.254', '23.226.17.26', '23.226.17.33', '23.226.17.4', '23.226.17.49', '23.226.17.5',
                      '23.226.17.55', '23.226.17.66', '23.226.17.7', '23.226.17.72', '23.226.17.78', '23.226.17.8',
                      '23.226.17.86', '23.226.17.90', '23.226.17.93', '23.230.177.105', '23.230.177.110',
                      '23.230.177.113',
                      '23.230.177.121', '23.230.177.130', '23.230.177.14', '23.230.177.143', '23.230.177.15',
                      '23.230.177.150', '23.230.177.154', '23.230.177.165', '23.230.177.173', '23.230.177.191',
                      '23.230.177.196', '23.230.177.203', '23.230.177.206', '23.230.177.208', '23.230.177.217',
                      '23.230.177.220', '23.230.177.221', '23.230.177.224', '23.230.177.228', '23.230.177.231',
                      '23.230.177.235', '23.230.177.237', '23.230.177.241', '23.230.177.27', '23.230.177.38',
                      '23.230.177.52', '23.230.177.61', '23.230.177.67', '23.230.177.72', '23.230.177.80',
                      '23.230.177.88',
                      '23.230.177.94', '23.230.177.99', '23.230.197.103', '23.230.197.106', '23.230.197.109',
                      '23.230.197.11', '23.230.197.12', '23.230.197.122', '23.230.197.124', '23.230.197.146',
                      '23.230.197.155', '23.230.197.156', '23.230.197.174', '23.230.197.179', '23.230.197.181',
                      '23.230.197.196', '23.230.197.2', '23.230.197.201', '23.230.197.207', '23.230.197.208',
                      '23.230.197.225', '23.230.197.227', '23.230.197.233', '23.230.197.236', '23.230.197.239',
                      '23.230.197.240', '23.230.197.244', '23.230.197.251', '23.230.197.50', '23.230.197.52',
                      '23.230.197.54', '23.230.197.60', '23.230.197.71', '23.230.197.80', '23.230.197.81',
                      '23.230.197.84',
                      '23.230.197.97', '23.230.74.102', '23.230.74.110', '23.230.74.116', '23.230.74.125',
                      '23.230.74.133',
                      '23.230.74.135', '23.230.74.14', '23.230.74.141', '23.230.74.149', '23.230.74.15',
                      '23.230.74.157',
                      '23.230.74.16', '23.230.74.170', '23.230.74.172', '23.230.74.174', '23.230.74.183',
                      '23.230.74.187',
                      '23.230.74.19', '23.230.74.198', '23.230.74.208', '23.230.74.212', '23.230.74.215',
                      '23.230.74.23',
                      '23.230.74.230', '23.230.74.231', '23.230.74.252', '23.230.74.30', '23.230.74.41', '23.230.74.57',
                      '23.230.74.58', '23.230.74.59', '23.230.74.6', '23.230.74.75', '23.230.74.81', '23.230.74.88',
                      '23.230.74.91', '23.27.222.108', '23.27.222.109', '23.27.222.134', '23.27.222.138',
                      '23.27.222.159',
                      '23.27.222.161', '23.27.222.164', '23.27.222.166', '23.27.222.178', '23.27.222.19',
                      '23.27.222.195',
                      '23.27.222.201', '23.27.222.202', '23.27.222.203', '23.27.222.208', '23.27.222.21',
                      '23.27.222.211',
                      '23.27.222.218', '23.27.222.223', '23.27.222.228', '23.27.222.234', '23.27.222.236',
                      '23.27.222.242',
                      '23.27.222.251', '23.27.222.253', '23.27.222.34', '23.27.222.61', '23.27.222.62', '23.27.222.69',
                      '23.27.222.70', '23.27.222.72', '23.27.222.73', '23.27.222.74', '23.27.222.81', '23.27.222.93',
                      '38.131.131.110', '38.131.131.114', '38.131.131.123', '38.131.131.125', '38.131.131.137',
                      '38.131.131.142', '38.131.131.145', '38.131.131.147', '38.131.131.15', '38.131.131.154',
                      '38.131.131.16', '38.131.131.17', '38.131.131.173', '38.131.131.18', '38.131.131.193',
                      '38.131.131.204', '38.131.131.207', '38.131.131.227', '38.131.131.229', '38.131.131.233',
                      '38.131.131.238', '38.131.131.246', '38.131.131.248', '38.131.131.250', '38.131.131.31',
                      '38.131.131.36', '38.131.131.50', '38.131.131.58', '38.131.131.64', '38.131.131.70',
                      '38.131.131.71',
                      '38.131.131.74', '38.131.131.83', '38.131.131.94', '38.131.131.99', '38.75.75.104',
                      '38.75.75.111',
                      '38.75.75.112', '38.75.75.119', '38.75.75.120', '38.75.75.123', '38.75.75.127', '38.75.75.139',
                      '38.75.75.14', '38.75.75.143', '38.75.75.155', '38.75.75.156', '38.75.75.158', '38.75.75.170',
                      '38.75.75.179', '38.75.75.188', '38.75.75.2', '38.75.75.201', '38.75.75.231', '38.75.75.232',
                      '38.75.75.241', '38.75.75.246', '38.75.75.251', '38.75.75.26', '38.75.75.29', '38.75.75.4',
                      '38.75.75.44', '38.75.75.49', '38.75.75.56', '38.75.75.58', '38.75.75.62', '38.75.75.72',
                      '38.75.75.76', '38.75.75.79', '38.75.75.88', '38.96.156.108', '38.96.156.112', '38.96.156.128',
                      '38.96.156.131', '38.96.156.14', '38.96.156.142', '38.96.156.143', '38.96.156.149',
                      '38.96.156.16',
                      '38.96.156.163', '38.96.156.165', '38.96.156.169', '38.96.156.186', '38.96.156.188',
                      '38.96.156.190',
                      '38.96.156.192', '38.96.156.194', '38.96.156.199', '38.96.156.218', '38.96.156.236',
                      '38.96.156.240',
                      '38.96.156.252', '38.96.156.28', '38.96.156.32', '38.96.156.35', '38.96.156.56', '38.96.156.57',
                      '38.96.156.6', '38.96.156.67', '38.96.156.77', '38.96.156.80', '38.96.156.83', '38.96.156.84',
                      '38.96.156.89', '38.96.156.92', '45.238.157.100', '45.238.157.104', '45.238.157.106',
                      '45.238.157.110', '45.238.157.116', '45.238.157.118', '45.238.157.119', '45.238.157.12',
                      '45.238.157.123', '45.238.157.132', '45.238.157.14', '45.238.157.149', '45.238.157.15',
                      '45.238.157.183', '45.238.157.186', '45.238.157.189', '45.238.157.2', '45.238.157.212',
                      '45.238.157.214', '45.238.157.217', '45.238.157.22', '45.238.157.228', '45.238.157.23',
                      '45.238.157.247', '45.238.157.43', '45.238.157.48', '45.238.157.51', '45.238.157.52',
                      '45.238.157.53',
                      '45.238.157.56', '45.238.157.61', '45.238.157.65', '45.238.157.72', '45.238.157.79',
                      '45.238.157.8',
                      '45.238.159.103', '45.238.159.107', '45.238.159.110', '45.238.159.114', '45.238.159.116',
                      '45.238.159.123', '45.238.159.126', '45.238.159.144', '45.238.159.148', '45.238.159.15',
                      '45.238.159.156', '45.238.159.165', '45.238.159.167', '45.238.159.183', '45.238.159.20',
                      '45.238.159.208', '45.238.159.217', '45.238.159.220', '45.238.159.23', '45.238.159.230',
                      '45.238.159.235', '45.238.159.237', '45.238.159.238', '45.238.159.24', '45.238.159.249',
                      '45.238.159.251', '45.238.159.32', '45.238.159.34', '45.238.159.51', '45.238.159.6',
                      '45.238.159.66',
                      '45.238.159.77', '45.238.159.79', '45.238.159.82', '45.238.159.91']
def by_product_id(prod_id_data, n):
    product_id_url = str(prod_id_data[0])
    #product_id_url = 'https://www.zalando.fr/adidas-originals-stan-smith-baskets-basses-footwear-whitegold-metallic-ad111a1go-a11.html'
    print(product_id_url)
    import datetime
    time = datetime.datetime.now()
    details_dict = {'SKU_ID': prod_id_data[1],
                    'Website': prod_id_data[3],
                    'Country': prod_id_data[2],
                    'RPC': '',
                    'MPC': '',
                    'Product_ID': '',
                    'Product_URL': prod_id_data[0],
                    # 'Product_URL': prod_id_data[3],
                    'Product_Name': '',
                    'Category_Path': '',
                    'Specification': '',
                    'Description': '',
                    'Currency': 'WON',
                    'List_Price': '',
                    'Promo_Price': '',
                    'Discount': '',
                    'Brand': '',
                    'Rating_Count': '',
                    'Review_Count': '',
                    'Image_URLs': '',
                    'Variant': '',
                    'Variant_ID': '',
                    'Colour_of_Variant': '',
                    'Colour_Grouping': '',
                    'Seller_Name': '',
                    'Stock_Count': '',
                    'Stock_Condition': '',
                    'Stock_Message': '',
                    'Sustainability_Badge': '',
                    'Reason_Code': '',
                    'Crawling_TimeStamp': time.strftime("%Y-%m-%dT%H:%M:%SZ"),
                    'Cache_Page_Link': '',
                    'Extra1': '-',
                    'Extra2': '-',
                    'Extra3': '-',
                    'Extra4': '-',
                    'Extra5': '-'
                    }
    try:
        p_auth = str("csimonra:h19VA2xZ")
        p_host = random.choice(p1)
        p_port = "29842"
        proxy = {
            'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
            'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
        }
        response = requests.get(product_id_url, headers=headers, proxies=proxy, timeout=30)

        print(response.status_code)
        # print(response.text)

        n_hitting = 0
        if response.status_code == 200:
            tree = html.fromstring(response.text)

            Reason_code = "Success-PF"

            cache_data = response.text
            html_file_path = html_cache_page_saving(sku=str(prod_id_data[1]), response=cache_data)
            try:
                # product_id = tree.xpath('//div[@class="OXMl2N xVaPPo f2qidi _56Chwa"]//*[contains(text(),"Référence")]//following-sibling::span/text()')[0]
                product_id = tree.xpath(
                    '//div[@class="OXMl2N xVaPPo f2qidi _56Chwa"]//*[contains(text(),"Référence:")]//following-sibling::span/text()')

                details_dict['SKU_ID'] = prod_id_data[1]
                details_dict['RPC'] = str(prod_id_data[1]).replace('Zalando_fr_', '')
                details_dict['MPC'] = 'Not Available'
                details_dict['Product_ID'] = str(prod_id_data[1]).replace('Zalando_fr_', '')
                details_dict['Product_URL'] = prod_id_data[0]
                # name = tree.xpath('//span[@class="EKabf7 R_QwOV"]/text()')[0]
                name = tree.xpath('//h1/span/text()')[0]
                # print(name)
                details_dict['Product_Name'] = general.clean(name)

                details_dict['Category_Path'] = 'Not Available'

                value = tree.xpath(
                    '//div[@class="z-pdp__escape-grid"]/div[@class="y4Yt_f NN8L-8 JT3_zV MxUWj-"]/div/div/div/div/span[2]/text()')
                # print(len(value))
                if value ==[]:
                    spe_text = str(response.text).split('</script><script data-re-asset type="application/json" class="re-1-12">')[-1].split('</script><script data-re-asset class="re-1-12">')[0]
                    value = re.findall(r'\{"__typename":"ProductAttributeKeyValue","key":"(.*?)","value":"',spe_text)
                    value1 = re.findall(r'","value":"(.*?)"}',spe_text)
                    details_info = []
                    if value !=[]:
                        for i in range(len(value)):
                            details_info.append(value[i]+':'+value1[i])
                            detail_info_pipe = '|'.join(details_info)
                            # Specification = detail_info_pipe
                            details_dict['Specification'] = detail_info_pipe
                else:
                    details_info = []
                    for i in range(len(value)):
                        details_info.append(str(tree.xpath(
                            '//div[@class="z-pdp__escape-grid"]/div[@class="y4Yt_f NN8L-8 JT3_zV MxUWj-"]/div/div/div/div/span[1]/text()')[
                                                    (i - 1) + (i + 1)]) + ':' + str(tree.xpath(
                            '//div[@class="z-pdp__escape-grid"]/div[@class="y4Yt_f NN8L-8 JT3_zV MxUWj-"]/div/div/div/div/span[2]/text()')[
                                                                                        i]))
                        detail_info_pipe = '|'.join(details_info)
                        # Specification = detail_info_pipe
                        details_dict['Specification'] = detail_info_pipe#,"clusters":[]}],"b
                details_dict['Specification'] = general.clean(details_dict['Specification'])
                # sizes = re.findall(r'tive","description":(.*),"head', response.text)
                sizes = re.findall(r'{"__typename":"ProductAttributeDescriptive","description":"(.*?)","heading"', response.text)

                if len(sizes) == 1:
                    details_dict['Description'] = sizes[0]
                else:
                    details_dict['Description'] = 'Not Available'
                details_dict['Description'] = general.clean(details_dict['Description'])
                try:
                    #disc_price = tree.xpath('//div[@class="_0xLoFW vSgP6A _7ckuOK"]/span[@class="RYghuO uqkIZw ka2E9k uMhVZi dgII7d _88STHx cMfkVL"]/text()')
                    #value_MRP = tree.xpath('//div[@class="_0xLoFW vSgP6A"]/span[@class="RYghuO uqkIZw ka2E9k uMhVZi FxZV-M weHhRC ZiDB59"]/text()')[0].split("\xa0€")
                    print(product_id_url)
                    # discount = tree.xpath('//div[@class="Bqz_1C"]/span/text()')[0]
                    discount = general.xpath(tree, '//div[@class="Bqz_1C"]/p/text()')
                    discount = discount.replace('Jusqu’à', '')
                    discount = discount.replace('de réduction', '').strip()
                    print(discount)
                    if discount == '' or 'g' in discount or 'ml' in discount:
                        promo_price = general.xpath(tree,
                                                    '//div[@class="Bqz_1C"]//div[@class="_0xLoFW vSgP6A"]/p',mode='set_tc')
                        #à partir de 99,95 €
                        promo_price = promo_price.replace('€|TVA incluse', '').replace('à partir de', '').strip().strip()
                        list_price = promo_price
                        promo_price = promo_price
                        discount = '-'
                    else:
                        promo_price = general.xpath(tree,'//div[@class="Bqz_1C"]/div[@class="_0xLoFW vSgP6A _7ckuOK"]/p/text()')

                        list_price = general.xpath(tree, '//div[@class="Bqz_1C"]//div[@class="_0xLoFW vSgP6A"]/p',
                                               mode='set_tc')

                        promo_price = promo_price.replace('à partir de', '').replace('€','').strip()
                        list_price = list_price.replace('€|TVA incluse', '').replace('à partir de', '').strip()

                    details_dict['List_Price'] = list_price
                    details_dict['Promo_Price'] = promo_price
                    details_dict['Discount'] = discount
                    print((promo_price, list_price, discount))
                except:
                    value_MRP = tree.xpath('//div[@class="_0xLoFW vSgP6A"]/span[1]/text()')[0].split("\xa0€")
                    details_dict['List_Price'] = value_MRP[0].replace('à partir de', '')
                    details_dict['Promo_Price'] = value_MRP[0].replace('à partir de', '')
                    details_dict['Discount'] = '-'
                # print(details_dict)

                brand = tree.xpath('//span[@class="RYghuO _7Cm1F9 ka2E9k uMhVZi dgII7d pVrzNP"]/h3/text()')[0]
                # print(brand)
                details_dict['Brand'] = brand

                try:
                    details_dict['Rating_Count'] = \
                    tree.xpath('//div[@class="VKvyEj _0xLoFW _7ckuOK mROyo1"]/div/div/span[1]/text()')[0]
                    review = tree.xpath('//div[@class="DT5BTM _2hG8pA M1UKyg KoaFhL"]/h5/text()')[0]
                    details_dict['Review_Count'] = re.findall(r'\d+', review)[0]

                except:
                    details_dict['Rating_Count'] = '-'
                    details_dict['Review_Count'] = '-'


                img = tree.xpath('//div[@class="JT3_zV KLaowZ"]/ul/li/div/button/div/div/img/@src')
                img_pipe = '|'.join(img)
                details_dict['Image_URLs'] = img_pipe

                try:
                    seller = tree.xpath(
                        '//div[@class="oX3eAU _2LebSa Rft9Ae oW5DLR f2qidi _56Chwa OXMl2N"]/p/text()')
                    if len(seller) == 1:
                        seller_name = seller[0]
                        seller_name = seller_name.replace('Vendu et envoyé par', '')

                        details_dict['Seller_Name'] = seller_name
                        details_dict['Seller_Name'] = general.clean(details_dict['Seller_Name'])
                    else:
                        details_dict['Seller_Name'] = '-'
                except:
                    details_dict['Seller_Name'] = '-'

                try:
                    color_g = tree.xpath('//div[@class="JT3_zV mo6ZnF"]/a/div/div/img/@alt')
                    print(len(color_g))
                    print(color_g)
                    if color_g:
                        color_group = color_g[0:int(len(color_g) / 2)]
                        color_g_pipe = '|'.join(color_group)
                        details_dict['Colour_Grouping'] = general.clean(color_g_pipe)
                    else:
                        details_dict['Colour_Grouping'] = '-'
                except:
                    details_dict['Colour_Grouping'] = '-'

                #details_dict['Colour_of_Variant'] = tree.xpath('//div[@class="hPWzFB"]/span[2]/text()')[0]
                details_dict['Colour_of_Variant'] = general.xpath(tree, '//div[@class="okmnKS H_-43B"]//p[@class="RYghuO u-6V88 ka2E9k uMhVZi dgII7d pVrzNP zN9KaA"]/text()')
                print(details_dict['Colour_of_Variant'] )
                details_dict['Colour_of_Variant'] = general.clean(details_dict['Colour_of_Variant'])
                # details_dict['Stock_Count'] = 'Not Available'

                sustanability = tree.xpath(
                    '//div[@class="b3yJDY"]/div[@class="_0xLoFW FCIprz lqniou DJxzzA FCIprz sqymKT WCjo-q"]/span/text()')
                if (len(sustanability)) == 1:
                    details_dict['Sustainability_Badge'] = 'Yes'
                else:
                    details_dict['Sustainability_Badge'] = "No"

                details_dict['Reason_Code'] = 'Success-PF'
                # details_dict['Stock_Message'] = 'Not Available'
                details_dict['Cache_Page_Link'] = html_file_path


                sizes = re.findall(r'simples":(.*)}],"condition', response.text)
                # print(sizes)
                new_size = sizes[0] + "}]"
                # print(new_size)

                new_load_data = json.loads(new_size)

                for element in new_load_data:

                    ele = element['size']

                    details_dict['Variant'] = general.clean(ele)
                    details_dict['Variant_ID'] = element['sku']

                    stock = element['offer']['stock']['quantity']
                    # details_dict['Variant_ID'] = tree.xpath('//input[@id="size-picker-AD115O0DA-A110003000"]/text()')[0]
                    # print("stock: " ,stock)
                    if stock == "OUT_OF_STOCK" or (stock == "0"):
                        Stock_quantity = "Out of Stock"
                        # print("Stock_quantity has 0")
                        details_dict['Stock_Condition'] = Stock_quantity
                    else:
                        Stock_quantity = "In stock"
                        details_dict['Stock_Condition'] = Stock_quantity
                    if stock != "OUT_OF_STOCK":
                        Stock_quantity = element['offer']['stock']['quantity']
                        if Stock_quantity == "MANY":
                            details_dict['Stock_Count'] = '-'
                        else:
                            details_dict['Stock_Count'] = Stock_quantity
                    else:
                        details_dict['Stock_Count'] = '-'

                    details_dict['Stock_Message'] = 'Not Available'
                    print(details_dict)
                    store_data_by_product_id(dict=[details_dict])

            except:
                # pass
                details_dict['Reason_Code'] = 'Success-PNF'
                print(details_dict)
                store_data_by_product_id(dict=[details_dict])

        elif n != 0:
            n = n - 1
            by_product_id(prod_id_data, n)
        else:
            details_dict['Reason_Code'] = 'Success-PNF'
            print(details_dict)
            #store_data_by_product_id(dict=[details_dict])
    except Exception as e:
        print('proxy error')
        details_dict['Reason_Code'] = 'Success-PNF'
        print(details_dict)
        #store_data_by_product_id(dict=[details_dict])


for data in extract_xl():
    print('inside 1')
    by_product_id(prod_id_data=data, n=3)

end = time.time()
print(end)
total_time = end - start
print(total_time)